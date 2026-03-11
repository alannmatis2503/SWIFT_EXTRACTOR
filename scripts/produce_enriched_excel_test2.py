#!/usr/bin/env python3
"""
Script complet d'analyse des transferts — dossier Test 2
==========================================================
Produit un fichier Excel avec :
  - Feuille 1 : Transferts_Executes (MT900 matchés avec MT103/MT202)
  - Feuille 2 : MT900_non_rapproches (96 MT900 non matchés)
  - Feuille 3 : Suspens (MT103/MT202 sans confirmation MT900)
  - Feuille 4 : Analyse_MT900_non_matches (détail de l'investigation)
"""

import sys
import os
import re
from pathlib import Path
from datetime import datetime

HF_SPACES_DIR = Path(__file__).resolve().parent.parent / "hf_spaces"
sys.path.insert(0, str(HF_SPACES_DIR))
os.chdir(str(HF_SPACES_DIR))

import logging
logging.basicConfig(level=logging.INFO, format="%(levelname)s - %(name)s - %(message)s")

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from extractor_manager import (
    extract_mt900_only,
    extract_dispatch,
    match_mt900_with_transfers,
    create_transfer_analysis_workbook,
)
from extractors import mt_multi
from extractors.mt202 import extract_text_from_pdf

DATA_DIR = Path(__file__).resolve().parent.parent / "data" / "raw" / "Test 2"
OUTPUT_DIR = Path(__file__).resolve().parent.parent / "data" / "output_test2"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def main():
    print("=" * 80)
    print("PRODUCTION DU FICHIER EXCEL — Analyse Transferts Test 2")
    print("=" * 80)

    # ====== Collecter les références brutes de TOUS les messages ======
    print("\n[1/5] Collecte des références brutes (92 MT103 + 370 MT202)...")
    all_raw_refs = {}
    for pdf_name in ['MT103.pdf', 'MT202.pdf']:
        pdf_path = DATA_DIR / pdf_name
        text = extract_text_from_pdf(pdf_path)
        blocks = mt_multi._split_messages(text)
        for i, blk in enumerate(blocks, 1):
            ref = re.search(r'Transaction Reference:\s*(.+?)(?:\n|$)', blk)
            id_match = re.search(r'Identifier:\s*fin\.(\d{3})', blk)
            has_nak = bool(re.search(r'\bNAK\b', blk[:600]))
            ref_s = ref.group(1).strip() if ref else '?'
            mt_type = f"fin.{id_match.group(1)}" if id_match else '?'

            # Extraire aussi montant, devise et date
            amt = re.search(r'Amount:\s*([\d.,]+)', blk)
            ccy = re.search(r'Currency:\s*(\w+)', blk)
            vd = re.search(r'Value Date:\s*(.+?)(?:\n|$)', blk)
            sender = re.search(r'Sender:\s*(\S+)', blk)
            receiver = re.search(r'Receiver:\s*(\S+)', blk)

            all_raw_refs[ref_s.upper()] = {
                'type': mt_type,
                'nak': has_nak,
                'pdf': pdf_name,
                'msg_num': i,
                'amount': amt.group(1) if amt else None,
                'currency': ccy.group(1) if ccy else None,
                'value_date': vd.group(1).strip() if vd else None,
                'sender': sender.group(1) if sender else None,
                'receiver': receiver.group(1) if receiver else None,
            }
    print(f"   → {len(all_raw_refs)} références brutes collectées")

    # ====== Extraction des MT900 ======
    print("\n[2/5] Extraction des MT900...")
    mt900_rows, _ = extract_mt900_only(DATA_DIR / "MT900.pdf")
    print(f"   → {len(mt900_rows)} MT900 extraits")

    # ====== Extraction des MT103/MT202 (mode standard = direction outgoing) ======
    print("\n[3/5] Extraction des MT103/MT202 (direction=outgoing)...")
    all_transfers = []
    for pdf_name in ["MT103.pdf", "MT202.pdf"]:
        rows, *_ = extract_dispatch(DATA_DIR / pdf_name, direction="outgoing")
        transfer_rows = [r for r in rows if r.get("type_MT") and ("103" in r["type_MT"] or "202" in r["type_MT"])]
        all_transfers.extend(transfer_rows)
        print(f"   → {len(transfer_rows)} messages extraits depuis {pdf_name}")
    print(f"   → Total: {len(all_transfers)} MT103/MT202")

    # ====== Matching ======
    print("\n[4/5] Matching MT900 ↔ MT103/MT202...")
    matched_rows, suspens_rows, exception_rows, unmatched_rows = match_mt900_with_transfers(mt900_rows, all_transfers)
    print(f"   → {len(matched_rows)} matchés, {len(unmatched_rows)} non matchés, {len(exception_rows)} exceptions, {len(suspens_rows)} suspens")

    # ====== Produire le fichier Excel standard ======
    print("\n[5/5] Génération du fichier Excel...")

    # D'abord créer le workbook standard via la fonction de l'app
    out_path_standard = create_transfer_analysis_workbook(
        matched_rows, suspens_rows, exception_rows, OUTPUT_DIR,
        unmatched_mt900_rows=unmatched_rows,
    )
    print(f"   → Fichier standard créé : {out_path_standard.name}")

    # ====== Enrichir avec la feuille d'analyse ======
    from openpyxl import load_workbook

    wb = load_workbook(str(out_path_standard))

    # --- Style ---
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    orange_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # ====== Feuille: Analyse_Correspondances ======
    ws = wb.create_sheet(title="Analyse_Correspondances")

    analysis_headers = [
        "N°", "Réf MT900 (F20)", "Réf Origine (F21)", "Montant",
        "Devise", "Date Réf.", "Sender BIC",
        "Correspondance trouvée?", "Type", "Source PDF",
        "N° Message", "Statut NAK?", "Catégorie"
    ]
    ws.append(analysis_headers)

    # Appliquer le style header
    for col in range(1, len(analysis_headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Remplir les données
    for idx, r in enumerate(unmatched_rows, 1):
        rel_ref = str(r.get("related_reference") or "").strip()
        rel_ref_upper = rel_ref.upper()
        montant = r.get("montant")
        devise = r.get("devise") or "USD"
        date_ref = r.get("date_reference")
        sender_bic = r.get("sender_bic") or ""

        # Chercher la correspondance
        correspondance = "❌ Aucune"
        corr_type = ""
        corr_pdf = ""
        corr_msg = ""
        corr_nak = ""
        categorie = ""

        if not rel_ref:
            categorie = "Sans F21"
            correspondance = "⚠ Pas de F21"
        elif rel_ref_upper in all_raw_refs:
            info = all_raw_refs[rel_ref_upper]
            correspondance = "✅ Trouvé (exact)"
            corr_type = info["type"]
            corr_pdf = info["pdf"]
            corr_msg = str(info["msg_num"])
            corr_nak = "OUI" if info["nak"] else "NON"
            if info["nak"]:
                categorie = "MT103 NAK (message rejeté)"
            else:
                categorie = "Correspondance disponible"
        else:
            # Partiel?
            found_partial = False
            for mref, info in all_raw_refs.items():
                if mref.startswith(rel_ref_upper) or rel_ref_upper.startswith(mref):
                    correspondance = f"≈ Partiel ({mref})"
                    corr_type = info["type"]
                    corr_pdf = info["pdf"]
                    corr_msg = str(info["msg_num"])
                    corr_nak = "OUI" if info["nak"] else "NON"
                    categorie = "MT103 NAK (partiel)" if info["nak"] else "Correspondance partielle"
                    found_partial = True
                    break
            if not found_partial:
                # Classifier le type d'opération
                if re.match(r'^\d+\.\d+$', rel_ref):
                    categorie = "Opération NOSTRO/interbancaire"
                elif "NONREF" in rel_ref_upper:
                    categorie = "NONREF"
                else:
                    categorie = "Référence introuvable"

        row_data = [
            idx, r.get("reference"), rel_ref, montant,
            devise, date_ref, sender_bic,
            correspondance, corr_type, corr_pdf,
            corr_msg, corr_nak, categorie
        ]
        ws.append(row_data)

        # Colorier la ligne selon la catégorie
        current_row = ws.max_row
        for col in range(1, len(analysis_headers) + 1):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            if "Trouvé" in correspondance or "Partiel" in correspondance:
                cell.fill = green_fill
            elif "NOSTRO" in categorie or "NONREF" in categorie:
                cell.fill = orange_fill
            else:
                cell.fill = red_fill

    # Ajuster les largeurs
    for col in range(1, len(analysis_headers) + 1):
        vals = [cell.value for cell in ws[get_column_letter(col)] if cell.value is not None]
        max_len = max((len(str(v)) for v in vals), default=10)
        ws.column_dimensions[get_column_letter(col)].width = min(45, max(12, max_len + 2))

    # ====== Feuille: Synthèse ======
    ws_synth = wb.create_sheet(title="Synthèse", index=0)

    # Title
    ws_synth.merge_cells('A1:D1')
    title_cell = ws_synth.cell(row=1, column=1)
    title_cell.value = "SYNTHÈSE — Analyse Transferts Exécutés (Test 2)"
    title_cell.font = Font(bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center")

    ws_synth.merge_cells('A2:D2')
    ws_synth.cell(row=2, column=1).value = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    ws_synth.cell(row=2, column=1).font = Font(italic=True, color="808080")
    ws_synth.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    # Section 1: Données sources
    row = 4
    ws_synth.cell(row=row, column=1).value = "DONNÉES SOURCES"
    ws_synth.cell(row=row, column=1).font = Font(bold=True, size=12, color="4472C4")
    row += 1
    for label, val in [
        ("MT900.pdf", f"{len(mt900_rows)} messages"),
        ("MT103.pdf", f"92 messages (26 non-NAK, 66 NAK)"),
        ("MT202.pdf", f"370 messages (336 non-NAK, 34 NAK)"),
    ]:
        ws_synth.cell(row=row, column=1).value = label
        ws_synth.cell(row=row, column=2).value = val
        ws_synth.cell(row=row, column=1).font = Font(bold=True)
        row += 1

    # Section 2: Résultats du matching
    row += 1
    ws_synth.cell(row=row, column=1).value = "RÉSULTATS DU MATCHING MT900"
    ws_synth.cell(row=row, column=1).font = Font(bold=True, size=12, color="4472C4")
    row += 1
    for label, val, pct in [
        ("MT900 matchés avec MT103/MT202", len(matched_rows), f"{len(matched_rows)/len(mt900_rows)*100:.1f}%"),
        ("MT900 NON matchés", len(unmatched_rows), f"{len(unmatched_rows)/len(mt900_rows)*100:.1f}%"),
        ("MT900 en exception", len(exception_rows), f"{len(exception_rows)/len(mt900_rows)*100:.1f}%"),
        ("MT103/MT202 en suspens", len(suspens_rows), ""),
    ]:
        ws_synth.cell(row=row, column=1).value = label
        ws_synth.cell(row=row, column=2).value = val
        ws_synth.cell(row=row, column=3).value = pct
        ws_synth.cell(row=row, column=1).font = Font(bold=True)
        row += 1

    # Détail des matchés par type
    from collections import Counter
    matched_types = Counter(r.get("matched_type") for r in matched_rows)
    row += 1
    ws_synth.cell(row=row, column=1).value = "DÉTAIL DES MATCHÉS"
    ws_synth.cell(row=row, column=1).font = Font(bold=True, size=12, color="4472C4")
    row += 1
    for mt_type, count in matched_types.most_common():
        ws_synth.cell(row=row, column=1).value = f"MT900 matchés avec {mt_type}"
        ws_synth.cell(row=row, column=2).value = count
        row += 1

    # Section 3: Analyse des MT900 non matchés
    row += 1
    ws_synth.cell(row=row, column=1).value = "ANALYSE DES 96 MT900 NON MATCHÉS"
    ws_synth.cell(row=row, column=1).font = Font(bold=True, size=12, color="4472C4")
    row += 1

    # Compter les catégories
    cat_counts = Counter()
    for r in unmatched_rows:
        rel_ref = str(r.get("related_reference") or "").strip().upper()
        if not rel_ref:
            cat_counts["Sans F21"] += 1
        elif rel_ref in all_raw_refs:
            info = all_raw_refs[rel_ref]
            if info["nak"]:
                cat_counts["MT103 NAK (message rejeté SWIFT)"] += 1
            else:
                cat_counts["Correspondance disponible (non NAK)"] += 1
        else:
            found = False
            for mref in all_raw_refs:
                if mref.startswith(rel_ref) or rel_ref.startswith(mref):
                    info = all_raw_refs[mref]
                    if info["nak"]:
                        cat_counts["MT103 NAK (partiel)"] += 1
                    else:
                        cat_counts["Correspondance partielle"] += 1
                    found = True
                    break
            if not found:
                if re.match(r'^\d+\.\d+$', rel_ref):
                    cat_counts["Opération NOSTRO/interbancaire"] += 1
                elif "NONREF" in rel_ref:
                    cat_counts["NONREF"] += 1
                else:
                    cat_counts["Référence introuvable"] += 1

    for cat, count in cat_counts.most_common():
        ws_synth.cell(row=row, column=1).value = cat
        ws_synth.cell(row=row, column=2).value = count
        ws_synth.cell(row=row, column=3).value = f"{count/len(unmatched_rows)*100:.1f}%"
        ws_synth.cell(row=row, column=1).font = Font(bold=True)
        row += 1

    # Conclusion
    row += 1
    ws_synth.cell(row=row, column=1).value = "CONCLUSION"
    ws_synth.cell(row=row, column=1).font = Font(bold=True, size=12, color="4472C4")
    row += 1

    nak_count = sum(v for k, v in cat_counts.items() if "NAK" in k)
    nostro_count = cat_counts.get("Opération NOSTRO/interbancaire", 0)
    other_count = len(unmatched_rows) - nak_count - nostro_count - cat_counts.get("NONREF", 0)

    ws_synth.cell(row=row, column=1).value = (
        f"Sur les {len(unmatched_rows)} MT900 non matchés : "
        f"{nak_count} ({nak_count/len(unmatched_rows)*100:.0f}%) correspondent à des MT103 NAK, "
        f"{nostro_count} ({nostro_count/len(unmatched_rows)*100:.0f}%) sont des opérations NOSTRO/interbancaires, "
        f"et {other_count} sont d'autres cas."
    )
    ws_synth.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws_synth.cell(row=row, column=1).alignment = Alignment(wrap_text=True)

    # Ajuster largeurs de la synthèse
    ws_synth.column_dimensions['A'].width = 45
    ws_synth.column_dimensions['B'].width = 20
    ws_synth.column_dimensions['C'].width = 15
    ws_synth.column_dimensions['D'].width = 15

    # Sauvegarder
    enriched_path = OUTPUT_DIR / f"analyse_transferts_test2_enrichie_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(str(enriched_path))
    print(f"\n   ✅ Fichier Excel enrichi : {enriched_path}")
    print(f"   → {enriched_path.resolve()}")

    # Résumé console
    print(f"\n{'=' * 60}")
    print("RÉSUMÉ")
    print(f"{'=' * 60}")
    print(f"Feuilles du fichier Excel :")
    for name in wb.sheetnames:
        print(f"  📊 {name}")
    print(f"\nAnalyse des {len(unmatched_rows)} MT900 non matchés :")
    for cat, count in cat_counts.most_common():
        print(f"  • {cat}: {count} ({count/len(unmatched_rows)*100:.1f}%)")
    print(f"\n→ Conclusion : {nak_count}/{len(unmatched_rows)} MT900 non matchés")
    print(f"  ont bien leur correspondance parmi les MT103 (statut NAK)")

if __name__ == "__main__":
    main()
