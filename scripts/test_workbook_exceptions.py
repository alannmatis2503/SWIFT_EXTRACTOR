#!/usr/bin/env python3
"""
Test de création de workbook avec exceptions EUR.
"""
import sys
from pathlib import Path
import tempfile

sys.path.insert(0, str(Path(__file__).parent.parent / "hf_spaces"))

from extractor_manager import extract_dispatch, create_workbook

def test_workbook_with_exceptions():
    """Tester la création du workbook avec les exceptions EUR."""
    pdf_path = Path(__file__).parent.parent / "data/raw/report-484-msg254-255.pdf"
    
    print(f"📄 Test avec: {pdf_path.name}\n")
    
    # Extraction
    rows, beac_rows, exc_323201, other_exc, missing = extract_dispatch(
        pdf_path,
        direction="incoming"
    )
    
    print("=" * 70)
    print("RÉSULTATS EXTRACTION:")
    print("=" * 70)
    print(f"Messages normaux: {len(rows)}")
    print(f"BEACCMCX091: {len(beac_rows)}")
    print(f"Exceptions 323201: {len(exc_323201)}")
    print(f"Autres exceptions: {len(other_exc)}")
    print()
    
    # Créer le workbook
    with tempfile.TemporaryDirectory() as tmpdir:
        out_path = create_workbook(
            rows,
            Path(tmpdir),
            direction="incoming",
            beaccmcx091_rows=beac_rows,
            exception_323201_rows=exc_323201,
            other_exceptions_rows=other_exc
        )
        
        print("=" * 70)
        print("WORKBOOK CRÉÉ:")
        print("=" * 70)
        print(f"Fichier: {out_path}")
        
        # Lire les feuilles
        from openpyxl import load_workbook
        wb = load_workbook(out_path)
        
        print(f"\nFeuilles dans le workbook: {wb.sheetnames}")
        
        # Vérifier la feuille "Autres_Exceptions"
        if "Autres_Exceptions" in wb.sheetnames:
            sheet = wb["Autres_Exceptions"]
            print(f"\n✅ Feuille 'Autres_Exceptions' trouvée")
            print(f"   Nombre de lignes (avec en-tête): {sheet.max_row}")
            
            # Afficher les données
            if sheet.max_row > 1:
                print(f"\n   Contenu:")
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 1):
                    print(f"     Ligne {row_idx}: Type={row[0]}, Référence={row[2] if len(row) > 2 else 'N/A'}, Devise={row[4] if len(row) > 4 else 'N/A'}")
        else:
            print(f"\n❌ Feuille 'Autres_Exceptions' NON TROUVÉE")
            print(f"   Feuilles disponibles: {wb.sheetnames}")

if __name__ == "__main__":
    test_workbook_with_exceptions()
