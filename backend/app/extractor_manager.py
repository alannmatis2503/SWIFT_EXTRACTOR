# backend/app/extractor_manager.py
import os
import sys
from pathlib import Path
import re
import logging
from datetime import datetime
from typing import List, Dict, Optional

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

from utils import logger

# import extractors (primary helpers)
from backend.app.extractors.mt202 import extract_for_mt202, extract_text_from_pdf as extract_text_mt202
try:
    from backend.app.extractors.mt103 import extract_for_mt103
    HAS_MT103 = True
except Exception:
    HAS_MT103 = False
    logger.info("mt103 extractor not available at import time; you can add it to EXTRACTOR_MAP later.")

try:
    from backend.app.extractors.mt910 import extract_for_mt910
    HAS_MT910 = True
except Exception:
    HAS_MT910 = False
    logger.info("mt910 extractor not available at import time.")

# try to import the multi-message extractor (optional)
try:
    from backend.app.extractors import mt_multi as mt_multi_module
    HAS_MT_MULTI = True
except Exception:
    mt_multi_module = None
    HAS_MT_MULTI = False
    logger.info("mt_multi extractor not available at import time; multi-file detection will fall back to single extractors.")

# regex to detect MT/FIN code
MT_DETECT = re.compile(r'\b(?:MT|FIN)[\s\-\_\.:\/]*(\d{3})\b', re.I)
IDENTIFIER_FIN_RE = re.compile(r'Identifier[:\s]*fin[\.:\s\-\/]*(\d{3})', re.I)

# map MT number -> extractor callable
EXTRACTOR_MAP = {
    "202": extract_for_mt202,
}
if HAS_MT103:
    EXTRACTOR_MAP["103"] = extract_for_mt103
if HAS_MT910:
    EXTRACTOR_MAP["910"] = extract_for_mt910

# -----------------------------
# BIC mapping / donor logic
# -----------------------------
# caching globals
_cached_mapping: Optional[Dict[str, str]] = None
_cached_mapping_path: Optional[str] = None

# heuristics: possible default file locations
_DEFAULT_XLS_PATHS = [
    "data/bic_codes.xlsx",
    "data/bic.xlsx",
    "bic_codes.xlsx",
    "bic.xlsx",
    "data/bfde98b8-0a94-4ba1-ab8a-eae27357cc7e.xlsx"  # the uploaded name you used earlier (kept as candidate)
]

def _find_columns(df):
    """
    Find likely code_col and name_col heuristically from DataFrame columns.
    """
    cols = list(df.columns)
    code_col = None
    name_col = None
    for c in cols:
        cu = c.upper()
        if ('BIC' in cu) or (('CODE' in cu) and ('BIC' in cu or 'SWIFT' in cu)):
            code_col = c
            break
    if not code_col:
        for c in cols:
            cu = c.upper()
            if 'CODE' in cu:
                code_col = c
                break

    for c in cols:
        cu = c.upper()
        if 'NOM' in cu or 'NAME' in cu or 'NOMS' in cu:
            name_col = c
            break
    if not name_col:
        candidate = None
        best_alpha = 0.0
        for c in cols:
            sample = ' '.join([str(x) for x in df[c].dropna().astype(str).head(20).tolist()])
            if not sample:
                continue
            alpha_frac = sum(ch.isalpha() for ch in sample) / max(1, len(sample))
            if alpha_frac > best_alpha:
                best_alpha = alpha_frac
                candidate = c
        name_col = candidate
    return code_col, name_col


def bundled_base_path() -> Path:
    """
    Retourne le chemin racine d'où lire les fichiers "embarqués".
    - Si l'app est packagée par PyInstaller (--onefile), les resources sont extraites
      temporairement dans sys._MEIPASS.
    - Sinon, retourne la racine du projet (deux niveaux au-dessus de ce fichier).
    """
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        return Path(sys._MEIPASS)
    # adjust parents count depending on file location; this file is backend/app/extractor_manager.py
    # parents[2] points to repo root (pdf-extractor/)
    return Path(__file__).resolve().parents[2]

def _user_override_bic_paths() -> List[Path]:
    """
    Emplacements où un admin/utilisateur peut déposer un bic_codes.xlsx modifiable.
    Ordre de priorité (testé dans load_bic_mapping) :
      1) variable d'environnement PDF_SWIFT_DATA_DIR si définie
      2) dossier commun ProgramData (Windows) -> {PROGRAMDATA}/PDF_Swift_Extractor/data
      3) dossier local de l'utilisateur -> %LOCALAPPDATA%/PDF_Swift_Extractor/data ou ~/ .pdf_swift_extractor/data
    """
    paths = []
    env = os.getenv("PDF_SWIFT_DATA_DIR")
    if env:
        paths.append(Path(env))

    # Windows common appdata (ProgramData)
    programdata = os.getenv("PROGRAMDATA")
    if programdata:
        paths.append(Path(programdata) / "PDF_Swift_Extractor" / "data")

    # Windows local appdata or cross-platform user dir
    localappdata = os.getenv("LOCALAPPDATA") or os.getenv("XDG_DATA_HOME")
    if localappdata:
        paths.append(Path(localappdata) / "PDF_Swift_Extractor" / "data")

    # fallback to user home hidden dir
    paths.append(Path.home() / ".pdf_swift_extractor" / "data")

    return paths

def load_bic_mapping(xlsx_path: Optional[str] = None, sheet_name: Optional[str] = 0) -> Dict[str, str]:
    """
    Charge (et met en cache) la table BIC -> nom de banque depuis un fichier Excel.
    Logique améliorée pour permettre une mise à jour manuelle après installation :
      - si xlsx_path explicit fourni -> utilisé (existing behaviour)
      - sinon : on cherche d'abord dans des emplacements externes éditables (ProgramData, user dir,
        variable d'env PDF_SWIFT_DATA_DIR)
      - sinon : on cherche dans le bundle embarqué (bundled_base_path()/data)
      - sinon : on retombe sur les chemins _DEFAULT_XLS_PATHS comme avant
    """
    global _cached_mapping, _cached_mapping_path
    import pandas as pd  # lazy import

    # 1) if explicit path provided, prefer it
    if xlsx_path:
        p = Path(xlsx_path)
        if not p.exists():
            raise FileNotFoundError(f"Provided xlsx_path not found: {xlsx_path}")
    else:
        # 2) check user-writable override locations (ProgramData, LOCALAPPDATA, env var)
        p = None
        for base in _user_override_bic_paths():
            candidate = base / "bic_codes.xlsx"
            if candidate.exists():
                p = candidate
                break
            # also accept alternative names
            alt = base / "bic.xlsx"
            if alt.exists():
                p = alt
                break

        # 3) check bundled data dir (this will work for --onedir and for files added with --add-data)
        if p is None:
            bundled = bundled_base_path() / "data"
            for name in ("bic_codes.xlsx", "bic.xlsx"):
                cand = bundled / name
                if cand.exists():
                    p = cand
                    break

        # 4) fallback to original candidate list (relative to current working dir)
        if p is None:
            for cand in _DEFAULT_XLS_PATHS:
                if Path(cand).exists():
                    p = Path(cand)
                    break

        if p is None:
            raise FileNotFoundError(
                "Aucun fichier Excel trouvé. Place your Excel mapping in one of: "
                + ", ".join(_DEFAULT_XLS_PATHS)
                + " or a writable location like %PROGRAMDATA%\\PDF_Swift_Extractor\\data\\bic_codes.xlsx "
                + "or set environment variable PDF_SWIFT_DATA_DIR to a folder containing bic_codes.xlsx"
            )

    # cache check
    pstr = str(Path(p).resolve())
    if _cached_mapping is not None and _cached_mapping_path == pstr:
        return _cached_mapping

    df = pd.read_excel(pstr, sheet_name=sheet_name, dtype=str)
    df = df.dropna(axis=1, how='all')

    code_col, name_col = _find_columns(df)
    if not code_col:
        raise ValueError(f"Impossible de détecter la colonne code BIC dans {pstr}. Colonnes: {list(df.columns)}")
    if not name_col:
        logger.warning("load_bic_mapping: impossible de détecter colonne 'nom' ; les valeurs de nom seront vides.")

    mapping: Dict[str, str] = {}
    for _, row in df.iterrows():
        code_val = (str(row.get(code_col) or "")).strip()
        name_val = (str(row.get(name_col) or "")).strip() if name_col else ""
        if not code_val or code_val.lower() in ("nan", "none"):
            continue
        code_clean = re.sub(r'\s+', '', code_val).upper()
        key = code_clean[:8]
        if not key:
            continue
        mapping[key] = name_val

    _cached_mapping = mapping
    _cached_mapping_path = pstr
    logger.info("load_bic_mapping: loaded %d entries from %s", len(mapping), pstr)
    return mapping


FALLBACK_11_RE = re.compile(r'\b([A-Z0-9]{11})\b')

# remplace la fonction get_donneur_from_f52 existante par ceci
IDENTIFIER_RE_AFTER_LABEL = re.compile(
    r"(?i)(?:IdentifierCode|Identifier Code|Identifiercode|Code d'identifiant|Code d identifiant|IDENTIFIERCODE)\s*[:\-\s]*\n?\s*([A-Z0-9]{11})"
)

# words that look like labels and should NOT be accepted as code
_BAD_LABEL_TOKENS = {
    "IDENTIFIER", "IDENTIFIERC", "PARTYIDENTI", "PARTYIDENT", "IDENTIFIANT", "IDENTIFIERCODE",
    "PARTY", "PARTYIDENTIFIER"
}


# label regex (variantes FR/EN)
_LABEL_RE = re.compile(
    r"(?i)(?:IdentifierCode|Identifier Code|Identifiercode|Code d'identifiant|Code d identifiant|identifiant de partie|IDENTIFIERCODE)\s*[:\-\s]*",
    re.M
)

_BAD_LABEL_PREFIXES = ("IDENTIF", "PARTYIDENT", "PARTY", "IDENTIFIANT")

def _find_identifier_after_label(text: str, lookahead_chars: int = 600) -> Optional[str]:
    """
    Cherche après un label 'IdentifierCode' un token alphanumérique 6..11 caractères,
    autorise lignes vides entre label et token, privilégie tokens contenant des lettres.
    """
    if not text:
        return None
    txt = text.replace('\r', '\n')
    m = _LABEL_RE.search(txt)
    if not m:
        return None
    start = m.end()
    tail = txt[start: start + lookahead_chars]
    # find candidates 6..11 chars
    toks = re.findall(r'\b([A-Z0-9]{6,11})\b', tail, flags=re.I)
    toks = [t.upper() for t in toks]
    # filter label-like tokens
    toks = [t for t in toks if not any(t.startswith(pref) for pref in _BAD_LABEL_PREFIXES)]
    if not toks:
        return None
    # prefer token with a letter (likely BIC), otherwise return first
    for t in toks:
        if re.search(r'[A-Z]', t):
            return t
    return toks[0]

def get_donneur_from_f52(f52_text: Optional[str], message_text: Optional[str] = None, xlsx_path: Optional[str] = None) -> Optional[str]:
    """
    Robust extract:
     - try inside F52A block: find label then next token (6..11 chars), preferring alpha tokens
     - if not found, try search on the whole message (cross-page)
     - if still not found, attempt a last-resort token in F52A that contains letters
     - then map first 8 chars to bank name using load_bic_mapping (if available)
    Returns: "CODE11/BANK NAME" if mapping found, else CODE (or None).
    """
    # normalize and remove xml-like tags
    def _norm(s):
        return re.sub(r'<[^>]+>', ' ', (s or "")).replace('\r', '\n')

    f52 = _norm(f52_text)
    full = _norm(message_text) if message_text else None

    # 1) try strictly in F52A block
    code = _find_identifier_after_label(f52, lookahead_chars=800)
    # 2) if not found, try whole message (cross-page)
    if not code and full:
        code = _find_identifier_after_label(full, lookahead_chars=1200)
    # 3) last-resort: try to find first alnum token with letters in F52A
    if not code and f52:
        m = re.search(r'\b([A-Z][A-Z0-9]{5,10})\b', f52, flags=re.I)
        if m:
            tok = m.group(1).upper()
            if not any(tok.startswith(pref) for pref in _BAD_LABEL_PREFIXES):
                code = tok

    if not code:
        return None

    # attempt mapping to bank name (uses cached loader)
    try:
        mapping = load_bic_mapping(xlsx_path=xlsx_path)
    except Exception:
        mapping = {}

    key8 = code[:8].upper()
    bank = mapping.get(key8)
    if bank:
        bank_clean = re.sub(r'\s{2,}', ' ', bank).strip()
        return f"{code}/{bank_clean}"
    return code


# -----------------------------
# Dispatcher / workbook logic (existing)
# -----------------------------
def detect_message_type(text: str) -> Optional[str]:
    """
    Detect the MT type (e.g. "202", "103", "910") from extracted text.
    Returns the numeric string (e.g. "202") or None.
    """
    if not text:
        return None

    m = MT_DETECT.search(text)
    if m:
        mt = m.group(1)
        logger.debug("detect_message_type: primary MT_DETECT matched -> %s", mt)
        return mt

    m2 = IDENTIFIER_FIN_RE.search(text)
    if m2:
        mt = m2.group(1)
        logger.debug("detect_message_type: IDENTIFIER_FIN_RE matched -> %s", mt)
        return mt

    logger.debug("detect_message_type: no MT type matched")
    return None


def extract_dispatch(pdf_path: Path, direction: str = "incoming") -> tuple[List[Dict], List[Dict], List[Dict], List[Dict], Dict[str, set]]:
    """
    Dispatcher intelligent :
      - si le PDF contient plusieurs messages -> utilise mt_multi.extract_messages_from_pdf
      - sinon -> utilise extract_single (retourne [row])
    Retourne toujours : (liste de rows, liste de BEACCMCX091 rows, liste de 323201 exception rows, 
                         liste d'autres exceptions (EUR/nivellement), dict de codes manquants)
    
    Args:
        pdf_path: Path to the PDF file
        direction: "incoming" or "outgoing" - determines beneficiary extraction logic
    """
    p = Path(pdf_path)
    # quick text extraction using existing helper
    text = ""
    try:
        text = extract_text_mt202(p)
    except Exception as e:
        logger.debug("extract_dispatch: extract_text_mt202 failed (%s), falling back to pdfplumber", e)
        try:
            import pdfplumber
            s = ""
            with pdfplumber.open(str(p)) as pdf:
                for page in pdf.pages[:2]:
                    s += "\n" + (page.extract_text() or "")
            text = s
        except Exception as e2:
            logger.warning("extract_dispatch: quick pdfplumber fallback failed for %s: %s", p.name, e2)
            text = ""

    missing_codes: Dict[str, set] = {"unmapped": set(), "empty": set()}
    beaccmcx091_rows: List[Dict] = []  # Liste séparée pour les messages BEACCMCX091
    exception_323201_rows: List[Dict] = []  # Liste pour les exceptions 323201
    other_exceptions_rows: List[Dict] = []  # Liste pour les autres exceptions (EUR/nivellement)

    # If multi-message extractor available, use its split logic to decide
    if HAS_MT_MULTI and mt_multi_module:
        try:
            blocks = mt_multi_module._split_messages(text)
            if blocks and len(blocks) > 1:
                logger.info("%s: detected %d messages (using mt_multi).", p.name, len(blocks))
                # Pass preloaded_text to avoid re-reading the PDF
                rows, beaccmcx091_rows, exception_323201_rows, other_exceptions_rows, missing_codes = mt_multi_module.extract_messages_from_pdf(p, direction=direction, preloaded_text=text)
                # ensure backward compatibility: set institution_name from donneur_dordre if missing
                for r in rows:
                    if "institution_name" not in r or not r.get("institution_name"):
                        r["institution_name"] = r.get("donneur_dordre") or r.get("donneur d'ordre") or None
                    for k in ["code_banque", "date_reference", "reference", "type_MT", "pays_iso3", "beneficiaire", "montant", "devise", "source_pdf", "commentaires", "correspondant"]:
                        if k not in r:
                            r[k] = None
                # apply same compatibility to BEACCMCX091 rows
                for r in beaccmcx091_rows:
                    if "institution_name" not in r or not r.get("institution_name"):
                        r["institution_name"] = r.get("donneur_dordre") or r.get("donneur d'ordre") or None
                    for k in ["code_banque", "date_reference", "reference", "type_MT", "pays_iso3", "beneficiaire", "montant", "devise", "source_pdf", "commentaires", "correspondant"]:
                        if k not in r:
                            r[k] = None
                # apply same compatibility to exception_323201 rows
                for r in exception_323201_rows:
                    if "institution_name" not in r or not r.get("institution_name"):
                        r["institution_name"] = r.get("donneur_dordre") or r.get("donneur d'ordre") or None
                    for k in ["code_banque", "date_reference", "reference", "type_MT", "pays_iso3", "beneficiaire", "montant", "devise", "source_pdf", "commentaires", "correspondant"]:
                        if k not in r:
                            r[k] = None
                # apply same compatibility to other_exceptions rows
                for r in other_exceptions_rows:
                    if "institution_name" not in r or not r.get("institution_name"):
                        r["institution_name"] = r.get("donneur_dordre") or r.get("donneur d'ordre") or None
                    for k in ["code_banque", "date_reference", "reference", "type_MT", "pays_iso3", "beneficiaire", "montant", "devise", "source_pdf", "commentaires", "correspondant"]:
                        if k not in r:
                            r[k] = None
                return rows, beaccmcx091_rows, exception_323201_rows, other_exceptions_rows, missing_codes
        except Exception as e:
            logger.exception("extract_dispatch: mt_multi detection/extraction failed for %s: %s", p.name, e)
            # fall through to single extractor

    # fallback: treat as single message
    single_row = extract_single(p, direction=direction)
    return [single_row], beaccmcx091_rows, exception_323201_rows, other_exceptions_rows, missing_codes


def _ensure_minimal_row(p: Path, mt_type: Optional[str] = None) -> Dict:
    """Return a minimal row template used when extraction not performed or failed."""
    return {
        "code_banque": None,
        "date_reference": None,
        "reference": None,
        "type_MT": f"fin.{mt_type}" if mt_type else None,
        "pays_iso3": None,
        "institution_name": None,
        "beneficiaire": None,
        "montant": None,
        "devise": None,
        "source_pdf": p.name
    }


def extract_single(pdf_path: Path, direction: str = "incoming") -> Dict:
    """
    Dispatch extraction for a single pdf_path (Path or str).
    Returns a dict with fields (internal keys). The create_workbook function maps
    'institution_name' -> "donneur d'ordre" when writing the summary sheet.
    
    Args:
        pdf_path: Path to the PDF file
        direction: "incoming" or "outgoing" - determines beneficiary extraction logic
    """
    p = Path(pdf_path)
    if not p.exists():
        logger.error("extract_single: file not found: %s", p)
        return _ensure_minimal_row(p)

    # read text (use helper from mt202 for consistent behavior)
    try:
        text = extract_text_mt202(p)
    except Exception as e:
        logger.exception("extract_single: extract_text_mt202 failed for %s: %s", p.name, e)
        # fallback quick text extraction
        try:
            import pdfplumber
            s = ""
            with pdfplumber.open(str(p)) as pdf:
                for page in pdf.pages[:2]:
                    s += "\n" + (page.extract_text() or "")
            text = s
        except Exception as e2:
            logger.exception("extract_single: fallback pdfplumber failed for %s: %s", p.name, e2)
            return _ensure_minimal_row(p)

    mt = detect_message_type(text)
    if not mt:
        logger.info("%s: MT type not found in text", p.name)
        row = _ensure_minimal_row(p, mt_type=None)
        row["source_pdf"] = p.name
        return row

    extractor = EXTRACTOR_MAP.get(mt)
    if not extractor:
        logger.info("%s: type detected -> %s but no extractor implemented", p.name, mt)
        row = _ensure_minimal_row(p, mt_type=mt)
        row["source_pdf"] = p.name
        return row

    try:
        # Pass direction to extractor if it supports it
        import inspect
        sig = inspect.signature(extractor)
        if 'direction' in sig.parameters:
            row = extractor(p, direction=direction)
        else:
            row = extractor(p)
            
        if not isinstance(row, dict):
            logger.error("%s: extractor returned non-dict result: %r", p.name, row)
            row = _ensure_minimal_row(p, mt_type=mt)
        else:
            required = ["code_banque", "date_reference", "reference", "type_MT", "pays_iso3",
                        "institution_name", "beneficiaire", "montant", "devise", "source_pdf"]
            for k in required:
                if k not in row:
                    row[k] = None
            if not row.get("institution_name") and row.get("donneur_dordre"):
                row["institution_name"] = row.get("donneur_dordre")
            if not row.get("type_MT"):
                row["type_MT"] = f"fin.{mt}"
            if not row.get("source_pdf"):
                row["source_pdf"] = p.name
        logger.info("%s: extracted via MT%s (direction: %s)", p.name, mt, direction)
        return row
    except Exception as e:
        logger.exception("Extraction failed for %s (MT%s): %s", p.name, mt, e)
        row = _ensure_minimal_row(p, mt_type=mt)
        row["error"] = str(e)
        return row


def _sanitize_sheet_title(name: str, max_len: int = 31) -> str:
    """Make a safe Excel sheet name (no invalid chars, limited length)."""
    if not name:
        name = "sheet"
    sanitized = re.sub(r'[:\\\/\?\*\[\]]+', '_', name)
    sanitized = sanitized.strip()
    if len(sanitized) > max_len:
        sanitized = sanitized[:max_len]
    if not sanitized:
        sanitized = "sheet"
    return sanitized


def create_workbook(rows: List[Dict], out_dir: Path, direction: str = "incoming", beaccmcx091_rows: Optional[List[Dict]] = None, exception_323201_rows: Optional[List[Dict]] = None, other_exceptions_rows: Optional[List[Dict]] = None, date_start: str = None, date_end: str = None) -> Path:
    """
    Create an Excel workbook with:
      - a 'summary' sheet containing one row per extracted file (display headers in French)
      - a 'BEACCMCX091' sheet if beaccmcx091_rows is provided
      - a 'Exceptions_323201' sheet if exception_323201_rows is provided
      - a 'Autres_Exceptions' sheet for EUR/nivellement exceptions
      - a 'Doublons_potentiels' sheet for potential duplicates (910 vs 202 with same ref+amount)
      - per-country summary sheets
      - one additional sheet per file with key/value pairs (debug-friendly)
    Returns the Path to the saved workbook.
    
    Args:
        rows: List of extracted data dictionaries
        out_dir: Output directory for the workbook
        direction: "incoming" or "outgoing" - used in filename
        beaccmcx091_rows: Optional list of BEACCMCX091 messages to display separately
        exception_323201_rows: Optional list of 323201 exception messages
        other_exceptions_rows: Optional list of other exceptions (EUR/nivellement)
        date_start: Optional start date for filtering (YYYY-MM-DD)
        date_end: Optional end date for filtering (YYYY-MM-DD)
    """
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Ajouter la plage de dates au nom du fichier si spécifiée
    date_suffix = ""
    if date_start and date_end:
        date_suffix = f"_{date_start}_to_{date_end}"
    elif date_start:
        date_suffix = f"_from_{date_start}"
    elif date_end:
        date_suffix = f"_to_{date_end}"
    
    out_path = out_dir / f"swift_extraction_{direction}{date_suffix}_{ts}.xlsx"

    wb = Workbook()
    summary = wb.active
    summary.title = "summary"

    # summary headers (user-facing) - avec nouvelles colonnes
    display_headers = [
        "code_banque",
        "date_reference",
        "reference",
        "type_MT",
        "pays_iso3",
        "Code du donneur d'ordre",
        "donneur d'ordre",
        "Bénéficiaire",
        "correspondant",
        "montant",
        "devise",
        "commentaires",
        "source_pdf"
    ]
    summary.append(display_headers)

    def _convert_date_to_excel(date_str):
        """Convertir une date ISO en objet datetime pour Excel."""
        if not date_str:
            return None
        try:
            # Format ISO: YYYY-MM-DD
            if isinstance(date_str, str) and re.match(r'^\d{4}-\d{2}-\d{2}$', date_str):
                return datetime.strptime(date_str, '%Y-%m-%d')
            return date_str
        except Exception:
            return date_str

    def _write_row_to_sheet(sheet, r, row_num=None):
        """Écrire une ligne de données dans une feuille."""
        code_donneur = r.get("code_donneur_dordre") or None
        donneur = r.get("donneur_dordre") or None
        if not donneur and "institution_name" in r:
            donneur = r.get("institution_name")
        beneficiaire = r.get("beneficiaire") or None
        correspondant = r.get("correspondant") or None
        commentaires = r.get("commentaires") or None
        date_ref = _convert_date_to_excel(r.get("date_reference"))
        
        row_data = [
            r.get("code_banque"),
            date_ref,
            r.get("reference"),
            r.get("type_MT"),
            r.get("pays_iso3"),
            code_donneur,
            donneur,
            beneficiaire,
            correspondant,
            r.get("montant"),
            r.get("devise"),
            commentaires,
            r.get("source_pdf")
        ]
        sheet.append(row_data)
        
        # Appliquer le format date à la colonne date_reference (colonne B = 2)
        if date_ref and isinstance(date_ref, datetime):
            current_row = sheet.max_row
            sheet.cell(row=current_row, column=2).number_format = 'DD/MM/YYYY'

    # NOUVELLE FONCTIONNALITÉ: Détection des doublons potentiels (910 vs 202)
    # ÉTAPE 1: Détecter les doublons AVANT d'écrire dans summary
    # Combiner toutes les rows pour la détection
    all_rows_for_duplicates = list(rows) + (beaccmcx091_rows or []) + (exception_323201_rows or []) + (other_exceptions_rows or [])
    
    # Créer des groupes par (référence, montant) pour détecter les doublons
    potential_duplicates = []
    seen_keys = {}
    rows_to_exclude_from_summary = set()  # IDs des messages à exclure de summary
    rows_to_mark_as_duplicate = set()  # IDs des messages à marquer "potentiel doublon"
    
    # Utiliser id() pour identifier de manière unique chaque dictionnaire
    for r in all_rows_for_duplicates:
        mt_type = r.get("type_MT") or ""
        montant = r.get("montant")
        
        # Pour les MT910, utiliser F20 comme référence (déjà fait dans l'extraction)
        reference = r.get("reference")
        
        if reference and montant is not None:
            # Normaliser la clé
            key = (str(reference).strip().upper(), float(montant) if montant else 0)
            
            if key in seen_keys:
                # Doublon potentiel trouvé
                prev_row = seen_keys[key]
                prev_type = prev_row.get("type_MT") or ""
                
                # Un doublon est si c'est un 910 vs 202 (ou l'inverse)
                is_910_vs_202 = (
                    ("910" in mt_type and "202" in prev_type) or
                    ("202" in mt_type and "910" in prev_type)
                )
                
                if is_910_vs_202:
                    # Ajouter les deux dans la liste des doublons potentiels
                    if prev_row not in potential_duplicates:
                        potential_duplicates.append(prev_row)
                    if r not in potential_duplicates:
                        potential_duplicates.append(r)
                    
                    # NOUVELLE LOGIQUE: Exclure le second du summary, marquer le premier
                    # On garde le premier message vu (prev_row) et on exclut le second (r)
                    rows_to_exclude_from_summary.add(id(r))
                    rows_to_mark_as_duplicate.add(id(prev_row))
            else:
                seen_keys[key] = r
    
    # ÉTAPE 2: Modifier les commentaires des messages à marquer
    for r in all_rows_for_duplicates:
        if id(r) in rows_to_mark_as_duplicate:
            current_comment = r.get("commentaires") or ""
            if current_comment:
                r["commentaires"] = f"{current_comment} / potentiel doublon"
            else:
                r["commentaires"] = "potentiel doublon"
    
    # ÉTAPE 3: Écrire les rows dans summary (en excluant les doublons)
    for r in rows:
        if id(r) not in rows_to_exclude_from_summary:
            _write_row_to_sheet(summary, r)

    # Add BEACCMCX091 sheet if there are any (right after main summary, before country summaries)
    if beaccmcx091_rows and len(beaccmcx091_rows) > 0:
        beac_sheet = wb.create_sheet(title="BEACCMCX091", index=1)
        beac_sheet.append(display_headers)
        
        for r in beaccmcx091_rows:
            if id(r) not in rows_to_exclude_from_summary:
                _write_row_to_sheet(beac_sheet, r)
        
        # Adjust column widths for BEACCMCX091 sheet
        try:
            for col_idx in range(1, len(display_headers) + 1):
                max_len = max(
                    (len(str(cell.value)) for cell in beac_sheet[get_column_letter(col_idx)] if cell.value is not None),
                    default=10
                )
                beac_sheet.column_dimensions[get_column_letter(col_idx)].width = min(60, max(12, max_len + 2))
        except Exception:
            pass

    # Add Exceptions_323201 sheet if there are any
    if exception_323201_rows and len(exception_323201_rows) > 0:
        exc_sheet = wb.create_sheet(title="Exceptions_323201", index=2 if beaccmcx091_rows else 1)
        exc_sheet.append(display_headers)
        
        for r in exception_323201_rows:
            if id(r) not in rows_to_exclude_from_summary:
                _write_row_to_sheet(exc_sheet, r)
        
        # Adjust column widths
        try:
            for col_idx in range(1, len(display_headers) + 1):
                max_len = max(
                    (len(str(cell.value)) for cell in exc_sheet[get_column_letter(col_idx)] if cell.value is not None),
                    default=10
                )
                exc_sheet.column_dimensions[get_column_letter(col_idx)].width = min(60, max(12, max_len + 2))
        except Exception:
            pass

    # Add Autres_Exceptions sheet if there are any (EUR/nivellement exceptions)
    sheet_index = 1
    if beaccmcx091_rows:
        sheet_index += 1
    if exception_323201_rows:
        sheet_index += 1
    
    if other_exceptions_rows and len(other_exceptions_rows) > 0:
        other_exc_sheet = wb.create_sheet(title="Autres_Exceptions", index=sheet_index)
        other_exc_sheet.append(display_headers)
        
        for r in other_exceptions_rows:
            if id(r) not in rows_to_exclude_from_summary:
                _write_row_to_sheet(other_exc_sheet, r)
        
        # Adjust column widths
        try:
            for col_idx in range(1, len(display_headers) + 1):
                max_len = max(
                    (len(str(cell.value)) for cell in other_exc_sheet[get_column_letter(col_idx)] if cell.value is not None),
                    default=10
                )
                other_exc_sheet.column_dimensions[get_column_letter(col_idx)].width = min(60, max(12, max_len + 2))
        except Exception:
            pass
    
    # ÉTAPE 4: Créer la feuille Doublons_potentiels (contient TOUS les doublons détectés)
    if potential_duplicates:
        sheet_index = 1
        if beaccmcx091_rows:
            sheet_index += 1
        if exception_323201_rows:
            sheet_index += 1
        if other_exceptions_rows:
            sheet_index += 1
        
        dup_sheet = wb.create_sheet(title="Doublons_potentiels", index=sheet_index)
        dup_sheet.append(display_headers)
        
        # Écrire TOUS les doublons (pas de filtrage ici)
        for r in potential_duplicates:
            _write_row_to_sheet(dup_sheet, r)
        
        # Adjust column widths
        try:
            for col_idx in range(1, len(display_headers) + 1):
                max_len = max(
                    (len(str(cell.value)) for cell in dup_sheet[get_column_letter(col_idx)] if cell.value is not None),
                    default=10
                )
                dup_sheet.column_dimensions[get_column_letter(col_idx)].width = min(60, max(12, max_len + 2))
        except Exception:
            pass

    # Add per-country summary sheets (after special sheets, before per-file sheets)
    countries = {}
    for r in rows:
        # Exclure les doublons aussi des feuilles par pays
        if id(r) not in rows_to_exclude_from_summary:
            country = r.get("pays_iso3")
            if country:
                if country not in countries:
                    countries[country] = []
                countries[country].append(r)
    
    # Create a sheet for each country with summary data
    for country_code in sorted(countries.keys()):
        try:
            country_rows = countries[country_code]
            country_sheet = wb.create_sheet(title=country_code)
            
            # Same headers as summary
            country_sheet.append(display_headers)
            
            # Add rows for this country
            for r in country_rows:
                _write_row_to_sheet(country_sheet, r)
            
            # Adjust column widths
            try:
                for col_idx in range(1, len(display_headers) + 1):
                    max_len = max(
                        (len(str(cell.value)) for cell in country_sheet[get_column_letter(col_idx)] if cell.value is not None),
                        default=10
                    )
                    country_sheet.column_dimensions[get_column_letter(col_idx)].width = min(60, max(12, max_len + 2))
            except Exception:
                pass
        except Exception as e:
            logger.debug("Failed to create country sheet for %s: %s", country_code, e)

    # create per-file sheets (key/value) - AFTER country summaries
    used_names = set()
    for r in rows:
        base = r.get("source_pdf", "sheet")
        title = _sanitize_sheet_title(str(base))
        original = title
        i = 1
        while title in used_names or title in wb.sheetnames:
            suffix = f"_{i}"
            max_base_len = 31 - len(suffix)
            title = (original[:max_base_len] + suffix) if len(original) > max_base_len else (original + suffix)
            i += 1
        used_names.add(title)
        ws = wb.create_sheet(title=title)

        ordered_keys = [
            "code_banque", "date_reference", "reference", "type_MT", "pays_iso3",
            "code_donneur_dordre", "donneur_dordre", "institution_name", "beneficiaire", 
            "correspondant", "montant", "devise", "commentaires", "source_pdf"
        ]
        written = set()
        for k in ordered_keys:
            if k in r:
                label = "Code du donneur d'ordre" if k == "code_donneur_dordre" else ("donneur d'ordre" if k in ("donneur_dordre", "institution_name") else ("Bénéficiaire" if k == "beneficiaire" else ("correspondant" if k == "correspondant" else ("commentaires" if k == "commentaires" else k))))
                ws.append([label, r.get(k)])
                written.add(k)
        for k, v in r.items():
            if k in written:
                continue
            label = "Code du donneur d'ordre" if k == "code_donneur_dordre" else ("donneur d'ordre" if k in ("donneur_dordre", "institution_name") else ("Bénéficiaire" if k == "beneficiaire" else ("correspondant" if k == "correspondant" else ("commentaires" if k == "commentaires" else k))))
            ws.append([label, v])

        # adjust column widths heuristically
        try:
            max_len_col1 = max((len(str(row[0])) for row in ws.values if row[0] is not None), default=10)
            max_len_col2 = max((len(str(row[1])) for row in ws.values if len(row) > 1 and row[1] is not None), default=10)
            ws.column_dimensions[get_column_letter(1)].width = min(60, max(12, max_len_col1 + 2))
            ws.column_dimensions[get_column_letter(2)].width = min(80, max(12, max_len_col2 + 8))
        except Exception:
            pass

    # Single final save after all sheets are created
    wb.save(out_path)
    logger.info("Workbook created with country sheets: %s", out_path)
    return out_path


def extract_mt900_only(pdf_path: Path, bic_xlsx: Optional[str] = None) -> tuple[List[Dict], Dict[str, set]]:
    """
    Extraire uniquement les MT900 d'un fichier PDF.
    
    Args:
        pdf_path: Chemin vers le fichier PDF
        bic_xlsx: Chemin optionnel vers le fichier Excel BIC
        
    Returns:
        tuple: (mt900_rows, missing_codes)
    """
    p = Path(pdf_path)
    
    if HAS_MT_MULTI and mt_multi_module:
        try:
            return mt_multi_module.extract_mt900_only(p, bic_xlsx=bic_xlsx)
        except Exception as e:
            logger.exception("extract_mt900_only: failed for %s: %s", p.name, e)
    
    return [], {"unmapped": set(), "empty": set()}


def match_mt900_with_transfers(mt900_rows: List[Dict], transfer_rows: List[Dict]) -> tuple[List[Dict], List[Dict], List[Dict], List[Dict]]:
    """
    Matcher les MT900 avec les MT103/MT202 via la référence d'origine (F21).
    Applique également les règles d'exception pour MT900.
    
    Logique:
    - Vérifier d'abord si le MT900 doit être en exception (T2PL, NIVLT)
    - Pour chaque MT900, chercher le MT103/MT202 dont F20 (référence) = F21 du MT900 (related_reference)
    - Compléter les infos du MT900 (donneur d'ordre, pays) depuis le MT103/MT202 matché
    - Les MT103/MT202 sans correspondance vont dans suspens
    - Les MT900 sans correspondance vont dans unmatched_mt900_rows
    
    Args:
        mt900_rows: Liste des MT900 extraits
        transfer_rows: Liste des MT103/MT202 extraits
        
    Returns:
        tuple: (matched_mt900_rows, suspens_rows, exception_mt900_rows, unmatched_mt900_rows)
    """
    # Fonction pour vérifier les exceptions MT900
    def _check_mt900_exception(row: Dict) -> Optional[str]:
        """
        Vérifier si le MT900 doit être mis en exception.
        """
        if not row or not row.get("type_MT"):
            return None
        
        if not row.get("type_MT").startswith("fin.900"):
            return None
        
        # Vérifier F20 (référence)
        reference = row.get("reference") or ""
        reference_upper = reference.upper()
        
        if "T2PL" in reference_upper:
            return "T2PL"
        
        # Vérifier F21 (référence d'origine)
        related_reference = row.get("related_reference") or ""
        related_ref_upper = related_reference.upper()
        
        if "NIVLT" in related_ref_upper or "NIVELLEMENT" in related_ref_upper:
            return "nivellement"
        
        return None
    
    # Créer un index des MT103/MT202 par référence pour le matching rapide
    ref_index: Dict[str, Dict] = {}
    for row in transfer_rows:
        ref = row.get("reference")
        if ref:
            ref_key = str(ref).strip().upper()
            ref_index[ref_key] = row
    
    # Matcher les MT900 avec les MT103/MT202
    matched_mt900_rows: List[Dict] = []  # MT900 avec correspondant trouvé
    unmatched_mt900_rows: List[Dict] = []  # MT900 sans correspondant
    exception_mt900_rows: List[Dict] = []
    matched_refs: set = set()
    
    for mt900_row in mt900_rows:
        # Vérifier d'abord si le MT900 doit être mis en exception
        exception_comment = _check_mt900_exception(mt900_row)
        if exception_comment:
            mt900_row["commentaires"] = exception_comment
            exception_mt900_rows.append(mt900_row)
            continue  # Ne pas matcher ce message, il va directement en exception
        
        # La référence d'origine F21 est utilisée pour matcher
        related_ref = mt900_row.get("related_reference")
        
        has_match = False
        if related_ref:
            ref_key = str(related_ref).strip().upper()
            
            if ref_key in ref_index:
                # Match trouvé ! Compléter les infos du MT900 avec celles du MT103/MT202
                matched_transfer = ref_index[ref_key]
                
                # Copier les infos manquantes du MT103/MT202 vers le MT900
                fields_to_copy = [
                    "code_donneur_dordre", "donneur_dordre", "beneficiaire", 
                    "pays_iso3", "correspondant"
                ]
                for field in fields_to_copy:
                    if not mt900_row.get(field) and matched_transfer.get(field):
                        mt900_row[field] = matched_transfer[field]
                
                # Ajouter info de matching
                mt900_row["matched_with"] = matched_transfer.get("reference")
                mt900_row["matched_type"] = matched_transfer.get("type_MT")
                # Ajouter la source du message matché pour la traçabilité
                mt900_row["matched_source"] = matched_transfer.get("source_pdf")
                
                matched_refs.add(ref_key)
                has_match = True
        
        if has_match:
            matched_mt900_rows.append(mt900_row)
        else:
            # MT900 sans correspondant trouvé
            mt900_row["commentaires"] = "pas de correspondant MT103/MT202"
            unmatched_mt900_rows.append(mt900_row)
    
    # Les MT103/MT202 non matchés vont dans suspens
    suspens_rows: List[Dict] = []
    for row in transfer_rows:
        ref = row.get("reference")
        if ref:
            ref_key = str(ref).strip().upper()
            if ref_key not in matched_refs:
                row["status"] = "suspens - pas de confirmation MT900"
                suspens_rows.append(row)
        else:
            # Pas de référence = suspens
            row["status"] = "suspens - référence manquante"
            suspens_rows.append(row)
    
    logger.info("match_mt900_with_transfers: %d matched, %d unmatched MT900, %d suspens, %d exceptions", 
                len(matched_mt900_rows), len(unmatched_mt900_rows), len(suspens_rows), len(exception_mt900_rows))
    
    return matched_mt900_rows, suspens_rows, exception_mt900_rows, unmatched_mt900_rows


def extract_transfer_analysis_dispatch(pdf_path: Path) -> tuple[List[Dict], List[Dict], Dict[str, set]]:
    """
    Dispatch pour l'analyse des transferts sortants exécutés.
    Utilise mt_multi.extract_transfer_analysis.
    
    Returns:
        tuple: (matched_900_rows, suspens_rows, missing_codes)
    """
    p = Path(pdf_path)
    
    if HAS_MT_MULTI and mt_multi_module:
        try:
            return mt_multi_module.extract_transfer_analysis(p)
        except Exception as e:
            logger.exception("extract_transfer_analysis_dispatch: failed for %s: %s", p.name, e)
    
    # Fallback: retourner des listes vides
    return [], [], {"unmapped": set(), "empty": set()}


def create_transfer_analysis_workbook(matched_rows: List[Dict], suspens_rows: List[Dict], exception_rows: List[Dict], out_dir: Path, date_start: str = None, date_end: str = None, unmatched_mt900_rows: List[Dict] = None) -> Path:
    """
    Créer un workbook Excel pour l'analyse des transferts sortants exécutés.
    
    Sheets:
    - "Transferts_Executes": MT900 matchés avec leurs infos complétées
    - "MT900_non_matches": MT900 sans correspondant MT103/MT202
    - "Suspens": MT202/MT103 sans confirmation MT900
    - "Exceptions": MT900 en exception (T2PL, nivellement)
    
    Args:
        matched_rows: Liste des MT900 matchés
        suspens_rows: Liste des MT202/MT103 sans correspondant
        exception_rows: Liste des MT900 en exception
        out_dir: Répertoire de sortie
        date_start: Date de début optionnelle
        date_end: Date de fin optionnelle
        unmatched_mt900_rows: Liste des MT900 sans correspondant MT103/MT202
    """
    if unmatched_mt900_rows is None:
        unmatched_mt900_rows = []
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    date_suffix = ""
    if date_start and date_end:
        date_suffix = f"_{date_start}_to_{date_end}"
    elif date_start:
        date_suffix = f"_from_{date_start}"
    elif date_end:
        date_suffix = f"_to_{date_end}"
    
    out_path = out_dir / f"analyse_transferts{date_suffix}_{ts}.xlsx"
    
    wb = Workbook()
    
    # Headers pour les feuilles
    transfer_headers = [
        "type_MT",
        "reference",
        "related_reference",
        "date_reference",
        "devise",
        "montant",
        "Code du donneur d'ordre",
        "donneur d'ordre",
        "Bénéficiaire",
        "pays_iso3",
        "correspondant",
        "Message correspondant",
        "source_pdf"
    ]
    
    suspens_headers = [
        "type_MT",
        "reference",
        "date_reference",
        "devise",
        "montant",
        "Code du donneur d'ordre",
        "donneur d'ordre",
        "Bénéficiaire",
        "pays_iso3",
        "correspondant",
        "status",
        "source_pdf"
    ]
    
    def _convert_date_to_excel(date_str):
        if not date_str:
            return None
        try:
            if isinstance(date_str, str) and re.match(r'^\d{4}-\d{2}-\d{2}$', date_str):
                return datetime.strptime(date_str, '%Y-%m-%d')
            return date_str
        except Exception:
            return date_str
    
    # Sheet 1: Transferts Exécutés (MT900 matchés)
    exec_sheet = wb.active
    exec_sheet.title = "Transferts_Executes"
    exec_sheet.append(transfer_headers)
    
    for r in matched_rows:
        date_ref = _convert_date_to_excel(r.get("date_reference"))
        
        # Construire le texte "Message correspondant" à partir des infos de matching
        # Format: "voir message N°X du fichier Y.pdf" si la source est au format attendu
        matched_source = r.get("matched_source") or ""
        matched_type = r.get("matched_type") or ""
        if matched_source:
            # La source est déjà au format "voir message N°X du fichier Y.pdf" ou juste "Y.pdf"
            message_correspondant = f"{matched_type} - {matched_source}" if matched_type else matched_source
        else:
            message_correspondant = None
        
        row_data = [
            r.get("type_MT"),
            r.get("reference"),
            r.get("related_reference"),
            date_ref,
            r.get("devise"),
            r.get("montant"),
            r.get("code_donneur_dordre"),
            r.get("donneur_dordre"),
            r.get("beneficiaire"),
            r.get("pays_iso3"),
            r.get("correspondant"),
            message_correspondant,
            r.get("source_pdf")
        ]
        exec_sheet.append(row_data)
        
        if date_ref and isinstance(date_ref, datetime):
            current_row = exec_sheet.max_row
            exec_sheet.cell(row=current_row, column=4).number_format = 'DD/MM/YYYY'
    
    # Ajuster largeurs de colonnes
    try:
        for col_idx in range(1, len(transfer_headers) + 1):
            max_len = max(
                (len(str(cell.value)) for cell in exec_sheet[get_column_letter(col_idx)] if cell.value is not None),
                default=10
            )
            exec_sheet.column_dimensions[get_column_letter(col_idx)].width = min(60, max(12, max_len + 2))
    except Exception:
        pass
    
    # Sheet 2: MT900_non_matches (MT900 sans correspondant MT103/MT202)
    if unmatched_mt900_rows:
        unmatched_headers = [
            "type_MT",
            "reference",
            "related_reference",
            "date_reference",
            "devise",
            "montant",
            "correspondant",
            "Commentaires",
            "source_pdf"
        ]
        unmatched_sheet = wb.create_sheet(title="MT900_non_matches")
        unmatched_sheet.append(unmatched_headers)
        
        for r in unmatched_mt900_rows:
            date_ref = _convert_date_to_excel(r.get("date_reference"))
            row_data = [
                r.get("type_MT"),
                r.get("reference"),
                r.get("related_reference"),
                date_ref,
                r.get("devise"),
                r.get("montant"),
                r.get("correspondant"),
                r.get("commentaires"),
                r.get("source_pdf")
            ]
            unmatched_sheet.append(row_data)
            
            if date_ref and isinstance(date_ref, datetime):
                current_row = unmatched_sheet.max_row
                unmatched_sheet.cell(row=current_row, column=4).number_format = 'DD/MM/YYYY'
        
        # Ajuster largeurs de colonnes
        try:
            for col_idx in range(1, len(unmatched_headers) + 1):
                max_len = max(
                    (len(str(cell.value)) for cell in unmatched_sheet[get_column_letter(col_idx)] if cell.value is not None),
                    default=10
                )
                unmatched_sheet.column_dimensions[get_column_letter(col_idx)].width = min(60, max(12, max_len + 2))
        except Exception:
            pass
    
    # Sheet 3: Suspens (MT202/MT103 sans correspondant)
    if suspens_rows:
        suspens_sheet = wb.create_sheet(title="Suspens")
        suspens_sheet.append(suspens_headers)
        
        for r in suspens_rows:
            date_ref = _convert_date_to_excel(r.get("date_reference"))
            row_data = [
                r.get("type_MT"),
                r.get("reference"),
                date_ref,
                r.get("devise"),
                r.get("montant"),
                r.get("code_donneur_dordre"),
                r.get("donneur_dordre"),
                r.get("beneficiaire"),
                r.get("pays_iso3"),
                r.get("correspondant"),
                r.get("status"),
                r.get("source_pdf")
            ]
            suspens_sheet.append(row_data)
            
            if date_ref and isinstance(date_ref, datetime):
                current_row = suspens_sheet.max_row
                suspens_sheet.cell(row=current_row, column=3).number_format = 'DD/MM/YYYY'
        
        # Ajuster largeurs de colonnes
        try:
            for col_idx in range(1, len(suspens_headers) + 1):
                max_len = max(
                    (len(str(cell.value)) for cell in suspens_sheet[get_column_letter(col_idx)] if cell.value is not None),
                    default=10
                )
                suspens_sheet.column_dimensions[get_column_letter(col_idx)].width = min(60, max(12, max_len + 2))
        except Exception:
            pass
    
    # Sheet 3: Exceptions (MT900 en exception)
    if exception_rows:
        exception_sheet = wb.create_sheet(title="Exceptions")
        
        # Headers pour les exceptions
        exception_headers = [
            "type_MT",
            "reference",
            "related_reference",
            "date_reference",
            "devise",
            "montant",
            "Code du donneur d'ordre",
            "donneur d'ordre",
            "Bénéficiaire",
            "pays_iso3",
            "correspondant",
            "Commentaires",
            "source_pdf"
        ]
        exception_sheet.append(exception_headers)
        
        for r in exception_rows:
            date_ref = _convert_date_to_excel(r.get("date_reference"))
            row_data = [
                r.get("type_MT"),
                r.get("reference"),
                r.get("related_reference"),
                date_ref,
                r.get("devise"),
                r.get("montant"),
                r.get("code_donneur_dordre"),
                r.get("donneur_dordre"),
                r.get("beneficiaire"),
                r.get("pays_iso3"),
                r.get("correspondant"),
                r.get("commentaires"),
                r.get("source_pdf")
            ]
            exception_sheet.append(row_data)
            
            if date_ref and isinstance(date_ref, datetime):
                current_row = exception_sheet.max_row
                exception_sheet.cell(row=current_row, column=4).number_format = 'DD/MM/YYYY'
        
        # Ajuster largeurs de colonnes
        try:
            for col_idx in range(1, len(exception_headers) + 1):
                max_len = max(
                    (len(str(cell.value)) for cell in exception_sheet[get_column_letter(col_idx)] if cell.value is not None),
                    default=10
                )
                exception_sheet.column_dimensions[get_column_letter(col_idx)].width = min(60, max(12, max_len + 2))
        except Exception:
            pass
    
    wb.save(out_path)
    logger.info("Transfer analysis workbook created: %s (matched: %d, unmatched_mt900: %d, suspens: %d, exceptions: %d)", 
                out_path, len(matched_rows), len(unmatched_mt900_rows), len(suspens_rows), len(exception_rows))
    return out_path