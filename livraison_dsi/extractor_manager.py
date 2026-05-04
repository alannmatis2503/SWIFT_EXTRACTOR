# backend/app/extractor_manager.py
import os
import sys
from pathlib import Path
import re
import logging
from datetime import datetime
from typing import List, Dict, Optional

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
from openpyxl.cell import WriteOnlyCell

from utils import logger

# import extractors (primary helpers)
from extractors.mt202 import extract_for_mt202, extract_text_from_pdf as extract_text_mt202
try:
    from extractors.mt103 import extract_for_mt103
    HAS_MT103 = True
except Exception:
    HAS_MT103 = False
    logger.info("mt103 extractor not available at import time; you can add it to EXTRACTOR_MAP later.")

try:
    from extractors.mt910 import extract_for_mt910
    HAS_MT910 = True
except Exception:
    HAS_MT910 = False
    logger.info("mt910 extractor not available at import time.")

# try to import the multi-message extractor (optional)
try:
    from extractors import mt_multi as mt_multi_module
    HAS_MT_MULTI = True
except Exception:
    mt_multi_module = None
    HAS_MT_MULTI = False
    logger.info("mt_multi extractor not available at import time; multi-file detection will fall back to single extractors.")

# bic_utils: mapping code Reglement -> {nom, BIC, pays}
try:
    from extractors import bic_utils as _bic_utils
    HAS_BIC_UTILS = True
except Exception:
    _bic_utils = None
    HAS_BIC_UTILS = False
    logger.info("bic_utils extractor not available at import time; reglement mapping will be skipped.")

# regex to detect MT/FIN code
MT_DETECT = re.compile(r'\b(?:MT|FIN)[\s\-\_\.:\/]*(\d{3})\b', re.I)
IDENTIFIER_FIN_RE = re.compile(r'Identifier[:\s]*fin[\.:\s\-\/]*(\d{3})', re.I)

# map MT number -> extractor callable
EXTRACTOR_MAP = {
    "202": extract_for_mt202,
}
if HAS_MT103:
    EXTRACTOR_MAP["103"] = extract_for_mt103
if HAS_MT910:
    EXTRACTOR_MAP["910"] = extract_for_mt910

# -----------------------------
# BIC mapping / donor logic
# -----------------------------
# caching globals
_cached_mapping: Optional[Dict[str, str]] = None
_cached_mapping_path: Optional[str] = None

# heuristics: possible default file locations
_DEFAULT_XLS_PATHS = [
    "data/bic_codes.xlsx",
    "data/bic.xlsx",
    "bic_codes.xlsx",
    "bic.xlsx",
    "data/bfde98b8-0a94-4ba1-ab8a-eae27357cc7e.xlsx"  # the uploaded name you used earlier (kept as candidate)
]

def _find_columns(df):
    """
    Find likely code_col and name_col heuristically from DataFrame columns.
    """
    cols = list(df.columns)
    code_col = None
    name_col = None
    for c in cols:
        cu = c.upper()
        if ('BIC' in cu) or (('CODE' in cu) and ('BIC' in cu or 'SWIFT' in cu)):
            code_col = c
            break
    if not code_col:
        for c in cols:
            cu = c.upper()
            if 'CODE' in cu:
                code_col = c
                break

    for c in cols:
        cu = c.upper()
        if 'NOM' in cu or 'NAME' in cu or 'NOMS' in cu:
            name_col = c
            break
    if not name_col:
        candidate = None
        best_alpha = 0.0
        for c in cols:
            sample = ' '.join([str(x) for x in df[c].dropna().astype(str).head(20).tolist()])
            if not sample:
                continue
            alpha_frac = sum(ch.isalpha() for ch in sample) / max(1, len(sample))
            if alpha_frac > best_alpha:
                best_alpha = alpha_frac
                candidate = c
        name_col = candidate
    return code_col, name_col


def bundled_base_path() -> Path:
    """
    Retourne le chemin racine d'où lire les fichiers "embarqués".
    - Si l'app est packagée par PyInstaller (--onefile), les resources sont extraites
      temporairement dans sys._MEIPASS.
    - Sinon, retourne la racine du projet (deux niveaux au-dessus de ce fichier).
    """
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        return Path(sys._MEIPASS)
    # adjust parents count depending on file location; this file is backend/app/extractor_manager.py
    # parents[2] points to repo root (pdf-extractor/)
    return Path(__file__).resolve().parents[2]

def _user_override_bic_paths() -> List[Path]:
    """
    Emplacements où un admin/utilisateur peut déposer un bic_codes.xlsx modifiable.
    Ordre de priorité (testé dans load_bic_mapping) :
      1) variable d'environnement PDF_SWIFT_DATA_DIR si définie
      2) dossier commun ProgramData (Windows) -> {PROGRAMDATA}/PDF_Swift_Extractor/data
      3) dossier local de l'utilisateur -> %LOCALAPPDATA%/PDF_Swift_Extractor/data ou ~/ .pdf_swift_extractor/data
    """
    paths = []
    env = os.getenv("PDF_SWIFT_DATA_DIR")
    if env:
        paths.append(Path(env))

    # Windows common appdata (ProgramData)
    programdata = os.getenv("PROGRAMDATA")
    if programdata:
        paths.append(Path(programdata) / "PDF_Swift_Extractor" / "data")

    # Windows local appdata or cross-platform user dir
    localappdata = os.getenv("LOCALAPPDATA") or os.getenv("XDG_DATA_HOME")
    if localappdata:
        paths.append(Path(localappdata) / "PDF_Swift_Extractor" / "data")

    # fallback to user home hidden dir
    paths.append(Path.home() / ".pdf_swift_extractor" / "data")

    return paths

def load_bic_mapping(xlsx_path: Optional[str] = None, sheet_name: Optional[str] = 0) -> Dict[str, str]:
    """
    Charge (et met en cache) la table BIC -> nom de banque depuis un fichier Excel.
    Logique améliorée pour permettre une mise à jour manuelle après installation :
      - si xlsx_path explicit fourni -> utilisé (existing behaviour)
      - sinon : on cherche d'abord dans des emplacements externes éditables (ProgramData, user dir,
        variable d'env PDF_SWIFT_DATA_DIR)
      - sinon : on cherche dans le bundle embarqué (bundled_base_path()/data)
      - sinon : on retombe sur les chemins _DEFAULT_XLS_PATHS comme avant
    """
    global _cached_mapping, _cached_mapping_path
    import pandas as pd  # lazy import

    # 1) if explicit path provided, prefer it
    if xlsx_path:
        p = Path(xlsx_path)
        if not p.exists():
            raise FileNotFoundError(f"Provided xlsx_path not found: {xlsx_path}")
    else:
        # 2) check user-writable override locations (ProgramData, LOCALAPPDATA, env var)
        p = None
        for base in _user_override_bic_paths():
            candidate = base / "bic_codes.xlsx"
            if candidate.exists():
                p = candidate
                break
            # also accept alternative names
            alt = base / "bic.xlsx"
            if alt.exists():
                p = alt
                break

        # 3) check bundled data dir (this will work for --onedir and for files added with --add-data)
        if p is None:
            bundled = bundled_base_path() / "data"
            for name in ("bic_codes.xlsx", "bic.xlsx"):
                cand = bundled / name
                if cand.exists():
                    p = cand
                    break

        # 4) fallback to original candidate list (relative to current working dir)
        if p is None:
            for cand in _DEFAULT_XLS_PATHS:
                if Path(cand).exists():
                    p = Path(cand)
                    break

        if p is None:
            raise FileNotFoundError(
                "Aucun fichier Excel trouvé. Place your Excel mapping in one of: "
                + ", ".join(_DEFAULT_XLS_PATHS)
                + " or a writable location like %PROGRAMDATA%\\PDF_Swift_Extractor\\data\\bic_codes.xlsx "
                + "or set environment variable PDF_SWIFT_DATA_DIR to a folder containing bic_codes.xlsx"
            )

    # cache check
    pstr = str(Path(p).resolve())
    if _cached_mapping is not None and _cached_mapping_path == pstr:
        return _cached_mapping

    df = pd.read_excel(pstr, sheet_name=sheet_name, dtype=str)
    df = df.dropna(axis=1, how='all')

    code_col, name_col = _find_columns(df)
    if not code_col:
        raise ValueError(f"Impossible de détecter la colonne code BIC dans {pstr}. Colonnes: {list(df.columns)}")
    if not name_col:
        logger.warning("load_bic_mapping: impossible de détecter colonne 'nom' ; les valeurs de nom seront vides.")

    mapping: Dict[str, str] = {}
    for _, row in df.iterrows():
        code_val = (str(row.get(code_col) or "")).strip()
        name_val = (str(row.get(name_col) or "")).strip() if name_col else ""
        if not code_val or code_val.lower() in ("nan", "none"):
            continue
        code_clean = re.sub(r'\s+', '', code_val).upper()
        key = code_clean[:8]
        if not key:
            continue
        mapping[key] = name_val

    _cached_mapping = mapping
    _cached_mapping_path = pstr
    logger.info("load_bic_mapping: loaded %d entries from %s", len(mapping), pstr)
    return mapping


FALLBACK_11_RE = re.compile(r'\b([A-Z0-9]{11})\b')

# remplace la fonction get_donneur_from_f52 existante par ceci
IDENTIFIER_RE_AFTER_LABEL = re.compile(
    r"(?i)(?:IdentifierCode|Identifier Code|Identifiercode|Code d'identifiant|Code d identifiant|IDENTIFIERCODE)\s*[:\-\s]*\n?\s*([A-Z0-9]{11})"
)

# words that look like labels and should NOT be accepted as code
_BAD_LABEL_TOKENS = {
    "IDENTIFIER", "IDENTIFIERC", "PARTYIDENTI", "PARTYIDENT", "IDENTIFIANT", "IDENTIFIERCODE",
    "PARTY", "PARTYIDENTIFIER"
}


# label regex (variantes FR/EN)
_LABEL_RE = re.compile(
    r"(?i)(?:IdentifierCode|Identifier Code|Identifiercode|Code d'identifiant|Code d identifiant|identifiant de partie|IDENTIFIERCODE)\s*[:\-\s]*",
    re.M
)

_BAD_LABEL_PREFIXES = ("IDENTIF", "PARTYIDENT", "PARTY", "IDENTIFIANT")

def _find_identifier_after_label(text: str, lookahead_chars: int = 600) -> Optional[str]:
    """
    Cherche après un label 'IdentifierCode' un token alphanumérique 6..11 caractères,
    autorise lignes vides entre label et token, privilégie tokens contenant des lettres.
    """
    if not text:
        return None
    txt = text.replace('\r', '\n')
    m = _LABEL_RE.search(txt)
    if not m:
        return None
    start = m.end()
    tail = txt[start: start + lookahead_chars]
    # find candidates 6..11 chars
    toks = re.findall(r'\b([A-Z0-9]{6,11})\b', tail, flags=re.I)
    toks = [t.upper() for t in toks]
    # filter label-like tokens
    toks = [t for t in toks if not any(t.startswith(pref) for pref in _BAD_LABEL_PREFIXES)]
    if not toks:
        return None
    # prefer token with a letter (likely BIC), otherwise return first
    for t in toks:
        if re.search(r'[A-Z]', t):
            return t
    return toks[0]

def get_donneur_from_f52(f52_text: Optional[str], message_text: Optional[str] = None, xlsx_path: Optional[str] = None) -> Optional[str]:
    """
    Robust extract:
     - try inside F52A block: find label then next token (6..11 chars), preferring alpha tokens
     - if not found, try search on the whole message (cross-page)
     - if still not found, attempt a last-resort token in F52A that contains letters
     - then map first 8 chars to bank name using load_bic_mapping (if available)
    Returns: "CODE11/BANK NAME" if mapping found, else CODE (or None).
    """
    # normalize and remove xml-like tags
    def _norm(s):
        return re.sub(r'<[^>]+>', ' ', (s or "")).replace('\r', '\n')

    f52 = _norm(f52_text)
    full = _norm(message_text) if message_text else None

    # 1) try strictly in F52A block
    code = _find_identifier_after_label(f52, lookahead_chars=800)
    # 2) if not found, try whole message (cross-page)
    if not code and full:
        code = _find_identifier_after_label(full, lookahead_chars=1200)
    # 3) last-resort: try to find first alnum token with letters in F52A
    if not code and f52:
        m = re.search(r'\b([A-Z][A-Z0-9]{5,10})\b', f52, flags=re.I)
        if m:
            tok = m.group(1).upper()
            if not any(tok.startswith(pref) for pref in _BAD_LABEL_PREFIXES):
                code = tok

    if not code:
        return None

    # attempt mapping to bank name (uses cached loader)
    try:
        mapping = load_bic_mapping(xlsx_path=xlsx_path)
    except Exception:
        mapping = {}

    key8 = code[:8].upper()
    bank = mapping.get(key8)
    if bank:
        bank_clean = re.sub(r'\s{2,}', ' ', bank).strip()
        return f"{code}/{bank_clean}"
    return code


# -----------------------------
# Dispatcher / workbook logic (existing)
# -----------------------------
def detect_message_type(text: str) -> Optional[str]:
    """
    Detect the MT type (e.g. "202", "103", "910") from extracted text.
    Returns the numeric string (e.g. "202") or None.
    """
    if not text:
        return None

    m = MT_DETECT.search(text)
    if m:
        mt = m.group(1)
        logger.debug("detect_message_type: primary MT_DETECT matched -> %s", mt)
        return mt

    m2 = IDENTIFIER_FIN_RE.search(text)
    if m2:
        mt = m2.group(1)
        logger.debug("detect_message_type: IDENTIFIER_FIN_RE matched -> %s", mt)
        return mt

    logger.debug("detect_message_type: no MT type matched")
    return None


def extract_dispatch(pdf_path: Path, direction: str = "incoming") -> tuple[List[Dict], List[Dict], List[Dict], List[Dict], List[Dict], List[Dict], Dict[str, set]]:
    """
    Dispatcher intelligent :
      - si le PDF contient plusieurs messages -> utilise mt_multi.extract_messages_from_pdf
      - sinon -> utilise extract_single (retourne [row])
    Retourne toujours : (liste de rows, liste de BEACCMCX091 rows, liste de 323201 exception rows, 
                         liste d'autres exceptions (EUR/nivellement), liste BANQUE DE FRANCE rows,
                         dict de codes manquants)
    
    Args:
        pdf_path: Path to the PDF file
        direction: "incoming" or "outgoing" - determines beneficiary extraction logic
    """
    p = Path(pdf_path)
    # quick text extraction using existing helper
    text = ""
    try:
        text = extract_text_mt202(p)
    except Exception as e:
        logger.debug("extract_dispatch: extract_text_mt202 failed (%s), falling back to pdfplumber", e)
        try:
            import pdfplumber
            s = ""
            with pdfplumber.open(str(p)) as pdf:
                for page in pdf.pages[:2]:
                    s += "\n" + (page.extract_text() or "")
            text = s
        except Exception as e2:
            logger.warning("extract_dispatch: quick pdfplumber fallback failed for %s: %s", p.name, e2)
            text = ""

    missing_codes: Dict[str, set] = {"unmapped": set(), "empty": set()}
    beaccmcx091_rows: List[Dict] = []  # Liste séparée pour les messages BEACCMCX091
    exception_323201_rows: List[Dict] = []  # Liste pour les exceptions 323201
    other_exceptions_rows: List[Dict] = []  # Liste pour les autres exceptions (EUR/nivellement)
    banque_de_france_rows: List[Dict] = []  # MT103 USD avec BANQUE DE FRANCE / FW021083459

    # If multi-message extractor available, use its split logic to decide
    if HAS_MT_MULTI and mt_multi_module:
        try:
            blocks = mt_multi_module._split_messages(text)
            if blocks and len(blocks) > 1:
                logger.info("%s: detected %d messages (using mt_multi).", p.name, len(blocks))
                # Pass preloaded_text to avoid re-reading the PDF
                rows, beaccmcx091_rows, exception_323201_rows, other_exceptions_rows, banque_de_france_rows, forex_rows, bdf_corr_exception_rows, missing_codes = mt_multi_module.extract_messages_from_pdf(p, direction=direction, preloaded_text=text)
                # ensure backward compatibility: set institution_name from donneur_dordre if missing
                for r in rows:
                    if "institution_name" not in r or not r.get("institution_name"):
                        r["institution_name"] = r.get("donneur_dordre") or r.get("donneur d'ordre") or None
                    for k in ["date_reference", "reference", "type_MT", "pays_iso3", "beneficiaire", "montant", "devise", "source_pdf", "commentaires", "correspondant"]:
                        if k not in r:
                            r[k] = None
                # apply same compatibility to BEACCMCX091 rows
                for r in beaccmcx091_rows:
                    if "institution_name" not in r or not r.get("institution_name"):
                        r["institution_name"] = r.get("donneur_dordre") or r.get("donneur d'ordre") or None
                    for k in ["date_reference", "reference", "type_MT", "pays_iso3", "beneficiaire", "montant", "devise", "source_pdf", "commentaires", "correspondant"]:
                        if k not in r:
                            r[k] = None
                # apply same compatibility to exception_323201 rows
                for r in exception_323201_rows:
                    if "institution_name" not in r or not r.get("institution_name"):
                        r["institution_name"] = r.get("donneur_dordre") or r.get("donneur d'ordre") or None
                    for k in ["date_reference", "reference", "type_MT", "pays_iso3", "beneficiaire", "montant", "devise", "source_pdf", "commentaires", "correspondant"]:
                        if k not in r:
                            r[k] = None
                # apply same compatibility to other_exceptions rows
                for r in other_exceptions_rows:
                    if "institution_name" not in r or not r.get("institution_name"):
                        r["institution_name"] = r.get("donneur_dordre") or r.get("donneur d'ordre") or None
                    for k in ["date_reference", "reference", "type_MT", "pays_iso3", "beneficiaire", "montant", "devise", "source_pdf", "commentaires", "correspondant"]:
                        if k not in r:
                            r[k] = None
                # apply same compatibility to banque_de_france rows
                for r in banque_de_france_rows:
                    if "institution_name" not in r or not r.get("institution_name"):
                        r["institution_name"] = r.get("donneur_dordre") or r.get("donneur d'ordre") or None
                    for k in ["date_reference", "reference", "type_MT", "pays_iso3", "beneficiaire", "montant", "devise", "source_pdf", "commentaires", "correspondant"]:
                        if k not in r:
                            r[k] = None
                # apply same compatibility to forex rows
                for r in forex_rows:
                    if "institution_name" not in r or not r.get("institution_name"):
                        r["institution_name"] = r.get("donneur_dordre") or r.get("donneur d'ordre") or None
                    for k in ["date_reference", "reference", "type_MT", "pays_iso3", "beneficiaire", "montant", "devise", "source_pdf", "commentaires", "correspondant"]:
                        if k not in r:
                            r[k] = None
                # apply same compatibility to bdf_corr_exception rows
                for r in bdf_corr_exception_rows:
                    if "institution_name" not in r or not r.get("institution_name"):
                        r["institution_name"] = r.get("donneur_dordre") or r.get("donneur d'ordre") or None
                    for k in ["date_reference", "reference", "type_MT", "pays_iso3", "beneficiaire", "montant", "devise", "source_pdf", "commentaires", "correspondant"]:
                        if k not in r:
                            r[k] = None
                return rows, beaccmcx091_rows, exception_323201_rows, other_exceptions_rows, banque_de_france_rows, forex_rows, bdf_corr_exception_rows, missing_codes
        except Exception as e:
            logger.exception("extract_dispatch: mt_multi detection/extraction failed for %s: %s", p.name, e)
            # fall through to single extractor

    # fallback: treat as single message
    single_row = extract_single(p, direction=direction)
    return [single_row], beaccmcx091_rows, exception_323201_rows, other_exceptions_rows, banque_de_france_rows, [], [], missing_codes


def _ensure_minimal_row(p: Path, mt_type: Optional[str] = None) -> Dict:
    """Return a minimal row template used when extraction not performed or failed."""
    return {
        "date_reference": None,
        "reference": None,
        "type_MT": f"fin.{mt_type}" if mt_type else None,
        "pays_iso3": None,
        "institution_name": None,
        "beneficiaire": None,
        "montant": None,
        "devise": None,
        "source_pdf": p.name
    }


def extract_single(pdf_path: Path, direction: str = "incoming") -> Dict:
    """
    Dispatch extraction for a single pdf_path (Path or str).
    Returns a dict with fields (internal keys). The create_workbook function maps
    'institution_name' -> "donneur d'ordre" when writing the summary sheet.
    
    Args:
        pdf_path: Path to the PDF file
        direction: "incoming" or "outgoing" - determines beneficiary extraction logic
    """
    p = Path(pdf_path)
    if not p.exists():
        logger.error("extract_single: file not found: %s", p)
        return _ensure_minimal_row(p)

    # read text (use helper from mt202 for consistent behavior)
    try:
        text = extract_text_mt202(p)
    except Exception as e:
        logger.exception("extract_single: extract_text_mt202 failed for %s: %s", p.name, e)
        # fallback quick text extraction
        try:
            import pdfplumber
            s = ""
            with pdfplumber.open(str(p)) as pdf:
                for page in pdf.pages[:2]:
                    s += "\n" + (page.extract_text() or "")
            text = s
        except Exception as e2:
            logger.exception("extract_single: fallback pdfplumber failed for %s: %s", p.name, e2)
            return _ensure_minimal_row(p)

    mt = detect_message_type(text)
    if not mt:
        logger.info("%s: MT type not found in text", p.name)
        row = _ensure_minimal_row(p, mt_type=None)
        row["source_pdf"] = p.name
        return row

    extractor = EXTRACTOR_MAP.get(mt)
    if not extractor:
        logger.info("%s: type detected -> %s but no extractor implemented", p.name, mt)
        row = _ensure_minimal_row(p, mt_type=mt)
        row["source_pdf"] = p.name
        return row

    try:
        # Pass direction to extractor if it supports it
        import inspect
        sig = inspect.signature(extractor)
        if 'direction' in sig.parameters:
            row = extractor(p, direction=direction)
        else:
            row = extractor(p)
            
        if not isinstance(row, dict):
            logger.error("%s: extractor returned non-dict result: %r", p.name, row)
            row = _ensure_minimal_row(p, mt_type=mt)
        else:
            required = ["date_reference", "reference", "type_MT", "pays_iso3",
                        "institution_name", "beneficiaire", "montant", "devise", "source_pdf"]
            for k in required:
                if k not in row:
                    row[k] = None
            if not row.get("institution_name") and row.get("donneur_dordre"):
                row["institution_name"] = row.get("donneur_dordre")
            if not row.get("type_MT"):
                row["type_MT"] = f"fin.{mt}"
            if not row.get("source_pdf"):
                row["source_pdf"] = p.name
        logger.info("%s: extracted via MT%s (direction: %s)", p.name, mt, direction)
        return row
    except Exception as e:
        logger.exception("Extraction failed for %s (MT%s): %s", p.name, mt, e)
        row = _ensure_minimal_row(p, mt_type=mt)
        row["error"] = str(e)
        return row


def _sanitize_sheet_title(name: str, max_len: int = 31) -> str:
    """Make a safe Excel sheet name (no invalid chars, limited length)."""
    if not name:
        name = "sheet"
    sanitized = re.sub(r'[:\\\/\?\*\[\]]+', '_', name)
    sanitized = sanitized.strip()
    if len(sanitized) > max_len:
        sanitized = sanitized[:max_len]
    if not sanitized:
        sanitized = "sheet"
    return sanitized


def create_workbook(rows: List[Dict], out_dir: Path, direction: str = "incoming", beaccmcx091_rows: Optional[List[Dict]] = None, exception_323201_rows: Optional[List[Dict]] = None, other_exceptions_rows: Optional[List[Dict]] = None, banque_de_france_rows: Optional[List[Dict]] = None, forex_rows: Optional[List[Dict]] = None, bdf_corr_exception_rows: Optional[List[Dict]] = None, date_start: str = None, date_end: str = None, nt_status_rejected_rows: Optional[List[Dict]] = None) -> Path:
    """
    Create an Excel workbook with:
      - a 'summary' sheet containing one row per extracted file (display headers in French)
      - a 'BEACCMCX091' sheet if beaccmcx091_rows is provided
      - a 'Exceptions_323201' sheet if exception_323201_rows is provided
      - a 'Autres_Exceptions' sheet for EUR/nivellement exceptions
      - a 'BANQUE DE FRANCE' sheet for MT103 USD with BDF patterns
      - a 'forex' sheet for MT910 incoming with forex donor codes
      - a 'Doublons_potentiels' sheet for potential duplicates (910 vs 202 with same ref+amount)
      - per-country summary sheets
      - one additional sheet per file with key/value pairs (debug-friendly)
    Returns the Path to the saved workbook.
    
    Args:
        rows: List of extracted data dictionaries
        out_dir: Output directory for the workbook
        direction: "incoming" or "outgoing" - used in filename
        beaccmcx091_rows: Optional list of BEACCMCX091 messages to display separately
        exception_323201_rows: Optional list of 323201 exception messages
        other_exceptions_rows: Optional list of other exceptions (EUR/nivellement)
        banque_de_france_rows: Optional list of MT103 USD BANQUE DE FRANCE messages
        forex_rows: Optional list of MT910 incoming forex messages
        date_start: Optional start date for filtering (YYYY-MM-DD)
        date_end: Optional end date for filtering (YYYY-MM-DD)
    """
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Ajouter la plage de dates au nom du fichier si spécifiée
    date_suffix = ""
    if date_start and date_end:
        date_suffix = f"_{date_start}_to_{date_end}"
    elif date_start:
        date_suffix = f"_from_{date_start}"
    elif date_end:
        date_suffix = f"_to_{date_end}"
    
    out_path = out_dir / f"swift_extraction_{direction}{date_suffix}_{ts}.xlsx"

    wb = Workbook()
    summary = wb.active
    summary.title = "summary"

    # summary headers (user-facing) - avec nouvelles colonnes
    display_headers = [
        "correspondant",
        "date_reference",
        "reference",
        "reference_origine",
        "type_MT",
        "pays_iso3",
        "Code du donneur d'ordre",
        "donneur d'ordre",
        "Bénéficiaire",
        "montant",
        "devise",
        "commentaires",
        "source_pdf"
    ]
    summary.append(display_headers)

    # Cache mémoïsé : datetime.strptime est très lent (~80µs/appel)
    # et la plupart des fichiers n'ont qu'une poignée de dates distinctes.
    _date_cache: Dict[str, Optional[datetime]] = {}
    _ISO_DATE_RE = re.compile(r'^\d{4}-\d{2}-\d{2}$')

    def _convert_date_to_excel(date_str):
        """Convertir une date ISO en datetime pour Excel (avec cache)."""
        if not date_str:
            return None
        # Si déjà un datetime, retour immédiat
        if isinstance(date_str, datetime):
            return date_str
        if not isinstance(date_str, str):
            return date_str
        cached = _date_cache.get(date_str)
        if cached is not None:
            return cached
        if _ISO_DATE_RE.match(date_str):
            try:
                dt = datetime.strptime(date_str, '%Y-%m-%d')
                _date_cache[date_str] = dt
                return dt
            except Exception:
                pass
        _date_cache[date_str] = date_str  # type: ignore[assignment]
        return date_str

    # Compteur de lignes par feuille pour éviter sheet.max_row (O(n) par appel)
    _sheet_row_counts = {}

    def _write_row_to_sheet(sheet, r, row_num=None, hyperlink_sheet_name=None):
        """Écrire une ligne de données dans une feuille.
        
        Args:
            sheet: La feuille Excel
            r: Le dictionnaire de données
            row_num: Numéro de ligne (optionnel)
            hyperlink_sheet_name: Nom de la feuille vers laquelle créer un lien hypertexte dans source_pdf
        """
        code_donneur = r.get("code_donneur_dordre") or None
        donneur = r.get("donneur_dordre") or None
        if not donneur and "institution_name" in r:
            donneur = r.get("institution_name")
        beneficiaire = r.get("beneficiaire") or None
        correspondant = r.get("correspondant") or None
        commentaires = r.get("commentaires") or None
        # Sanitize: ensure no tuples/lists leak into Excel cells
        if isinstance(code_donneur, (list, tuple)): code_donneur = code_donneur[0] if code_donneur else None
        if isinstance(donneur, (list, tuple)): donneur = donneur[0] if donneur else None
        if isinstance(beneficiaire, (list, tuple)): beneficiaire = beneficiaire[0] if beneficiaire else None
        if isinstance(correspondant, (list, tuple)): correspondant = correspondant[0] if correspondant else None
        if isinstance(commentaires, (list, tuple)): commentaires = commentaires[0] if commentaires else None
        date_ref = _convert_date_to_excel(r.get("date_reference"))
        reference_origine = r.get("reference_origine") or None
        source_pdf = r.get("source_pdf") or ""
        
        row_data = [
            correspondant,
            date_ref,
            r.get("reference"),
            reference_origine,
            r.get("type_MT"),
            r.get("pays_iso3"),
            code_donneur,
            donneur,
            beneficiaire,
            r.get("montant"),
            r.get("devise"),
            commentaires,
            source_pdf
        ]
        # NB: Les champs (code_donneur, donneur, beneficiaire, correspondant,
        # commentaires) sont déjà désérialisés ci-dessus si liste/tuple. La
        # sanitization globale ci-dessous a été retirée pour éviter ~5M
        # appels isinstance() sur les gros volumes (gain ~3-4s).
        sheet.append(row_data)
        
        # Utiliser un compteur local au lieu de sheet.max_row (O(n) par appel → O(n²))
        sheet_id = id(sheet)
        if sheet_id not in _sheet_row_counts:
            _sheet_row_counts[sheet_id] = 2  # Row 1 = header, first data row = 2
        else:
            _sheet_row_counts[sheet_id] += 1
        current_row = _sheet_row_counts[sheet_id]
        
        # Appliquer le format date à la colonne date_reference (colonne B = 2)
        if date_ref and isinstance(date_ref, datetime):
            sheet.cell(row=current_row, column=2).number_format = 'DD/MM/YYYY'
        
        # Ajouter un lien hypertexte vers la feuille détaillée du message (colonne M = 13)
        if hyperlink_sheet_name and source_pdf:
            try:
                # Échapper le nom de la feuille pour gérer les caractères spéciaux
                escaped_sheet_name = hyperlink_sheet_name.replace("'", "''")
                cell = sheet.cell(row=current_row, column=13)
                cell.hyperlink = f"#'{escaped_sheet_name}'!A1"
                cell.style = "Hyperlink"
            except Exception:
                pass  # Si le lien échoue, on garde juste le texte

    # FONCTIONNALITÉ: Détection des doublons potentiels (910, 103, 202 — tous types confondus)
    # ÉTAPE 1: Détecter les doublons AVANT d'écrire dans summary
    # Combiner toutes les rows pour la détection
    all_rows_for_duplicates = list(rows) + (beaccmcx091_rows or []) + (exception_323201_rows or []) + (other_exceptions_rows or []) + (banque_de_france_rows or []) + (forex_rows or []) + (bdf_corr_exception_rows or [])
    
    # Créer des groupes par (référence, montant) pour détecter les doublons
    potential_duplicates = []
    seen_keys = {}
    rows_to_exclude_from_summary = set()  # IDs des messages à exclure de summary
    rows_to_mark_as_duplicate = set()  # IDs des messages à marquer "potentiel doublon"
    
    # Utiliser id() pour identifier de manière unique chaque dictionnaire
    for r in all_rows_for_duplicates:
        mt_type = r.get("type_MT") or ""
        montant = r.get("montant")
        
        # Pour les MT910, utiliser F20 comme référence (déjà fait dans l'extraction)
        reference = r.get("reference")
        row_direction = r.get("direction", "")
        
        if reference and montant is not None:
            # Normaliser la clé — inclure direction pour ne pas considérer
            # un entrant et un sortant de même ref/montant comme doublons
            key = (str(reference).strip().upper(), float(montant) if montant else 0, row_direction)
            
            if key in seen_keys:
                # Doublon potentiel trouvé entre deux messages de types quelconques
                prev_row = seen_keys[key]
                
                # Ajouter les deux dans la liste des doublons potentiels
                if prev_row not in potential_duplicates:
                    potential_duplicates.append(prev_row)
                if r not in potential_duplicates:
                    potential_duplicates.append(r)
                
                # Exclure le second du summary, marquer le premier
                # On garde le premier message vu (prev_row) et on exclut le second (r)
                rows_to_exclude_from_summary.add(id(r))
                rows_to_mark_as_duplicate.add(id(prev_row))
            else:
                seen_keys[key] = r
    
    # ÉTAPE 2: Modifier les commentaires des messages à marquer
    for r in all_rows_for_duplicates:
        if id(r) in rows_to_mark_as_duplicate:
            current_comment = r.get("commentaires") or ""
            if current_comment:
                r["commentaires"] = f"{current_comment} / potentiel doublon"
            else:
                r["commentaires"] = "potentiel doublon"
    
    # ÉTAPE 2.5: Pré-calculer les noms de feuilles pour les liens hypertextes
    # et détecter BEAC dans le donneur d'ordre pour assigner le pays
    row_id_to_sheet_name = {}
    
    # D'abord, collecter tous les pays pour réserver leurs noms
    # et détecter BEAC dans le donneur d'ordre pour assigner le pays si manquant
    for r in rows:
        donneur = r.get("donneur_dordre") or r.get("institution_name") or ""
        if isinstance(donneur, (list, tuple)):
            donneur = donneur[0] if donneur else ""
        if not isinstance(donneur, str):
            donneur = str(donneur) if donneur else ""
        if donneur and "BEAC" in donneur.upper():
            if not r.get("pays_iso3"):
                r["pays_iso3"] = "BEAC"
    
    # Collecter les pays uniques (pour réserver leurs noms de feuilles)
    all_countries = set()
    for r in rows:
        if id(r) not in rows_to_exclude_from_summary:
            country = r.get("pays_iso3")
            if country:
                # BEAC sera renommé en Operations_BEAC
                all_countries.add("Operations_BEAC" if country == "BEAC" else country)
    
    # Skip per-row detail sheets for Excel extraction mode (too many rows)
    # Skip per-row detail sheets pour les modes à grande volumétrie
    # OU pour tout volume élevé (safety-net : créer 30k+ feuilles fait O(n²)
    # avec wb.create_sheet et Excel ne gère pas plus de ~10k onglets en pratique).
    skip_per_row_sheets = (
        direction.startswith("excel_extraction")
        or direction.startswith("eastnet_")
        or len(rows) > 500
    )

    if not skip_per_row_sheets:
        # Noms de feuilles réservés (déjà utilisés par d'autres onglets)
        reserved_names = {
            "summary", "BEACCMCX091", "Exceptions_323201", 
            "Autres_Exceptions", "Doublons_potentiels"
        }
        reserved_names.update(all_countries)
        
        # Calculer les noms de feuilles en évitant les noms réservés
        used_names_precompute = set(reserved_names)
        for r in rows:
            # Pré-calculer le nom de la feuille
            base = r.get("source_pdf", "sheet")
            title = _sanitize_sheet_title(str(base))
            original = title
            i = 1
            while title in used_names_precompute:
                suffix = f"_{i}"
                max_base_len = 31 - len(suffix)
                title = (original[:max_base_len] + suffix) if len(original) > max_base_len else (original + suffix)
                i += 1
            used_names_precompute.add(title)
            row_id_to_sheet_name[id(r)] = title
    
    # ÉTAPE 3: Écrire les rows dans summary (en excluant les doublons) avec liens hypertextes
    for r in rows:
        if id(r) not in rows_to_exclude_from_summary:
            sheet_name = row_id_to_sheet_name.get(id(r))
            _write_row_to_sheet(summary, r, hyperlink_sheet_name=sheet_name)

    # Add BEACCMCX091 sheet if there are any (right after main summary, before country summaries)
    if beaccmcx091_rows and len(beaccmcx091_rows) > 0:
        beac_sheet = wb.create_sheet(title="BEACCMCX091", index=1)
        beac_sheet.append(display_headers)

        # Note : on n'applique pas `rows_to_exclude_from_summary` ici car les
        # "doublons" détectés par clé (F20, montant, direction) correspondent
        # le plus souvent à des relations légitimes entre messages SWIFT liés
        # (MT103/MT202/MT910 partageant la même F20). Les exclure ici viderait
        # la feuille d'exception qui doit refléter l'intégralité des messages
        # ayant déclenché la règle.
        for r in beaccmcx091_rows:
            _write_row_to_sheet(beac_sheet, r)
        
        # Adjust column widths for BEACCMCX091 sheet
        try:
            for col_idx in range(1, len(display_headers) + 1):
                max_len = max(
                    (len(str(cell.value)) for cell in beac_sheet[get_column_letter(col_idx)] if cell.value is not None),
                    default=10
                )
                beac_sheet.column_dimensions[get_column_letter(col_idx)].width = min(60, max(12, max_len + 2))
        except Exception:
            pass

    # Add Exceptions_323201 sheet if there are any
    if exception_323201_rows and len(exception_323201_rows) > 0:
        exc_sheet = wb.create_sheet(title="Exceptions_323201", index=2 if beaccmcx091_rows else 1)
        exc_sheet.append(display_headers)

        # Pas de filtre rows_to_exclude_from_summary : voir commentaire feuille BEACCMCX091.
        for r in exception_323201_rows:
            _write_row_to_sheet(exc_sheet, r)
        
        # Adjust column widths
        try:
            for col_idx in range(1, len(display_headers) + 1):
                max_len = max(
                    (len(str(cell.value)) for cell in exc_sheet[get_column_letter(col_idx)] if cell.value is not None),
                    default=10
                )
                exc_sheet.column_dimensions[get_column_letter(col_idx)].width = min(60, max(12, max_len + 2))
        except Exception:
            pass

    # Add Autres_Exceptions sheet if there are any (EUR/nivellement exceptions)
    sheet_index = 1
    if beaccmcx091_rows:
        sheet_index += 1
    if exception_323201_rows:
        sheet_index += 1
    
    if other_exceptions_rows and len(other_exceptions_rows) > 0:
        other_exc_sheet = wb.create_sheet(title="Autres_Exceptions", index=sheet_index)
        other_exc_sheet.append(display_headers)

        # Pas de filtre rows_to_exclude_from_summary : voir commentaire feuille BEACCMCX091.
        for r in other_exceptions_rows:
            _write_row_to_sheet(other_exc_sheet, r)
        
        # Adjust column widths
        try:
            for col_idx in range(1, len(display_headers) + 1):
                max_len = max(
                    (len(str(cell.value)) for cell in other_exc_sheet[get_column_letter(col_idx)] if cell.value is not None),
                    default=10
                )
                other_exc_sheet.column_dimensions[get_column_letter(col_idx)].width = min(60, max(12, max_len + 2))
        except Exception:
            pass
    
    # ÉTAPE 4: Créer la feuille BANQUE DE FRANCE (MT103 USD avec BANQUE DE FRANCE / FW021083459)
    if banque_de_france_rows is None:
        banque_de_france_rows = []
    if banque_de_france_rows:
        sheet_index = 1
        if beaccmcx091_rows:
            sheet_index += 1
        if exception_323201_rows:
            sheet_index += 1
        if other_exceptions_rows:
            sheet_index += 1
        
        bdf_sheet = wb.create_sheet(title="BANQUE DE FRANCE", index=sheet_index)
        bdf_sheet.append(display_headers)
        
        for r in banque_de_france_rows:
            _write_row_to_sheet(bdf_sheet, r)
        
        # Adjust column widths
        try:
            for col_idx in range(1, len(display_headers) + 1):
                max_len = max(
                    (len(str(cell.value)) for cell in bdf_sheet[get_column_letter(col_idx)] if cell.value is not None),
                    default=10
                )
                bdf_sheet.column_dimensions[get_column_letter(col_idx)].width = min(60, max(12, max_len + 2))
        except Exception:
            pass
    
    # ÉTAPE 5: Créer la feuille forex (MT910 entrants avec code donneur dans la feuille forex)
    if forex_rows is None:
        forex_rows = []
    if forex_rows:
        sheet_index = 1
        if beaccmcx091_rows:
            sheet_index += 1
        if exception_323201_rows:
            sheet_index += 1
        if other_exceptions_rows:
            sheet_index += 1
        if banque_de_france_rows:
            sheet_index += 1
        
        forex_sheet = wb.create_sheet(title="forex", index=sheet_index)
        forex_sheet.append(display_headers)
        
        for r in forex_rows:
            _write_row_to_sheet(forex_sheet, r)
        
        # Adjust column widths
        try:
            for col_idx in range(1, len(display_headers) + 1):
                max_len = max(
                    (len(str(cell.value)) for cell in forex_sheet[get_column_letter(col_idx)] if cell.value is not None),
                    default=10
                )
                forex_sheet.column_dimensions[get_column_letter(col_idx)].width = min(60, max(12, max_len + 2))
        except Exception:
            pass
    
    # ÉTAPE 5bis: Créer la feuille Exceptions_Correspondants (MT202 sortants avec exceptions BdF)
    if bdf_corr_exception_rows is None:
        bdf_corr_exception_rows = []
    if bdf_corr_exception_rows:
        sheet_index = 1
        if beaccmcx091_rows:
            sheet_index += 1
        if exception_323201_rows:
            sheet_index += 1
        if other_exceptions_rows:
            sheet_index += 1
        if banque_de_france_rows:
            sheet_index += 1
        if forex_rows:
            sheet_index += 1

        bdf_exc_sheet = wb.create_sheet(title="Exceptions_Correspondants", index=sheet_index)
        bdf_exc_sheet.append(display_headers)

        for r in bdf_corr_exception_rows:
            _write_row_to_sheet(bdf_exc_sheet, r)

        # Adjust column widths
        try:
            for col_idx in range(1, len(display_headers) + 1):
                max_len = max(
                    (len(str(cell.value)) for cell in bdf_exc_sheet[get_column_letter(col_idx)] if cell.value is not None),
                    default=10
                )
                bdf_exc_sheet.column_dimensions[get_column_letter(col_idx)].width = min(60, max(12, max_len + 2))
        except Exception:
            pass

    # ÉTAPE 6: Créer la feuille Doublons_potentiels (contient TOUS les doublons détectés)
    if potential_duplicates:
        sheet_index = 1
        if beaccmcx091_rows:
            sheet_index += 1
        if exception_323201_rows:
            sheet_index += 1
        if other_exceptions_rows:
            sheet_index += 1
        if banque_de_france_rows:
            sheet_index += 1
        if forex_rows:
            sheet_index += 1
        if bdf_corr_exception_rows:
            sheet_index += 1
        
        dup_sheet = wb.create_sheet(title="Doublons_potentiels", index=sheet_index)
        dup_sheet.append(display_headers)
        
        # Écrire TOUS les doublons (pas de filtrage ici)
        for r in potential_duplicates:
            _write_row_to_sheet(dup_sheet, r)
        
        # Adjust column widths
        try:
            for col_idx in range(1, len(display_headers) + 1):
                max_len = max(
                    (len(str(cell.value)) for cell in dup_sheet[get_column_letter(col_idx)] if cell.value is not None),
                    default=10
                )
                dup_sheet.column_dimensions[get_column_letter(col_idx)].width = min(60, max(12, max_len + 2))
        except Exception:
            pass

    # Add per-country summary sheets (after special sheets, before per-file sheets)
    # Inclut maintenant BEAC comme "pays" pour les opérations de la BEAC
    countries = {}
    for r in rows:
        # Exclure les doublons aussi des feuilles par pays
        if id(r) not in rows_to_exclude_from_summary:
            country = r.get("pays_iso3")
            if country:
                if country not in countries:
                    countries[country] = []
                countries[country].append(r)
    
    # Create a sheet for each country with summary data
    # BEAC sera créé comme une feuille "Operations_BEAC" si présent
    for country_code in sorted(countries.keys()):
        try:
            country_rows = countries[country_code]
            # Trier par correspondant (puis donneur_dordre) pour regrouper
            country_rows.sort(key=lambda r: (str(r.get("correspondant") or "").upper(), str(r.get("donneur_dordre") or "").upper()))
            # Renommer la feuille BEAC en "Operations_BEAC" pour plus de clarté
            sheet_title = "Operations_BEAC" if country_code == "BEAC" else country_code
            country_sheet = wb.create_sheet(title=sheet_title)
            
            # Same headers as summary
            country_sheet.append(display_headers)
            
            # Add rows for this country with hyperlinks
            for r in country_rows:
                sheet_name = row_id_to_sheet_name.get(id(r))
                _write_row_to_sheet(country_sheet, r, hyperlink_sheet_name=sheet_name)
            
            # Adjust column widths
            try:
                for col_idx in range(1, len(display_headers) + 1):
                    max_len = max(
                        (len(str(cell.value)) for cell in country_sheet[get_column_letter(col_idx)] if cell.value is not None),
                        default=10
                    )
                    country_sheet.column_dimensions[get_column_letter(col_idx)].width = min(60, max(12, max_len + 2))
            except Exception:
                pass
        except Exception as e:
            logger.debug("Failed to create country sheet for %s: %s", country_code, e)

    # create per-file sheets (key/value) - AFTER country summaries
    # Skip for Excel extraction mode (rows are individual transactions, not SWIFT messages)
    if not skip_per_row_sheets:
        # Utiliser les noms pré-calculés pour garantir la cohérence avec les hyperliens
        for r in rows:
            title = row_id_to_sheet_name.get(id(r))
            if not title:
                # Fallback si pas de nom pré-calculé (ne devrait pas arriver)
                base = r.get("source_pdf", "sheet")
                title = _sanitize_sheet_title(str(base))
            
            # Vérifier que le nom n'existe pas déjà dans le workbook (sécurité)
            if title in wb.sheetnames:
                # Ajouter un suffixe unique
                i = 1
                original = title
                while title in wb.sheetnames:
                    suffix = f"_{i}"
                    max_base_len = 31 - len(suffix)
                    title = (original[:max_base_len] + suffix) if len(original) > max_base_len else (original + suffix)
                    i += 1
            ws = wb.create_sheet(title=title)

            # Lien retour vers la feuille summary (première ligne)
            ws.append(["⬅ Retour au summary", ""])
            try:
                back_cell = ws.cell(row=1, column=1)
                back_cell.hyperlink = "#summary!A1"
                back_cell.style = "Hyperlink"
            except Exception:
                pass
            ws.append([])  # Ligne vide de séparation

            ordered_keys = [
                "correspondant", "date_reference", "reference", "reference_origine", "type_MT", "pays_iso3",
                "code_donneur_dordre", "donneur_dordre", "institution_name", "beneficiaire", 
                "montant", "devise", "commentaires", "source_pdf"
            ]
            written = set()
            for k in ordered_keys:
                if k in r:
                    label = "Code du donneur d'ordre" if k == "code_donneur_dordre" else ("donneur d'ordre" if k in ("donneur_dordre", "institution_name") else ("Bénéficiaire" if k == "beneficiaire" else ("correspondant" if k == "correspondant" else ("commentaires" if k == "commentaires" else k))))
                    ws.append([label, r.get(k)])
                    written.add(k)
            for k, v in r.items():
                if k in written:
                    continue
                label = "Code du donneur d'ordre" if k == "code_donneur_dordre" else ("donneur d'ordre" if k in ("donneur_dordre", "institution_name") else ("Bénéficiaire" if k == "beneficiaire" else ("correspondant" if k == "correspondant" else ("commentaires" if k == "commentaires" else k))))
                ws.append([label, v])

            # adjust column widths heuristically
            try:
                max_len_col1 = max((len(str(row[0])) for row in ws.values if row[0] is not None), default=10)
                max_len_col2 = max((len(str(row[1])) for row in ws.values if len(row) > 1 and row[1] is not None), default=10)
                ws.column_dimensions[get_column_letter(1)].width = min(60, max(12, max_len_col1 + 2))
                ws.column_dimensions[get_column_letter(2)].width = min(80, max(12, max_len_col2 + 8))
            except Exception:
                pass

    # Ajouter un lien retour vers summary dans toutes les feuilles (sauf summary elle-même)
    # ATTENTION: ws.insert_rows(1, 1) est O(n) en cellules → catastrophique pour les
    # gros volumes (déplacement de 30k+ cellules par feuille). Pour les flux à
    # grande volumétrie (EastNet / Excel extraction), on saute cette étape :
    # les onglets restent navigables via la barre d'onglets d'Excel.
    if not skip_per_row_sheets:
        for sheet_name in wb.sheetnames:
            if sheet_name == "summary":
                continue
            ws = wb[sheet_name]
            # Vérifier si la feuille a déjà un lien retour (les per-file sheets l'ont déjà)
            first_val = ws.cell(row=1, column=1).value
            if first_val and "Retour" in str(first_val):
                continue  # Déjà ajouté
            # Pour les feuilles tabulaires (pays, exceptions, doublons) : insérer en première ligne avant les headers
            ws.insert_rows(1, 1)
            back_cell = ws.cell(row=1, column=1)
            back_cell.value = "⬅ Retour au summary"
            try:
                back_cell.hyperlink = "#summary!A1"
                back_cell.style = "Hyperlink"
            except Exception:
                pass

    # ────────────────────────────────────────────────────────────────────
    # Feuille "Sortants_Rejetes_NtStatus" — sortants RJE écartés au filtre
    # Nt.Status (mode 2 RJE pour CITI USD / BDF). Insérée en dernière position.
    # ────────────────────────────────────────────────────────────────────
    if nt_status_rejected_rows:
        rej_sheet = wb.create_sheet(title="Sortants_Rejetes_NtStatus")
        rej_headers = [
            "type_MT",
            "reference",
            "date_reference",
            "devise",
            "montant",
            "Code du donneur d'ordre",
            "donneur d'ordre",
            "Bénéficiaire",
            "pays_iso3",
            "correspondant",
            "Nt.Status (CSV)",
            "Status (CSV)",
            "Identifier (CSV)",
            "raison_rejet",
            "source_pdf",
        ]
        # Ligne 1 : lien retour ; ligne 2 : headers
        back_cell = rej_sheet.cell(row=1, column=1)
        back_cell.value = "⬅ Retour au summary"
        try:
            back_cell.hyperlink = "#summary!A1"
            back_cell.style = "Hyperlink"
        except Exception:
            pass
        rej_sheet.append(rej_headers)
        for r in nt_status_rejected_rows:
            rej_sheet.append([
                r.get("type_MT"),
                r.get("reference"),
                r.get("date_reference"),
                r.get("devise"),
                r.get("montant"),
                r.get("code_donneur_dordre"),
                r.get("donneur_dordre") or r.get("institution_name"),
                r.get("beneficiaire"),
                r.get("pays_iso3"),
                r.get("correspondant"),
                r.get("_nt_status"),
                r.get("_csv_status"),
                r.get("_csv_identifier"),
                r.get("_rejet_raison"),
                r.get("source_pdf"),
            ])
        try:
            for col_idx in range(1, len(rej_headers) + 1):
                max_len = max(
                    (len(str(cell.value)) for cell in rej_sheet[get_column_letter(col_idx)] if cell.value is not None),
                    default=10,
                )
                rej_sheet.column_dimensions[get_column_letter(col_idx)].width = min(60, max(12, max_len + 2))
        except Exception:
            pass

    # Single final save after all sheets are created
    wb.save(out_path)
    logger.info("Workbook created with country sheets: %s", out_path)
    return out_path


def extract_mt900_only(pdf_path: Path, bic_xlsx: Optional[str] = None) -> tuple[List[Dict], Dict[str, set]]:
    """
    Extraire uniquement les MT900 d'un fichier PDF.
    
    Args:
        pdf_path: Chemin vers le fichier PDF
        bic_xlsx: Chemin optionnel vers le fichier Excel BIC
        
    Returns:
        tuple: (mt900_rows, missing_codes)
    """
    p = Path(pdf_path)
    
    if HAS_MT_MULTI and mt_multi_module:
        try:
            return mt_multi_module.extract_mt900_only(p, bic_xlsx=bic_xlsx)
        except Exception as e:
            logger.exception("extract_mt900_only: failed for %s: %s", p.name, e)
    
    return [], {"unmapped": set(), "empty": set()}


def match_mt900_with_transfers(mt900_rows: List[Dict], transfer_rows: List[Dict]) -> tuple[List[Dict], List[Dict], List[Dict], List[Dict]]:
    """
    Matcher les MT900 avec les MT103/MT202 via la référence d'origine (F21).
    Applique également les règles d'exception pour MT900.
    
    Logique:
    - Vérifier d'abord si le MT900 doit être en exception (T2PL, NIVLT)
    - Pour chaque MT900, chercher le MT103/MT202 dont F20 (référence) = F21 du MT900 (related_reference)
    - Compléter les infos du MT900 (donneur d'ordre, pays) depuis le MT103/MT202 matché
    - Les MT103/MT202 sans correspondance vont dans suspens
    - Les MT900 sans correspondance vont dans unmatched_mt900_rows
    
    Args:
        mt900_rows: Liste des MT900 extraits
        transfer_rows: Liste des MT103/MT202 extraits
        
    Returns:
        tuple: (matched_mt900_rows, suspens_rows, exception_mt900_rows, unmatched_mt900_rows)
    """
    # Fonction pour vérifier les exceptions MT900
    def _check_mt900_exception(row: Dict) -> Optional[str]:
        """
        Vérifier si le MT900 doit être mis en exception.
        """
        if not row or not row.get("type_MT"):
            return None
        
        if not row.get("type_MT").startswith("fin.900"):
            return None
        
        # Vérifier F20 (référence)
        reference = row.get("reference") or ""
        reference_upper = reference.upper()
        
        if "T2PL" in reference_upper:
            return "T2PL"
        
        # Vérifier F21 (référence d'origine)
        related_reference = row.get("related_reference") or ""
        related_ref_upper = related_reference.upper()
        
        if "NIVLT" in related_ref_upper or "NIVELLEMENT" in related_ref_upper:
            return "nivellement"

        # Souveraineté (BDF) : F21 contient "SOUV" (typiquement SOUVxxxxxxxxxxx)
        if "SOUV" in related_ref_upper:
            return "souverainete"
        
        return None
    
    # Créer un index des MT103/MT202 par référence pour le matching rapide
    # Index "strict" : référence telle quelle (uppercase, trimmed)
    # Index "loose" : référence avec tous les non-alphanumériques retirés
    #   (utile quand un correspondant utilise '.' au lieu de '/' comme séparateur,
    #   ex: MT900 Citi rel_ref="00290565.251104" ↔ MT202 BEAC ref="00290565/251104")
    import re as _re
    def _loose_key(s: str) -> str:
        return _re.sub(r'[^A-Z0-9]', '', str(s).upper()) if s else ''

    ref_index: Dict[str, Dict] = {}
    ref_index_loose: Dict[str, Dict] = {}
    for row in transfer_rows:
        ref = row.get("reference")
        if ref:
            ref_key = str(ref).strip().upper()
            ref_index[ref_key] = row
            lk = _loose_key(ref)
            if lk:
                ref_index_loose[lk] = row
    
    # Matcher les MT900 avec les MT103/MT202
    matched_mt900_rows: List[Dict] = []  # MT900 avec correspondant trouvé
    unmatched_mt900_rows: List[Dict] = []  # MT900 sans correspondant
    exception_mt900_rows: List[Dict] = []
    matched_refs: set = set()
    
    for mt900_row in mt900_rows:
        # Vérifier d'abord si le MT900 doit être mis en exception
        exception_comment = _check_mt900_exception(mt900_row)
        if exception_comment:
            mt900_row["commentaires"] = exception_comment
            exception_mt900_rows.append(mt900_row)
            continue  # Ne pas matcher ce message, il va directement en exception
        
        # La référence d'origine F21 est utilisée pour matcher
        related_ref = mt900_row.get("related_reference")
        
        has_match = False
        if related_ref:
            ref_key = str(related_ref).strip().upper()
            matched_transfer = ref_index.get(ref_key)
            matched_key_used = ref_key
            # Fallback : matching "loose" sur version alphanumérique-only
            #   ex: "00290565.251104" ↔ "00290565/251104"
            if matched_transfer is None:
                lk = _loose_key(related_ref)
                if lk and lk in ref_index_loose:
                    matched_transfer = ref_index_loose[lk]
                    matched_key_used = (matched_transfer.get("reference") or "").strip().upper()

            if matched_transfer is not None:
                
                # Copier les infos manquantes du MT103/MT202 vers le MT900
                fields_to_copy = [
                    "code_donneur_dordre", "donneur_dordre", "beneficiaire", 
                    "pays_iso3"
                ]
                for field in fields_to_copy:
                    if not mt900_row.get(field) and matched_transfer.get(field):
                        mt900_row[field] = matched_transfer[field]
                
                # Forcer le correspondant du MT202/MT103 (receiver BIC) sur le MT900 matché
                if matched_transfer.get("correspondant"):
                    mt900_row["correspondant"] = matched_transfer["correspondant"]
                
                # Ajouter info de matching
                mt900_row["matched_with"] = matched_transfer.get("reference")
                mt900_row["matched_type"] = matched_transfer.get("type_MT")
                # Ajouter la source du message matché pour la traçabilité
                mt900_row["matched_source"] = matched_transfer.get("source_pdf")

                # Hériter de la catégorie du transfert matché afin que la
                # feuille de destination (main / BEACCMCX091 / etc.) soit correcte.
                if matched_transfer.get("_category"):
                    mt900_row["_category"] = matched_transfer["_category"]
                
                matched_refs.add(matched_key_used)
                has_match = True
        
        if has_match:
            matched_mt900_rows.append(mt900_row)
        else:
            # MT900 sans correspondant trouvé - utiliser sender_bic comme correspondant
            if not mt900_row.get("correspondant") and mt900_row.get("sender_bic"):
                mt900_row["correspondant"] = mt900_row["sender_bic"]
            mt900_row["commentaires"] = "pas de correspondant MT103/MT202"
            unmatched_mt900_rows.append(mt900_row)
    
    # Les MT103/MT202 non matchés vont dans suspens
    suspens_rows: List[Dict] = []
    for row in transfer_rows:
        ref = row.get("reference")
        if ref:
            ref_key = str(ref).strip().upper()
            if ref_key not in matched_refs:
                row["status"] = "suspens - pas de confirmation MT900"
                suspens_rows.append(row)
        else:
            # Pas de référence = suspens
            row["status"] = "suspens - référence manquante"
            suspens_rows.append(row)
    
    logger.info("match_mt900_with_transfers: %d matched, %d unmatched MT900, %d suspens, %d exceptions", 
                len(matched_mt900_rows), len(unmatched_mt900_rows), len(suspens_rows), len(exception_mt900_rows))
    
    return matched_mt900_rows, suspens_rows, exception_mt900_rows, unmatched_mt900_rows


def extract_transfer_analysis_dispatch(pdf_path: Path) -> tuple[List[Dict], List[Dict], Dict[str, set]]:
    """
    Dispatch pour l'analyse des transferts sortants exécutés.
    Utilise mt_multi.extract_transfer_analysis.
    
    Returns:
        tuple: (matched_900_rows, suspens_rows, missing_codes)
    """
    p = Path(pdf_path)
    
    if HAS_MT_MULTI and mt_multi_module:
        try:
            return mt_multi_module.extract_transfer_analysis(p)
        except Exception as e:
            logger.exception("extract_transfer_analysis_dispatch: failed for %s: %s", p.name, e)
    
    # Fallback: retourner des listes vides
    return [], [], {"unmapped": set(), "empty": set()}


def create_transfer_analysis_workbook(matched_rows: List[Dict], suspens_rows: List[Dict], exception_rows: List[Dict], out_dir: Path, date_start: str = None, date_end: str = None, unmatched_mt900_rows: List[Dict] = None, xlsx_path: Optional[str] = None, duplicate_mt900_rows: List[Dict] = None, nt_status_rejected_rows: Optional[List[Dict]] = None) -> Path:
    """
    Créer un workbook Excel pour l'analyse des transferts sortants exécutés.

    IMPLEMENTATION : mode write_only pour streamer les lignes directement vers
    le XML zippé. Sur 35k lignes, gain mesuré ×16 vs mode normal (7s vs 127s).

    Sheets:
    - "Transferts_Executes": TOUS les MT900 du flux principal (matchés + non rapprochés),
      enrichis quand possible depuis bic_codes.xlsx via les 4 premiers caractères de
      related_reference (colonne Reglement) -> code BIC, nom institution, pays.
    - "Transferts_Executes_Matches": MT900 main rapprochés avec leur MT202/MT103 d'origine.
    - "MT900_Doublons": MT900 retransmis par le réseau SWIFT FIN.
    - "MT900_non_rapproches": MT900 sans correspondant MT103/MT202.
    - "Suspens": MT202/MT103 sans confirmation MT900.
    - Feuilles d'exception (BEACCMCX091, etc.).
    - "Sortants_Rejetes_NtStatus": sortants RJE écartés par filtre Nt.Status.
    """
    if unmatched_mt900_rows is None:
        unmatched_mt900_rows = []
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    date_suffix = ""
    if date_start and date_end:
        date_suffix = f"_{date_start}_to_{date_end}"
    elif date_start:
        date_suffix = f"_from_{date_start}"
    elif date_end:
        date_suffix = f"_to_{date_end}"

    out_path = out_dir / f"analyse_transferts{date_suffix}_{ts}.xlsx"

    wb = Workbook(write_only=True)

    # ── Helpers communs ────────────────────────────────────────────────
    _DATE_FMT = 'DD/MM/YYYY'

    def _convert_date_to_excel(date_str):
        if not date_str:
            return None
        try:
            if isinstance(date_str, str) and re.match(r'^\d{4}-\d{2}-\d{2}$', date_str):
                return datetime.strptime(date_str, '%Y-%m-%d')
            return date_str
        except Exception:
            return date_str

    def _dc(ws, value):
        """Renvoie un WriteOnlyCell formaté date si datetime, sinon la valeur brute."""
        if value is not None and isinstance(value, datetime):
            c = WriteOnlyCell(ws, value=value)
            c.number_format = _DATE_FMT
            return c
        return value

    def _flush_sheet(ws, headers, data_rows, max_w: int = 60, sample_limit: int = 300):
        """Écrit headers + lignes en streaming et calcule les largeurs sur les
        `sample_limit` premières lignes (header inclus). Compatible write_only.

        `data_rows` est un itérable de listes de valeurs (datetime accepté).
        """
        n_cols = len(headers)
        max_lens = [len(str(h)) for h in headers]
        ws.append(list(headers))
        for idx, row in enumerate(data_rows):
            if idx < sample_limit:
                for i, v in enumerate(row):
                    if v is None or i >= n_cols:
                        continue
                    # WriteOnlyCell : récupérer la valeur sous-jacente
                    raw = getattr(v, "value", v)
                    if raw is None:
                        continue
                    L = len(str(raw))
                    if L > max_lens[i]:
                        max_lens[i] = L
            ws.append(row)
        for i, L in enumerate(max_lens, start=1):
            ws.column_dimensions[get_column_letter(i)].width = min(max_w, max(12, L + 2))

    # Mapping catégorie → titre de feuille
    _CATEGORY_TITLES = {
        "main": "Transferts_Executes",
        "beaccmcx091": "BEACCMCX091",
        "exception_323201": "Exceptions_323201",
        "other_exception": "Autres_Exceptions",
        "banque_de_france": "BANQUE DE FRANCE",
        "forex": "forex",
        "bdf_corr_exception": "Exceptions_Correspondants",
    }

    # Dispatcher matched & suspens par catégorie
    matched_by_cat: Dict[str, List[Dict]] = {}
    suspens_by_cat: Dict[str, List[Dict]] = {}
    for r in matched_rows:
        cat = r.get("_category") or "main"
        matched_by_cat.setdefault(cat, []).append(r)
    for r in suspens_rows:
        cat = r.get("_category") or "main"
        suspens_by_cat.setdefault(cat, []).append(r)

    # Headers
    transfer_headers = [
        "type_MT", "reference", "related_reference", "date_reference",
        "devise", "montant", "Code du donneur d'ordre", "donneur d'ordre",
        "Bénéficiaire", "pays_iso3", "correspondant", "Message correspondant",
        "source_pdf",
    ]
    suspens_headers = [
        "type_MT", "reference", "date_reference", "devise", "montant",
        "Code du donneur d'ordre", "donneur d'ordre", "Bénéficiaire",
        "pays_iso3", "correspondant", "status", "source_pdf",
    ]
    exc_combined_headers = [
        "type_MT", "reference", "related_reference", "date_reference",
        "devise", "montant", "Code du donneur d'ordre", "donneur d'ordre",
        "Bénéficiaire", "pays_iso3", "correspondant", "Statut MT900",
        "Message correspondant", "source_pdf",
    ]
    exec_headers_all = [
        "type_MT", "reference", "related_reference", "Code Reglement",
        "date_reference", "devise", "montant", "Code du donneur d'ordre",
        "donneur d'ordre", "Bénéficiaire", "pays_iso3", "correspondant",
        "source_pdf",
    ]

    def _build_message_correspondant(r: Dict) -> Optional[str]:
        matched_source = r.get("matched_source") or ""
        matched_type = r.get("matched_type") or ""
        if matched_source:
            return f"{matched_type} - {matched_source}" if matched_type else matched_source
        return None

    def _enrich_from_related_reference(r: Dict) -> Dict:
        rr = (r.get("related_reference") or "").strip()
        code_4 = rr[:4] if len(rr) >= 4 else None
        info = None
        if code_4 and code_4.isdigit() and HAS_BIC_UTILS and _bic_utils is not None:
            try:
                info = _bic_utils.map_reglement_code(code_4, xlsx_path=xlsx_path)
            except Exception:
                info = None
        return {
            "code_reglement": code_4,
            "code_bic": (info or {}).get("bic"),
            "nom": (info or {}).get("name"),
            "pays": (info or {}).get("country"),
        }

    # ─── Sheet 1 : Transferts_Executes ─────────────────────────────────
    exec_sheet = wb.create_sheet(title="Transferts_Executes")
    main_mt900_matched = [
        r for r in matched_by_cat.get("main", [])
        if (r.get("type_MT") or "").lower() == "fin.900"
    ]
    all_main_mt900 = list(main_mt900_matched) + list(unmatched_mt900_rows or [])

    def _exec_rows():
        for r in all_main_mt900:
            enr = _enrich_from_related_reference(r)
            d = _convert_date_to_excel(r.get("date_reference"))
            yield [
                r.get("type_MT"), r.get("reference"), r.get("related_reference"),
                enr["code_reglement"], _dc(exec_sheet, d), r.get("devise"),
                r.get("montant"), enr["code_bic"], enr["nom"],
                r.get("beneficiaire"), enr["pays"], r.get("correspondant"),
                r.get("source_pdf"),
            ]
    _flush_sheet(exec_sheet, exec_headers_all, _exec_rows())

    # ─── Sheet 2 : Transferts_Executes_Matches ─────────────────────────
    matches_sheet = wb.create_sheet(title="Transferts_Executes_Matches")

    def _matches_rows():
        for r in matched_by_cat.get("main", []):
            d = _convert_date_to_excel(r.get("date_reference"))
            ms = r.get("matched_source") or ""
            mt = r.get("matched_type") or ""
            msg_corr = (f"{mt} - {ms}" if mt else ms) if ms else None
            yield [
                r.get("type_MT"), r.get("reference"), r.get("related_reference"),
                _dc(matches_sheet, d), r.get("devise"), r.get("montant"),
                r.get("code_donneur_dordre"), r.get("donneur_dordre"),
                r.get("beneficiaire"), r.get("pays_iso3"), r.get("correspondant"),
                msg_corr, r.get("source_pdf"),
            ]
    _flush_sheet(matches_sheet, transfer_headers, _matches_rows())

    # ─── Sheet 3 : MT900_Doublons ──────────────────────────────────────
    if duplicate_mt900_rows:
        dup_headers = [
            "type_MT", "reference", "related_reference", "date_reference",
            "devise", "montant", "correspondant", "source_pdf",
            "reference_conservee", "source_conservee", "DLM", "cle_dedup",
        ]
        dup_sheet = wb.create_sheet(title="MT900_Doublons")

        def _dup_rows():
            for r in duplicate_mt900_rows:
                d = _convert_date_to_excel(r.get("date_reference"))
                yield [
                    r.get("type_MT"), r.get("reference"), r.get("related_reference"),
                    _dc(dup_sheet, d), r.get("devise"), r.get("montant"),
                    r.get("correspondant"), r.get("source_pdf"),
                    r.get("_duplicate_of_reference"), r.get("_duplicate_of_source"),
                    "OUI" if r.get("_is_dlm") else "NON", r.get("_dedupe_key"),
                ]
        _flush_sheet(dup_sheet, dup_headers, _dup_rows())

    # ─── Sheet 4 : MT900_non_rapproches ────────────────────────────────
    effective_unmatched = list(unmatched_mt900_rows or [])
    if effective_unmatched:
        unmatched_headers = [
            "type_MT", "reference", "related_reference", "date_reference",
            "devise", "montant", "correspondant", "Commentaires", "source_pdf",
        ]
        unmatched_sheet = wb.create_sheet(title="MT900_non_rapproches")

        def _unm_rows():
            for r in effective_unmatched:
                d = _convert_date_to_excel(r.get("date_reference"))
                yield [
                    r.get("type_MT"), r.get("reference"), r.get("related_reference"),
                    _dc(unmatched_sheet, d), r.get("devise"), r.get("montant"),
                    r.get("correspondant"), r.get("commentaires"), r.get("source_pdf"),
                ]
        _flush_sheet(unmatched_sheet, unmatched_headers, _unm_rows())

    # ─── Sheet 5 : Suspens ─────────────────────────────────────────────
    main_suspens = suspens_by_cat.get("main", [])
    if main_suspens:
        suspens_sheet = wb.create_sheet(title="Suspens")

        def _susp_rows():
            for r in main_suspens:
                d = _convert_date_to_excel(r.get("date_reference"))
                yield [
                    r.get("type_MT"), r.get("reference"), _dc(suspens_sheet, d),
                    r.get("devise"), r.get("montant"), r.get("code_donneur_dordre"),
                    r.get("donneur_dordre"), r.get("beneficiaire"), r.get("pays_iso3"),
                    r.get("correspondant"), r.get("status"), r.get("source_pdf"),
                ]
        _flush_sheet(suspens_sheet, suspens_headers, _susp_rows())

    # ─── Sheet 6 : Exceptions (MT900) ──────────────────────────────────
    if exception_rows:
        exception_headers = [
            "type_MT", "reference", "related_reference", "date_reference",
            "devise", "montant", "Code du donneur d'ordre", "donneur d'ordre",
            "Bénéficiaire", "pays_iso3", "correspondant", "Commentaires",
            "source_pdf",
        ]
        exception_sheet = wb.create_sheet(title="Exceptions")

        def _exc_rows():
            for r in exception_rows:
                d = _convert_date_to_excel(r.get("date_reference"))
                yield [
                    r.get("type_MT"), r.get("reference"), r.get("related_reference"),
                    _dc(exception_sheet, d), r.get("devise"), r.get("montant"),
                    r.get("code_donneur_dordre"), r.get("donneur_dordre"),
                    r.get("beneficiaire"), r.get("pays_iso3"), r.get("correspondant"),
                    r.get("commentaires"), r.get("source_pdf"),
                ]
        _flush_sheet(exception_sheet, exception_headers, _exc_rows())

    # ─── Feuilles par catégorie d'exception ────────────────────────────
    for cat_key, sheet_title in _CATEGORY_TITLES.items():
        if cat_key == "main":
            continue
        n_match = len(matched_by_cat.get(cat_key, []))
        n_susp = len(suspens_by_cat.get(cat_key, []))
        if n_match == 0 and n_susp == 0:
            continue
        ws_cat = wb.create_sheet(title=sheet_title)

        def _cat_rows(_cat=cat_key, _ws=ws_cat):
            for r in matched_by_cat.get(_cat, []):
                d = _convert_date_to_excel(r.get("date_reference"))
                yield [
                    r.get("type_MT"), r.get("reference"), r.get("related_reference"),
                    _dc(_ws, d), r.get("devise"), r.get("montant"),
                    r.get("code_donneur_dordre"), r.get("donneur_dordre"),
                    r.get("beneficiaire"), r.get("pays_iso3"), r.get("correspondant"),
                    "Exécuté (MT900 reçu)", _build_message_correspondant(r),
                    r.get("source_pdf"),
                ]
            for r in suspens_by_cat.get(_cat, []):
                d = _convert_date_to_excel(r.get("date_reference"))
                yield [
                    r.get("type_MT"), r.get("reference"), None,
                    _dc(_ws, d), r.get("devise"), r.get("montant"),
                    r.get("code_donneur_dordre"), r.get("donneur_dordre"),
                    r.get("beneficiaire"), r.get("pays_iso3"), r.get("correspondant"),
                    r.get("status") or "En attente MT900", None,
                    r.get("source_pdf"),
                ]
        _flush_sheet(ws_cat, exc_combined_headers, _cat_rows())
        logger.info("Sheet %s: %d exécutés + %d en attente", sheet_title, n_match, n_susp)

    # ─── Feuille Sortants_Rejetes_NtStatus ─────────────────────────────
    if nt_status_rejected_rows:
        rej_sheet = wb.create_sheet(title="Sortants_Rejetes_NtStatus")
        rej_headers = [
            "type_MT", "reference", "date_reference", "devise", "montant",
            "Code du donneur d'ordre", "donneur d'ordre", "Bénéficiaire",
            "pays_iso3", "correspondant", "Nt.Status (CSV)", "Status (CSV)",
            "Identifier (CSV)", "raison_rejet", "source_pdf",
        ]

        def _rej_rows():
            for r in nt_status_rejected_rows:
                d = r.get("date_reference")
                # rej_sheet : la date n'est pas convertie (gardée telle quelle pour
                # préserver la traçabilité brute, comme dans l'ancienne version).
                yield [
                    r.get("type_MT"), r.get("reference"), d,
                    r.get("devise"), r.get("montant"), r.get("code_donneur_dordre"),
                    r.get("donneur_dordre") or r.get("institution_name"),
                    r.get("beneficiaire"), r.get("pays_iso3"), r.get("correspondant"),
                    r.get("_nt_status"), r.get("_csv_status"), r.get("_csv_identifier"),
                    r.get("_rejet_raison"), r.get("source_pdf"),
                ]
        _flush_sheet(rej_sheet, rej_headers, _rej_rows())
        logger.info("Sheet Sortants_Rejetes_NtStatus: %d lignes (mode 3 transfer analysis)",
                    len(nt_status_rejected_rows))

    wb.save(out_path)
    logger.info("Transfer analysis workbook created: %s (matched: %d, unmatched_mt900: %d, suspens: %d, exceptions: %d)",
                out_path, len(matched_rows), len(unmatched_mt900_rows), len(suspens_rows), len(exception_rows))
    return out_path


# ================================================================
# MODE 4 : Rapprochement MT950
# ================================================================

def create_mt950_reconciliation_workbook(
    rapproches: List[Dict],
    non_rapproches_messages: List[Dict],
    non_rapproches_f61: List[Dict],
    out_dir: Path,
    sub_mode: str = "entrants",
    date_start: str = None,
    date_end: str = None,
) -> Path:
    """
    Créer un workbook Excel pour le rapprochement MT950.

    Sheets :
    1. Rapproches          – F61 matchés côte-à-côte avec le message
    2. Msg_non_rapproches  – Messages sans correspondance F61
    3. F61_non_rapproches  – Écritures F61 sans correspondance message
    """
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    date_suffix = ""
    if date_start and date_end:
        date_suffix = f"_{date_start}_to_{date_end}"
    elif date_start:
        date_suffix = f"_from_{date_start}"
    elif date_end:
        date_suffix = f"_to_{date_end}"

    out_path = out_dir / f"rapprochement_mt950_{sub_mode}{date_suffix}_{ts}.xlsx"

    wb = Workbook()

    def _convert_date_to_excel(date_str):
        if not date_str:
            return None
        try:
            if isinstance(date_str, str) and re.match(r'^\d{4}-\d{2}-\d{2}$', date_str):
                return datetime.strptime(date_str, '%Y-%m-%d')
            return date_str
        except Exception:
            return date_str

    def _adjust_widths(sheet, num_cols):
        try:
            for col_idx in range(1, num_cols + 1):
                vals = [cell.value for cell in sheet[get_column_letter(col_idx)] if cell.value is not None]
                max_len = max((len(str(v)) for v in vals), default=10)
                sheet.column_dimensions[get_column_letter(col_idx)].width = min(60, max(12, max_len + 2))
        except Exception:
            pass

    # ---- Sheet 1 : Rapproches ----
    rap_sheet = wb.active
    rap_sheet.title = "Rapproches"

    rap_headers = [
        # F61 traceback
        "N° F61",
        # MT950 side
        "MT950 C/D",
        "MT950 Réf. Propriétaire",
        "MT950 Réf. Institution",
        "MT950 Date Valeur",
        "MT950 Détails",
        # Message side
        "Type MT",
        "Référence",
        "Date Référence",
        "Montant",
        "Devise",
        "Code donneur d'ordre",
        "Donneur d'ordre",
        "Bénéficiaire",
        "Pays ISO3",
        "Correspondant",
        "Commentaires",
        "Source PDF",
    ]
    rap_sheet.append(rap_headers)

    for r in rapproches:
        vd_raw = r.get("mt950_value_date") or ""
        # Format YYMMDD -> DD/MM/YYYY display
        if len(vd_raw) == 6:
            try:
                vd_display = f"{vd_raw[4:6]}/{vd_raw[2:4]}/20{vd_raw[0:2]}"
            except Exception:
                vd_display = vd_raw
        else:
            vd_display = vd_raw

        date_ref = _convert_date_to_excel(r.get("msg_date_reference"))

        row_data = [
            r.get("f61_index"),
            r.get("mt950_cd"),
            r.get("mt950_ref_owner"),
            r.get("mt950_ref_serv"),
            vd_display,
            r.get("mt950_supplementary_details"),
            r.get("msg_type_MT"),
            r.get("msg_reference"),
            date_ref,
            r.get("msg_montant"),
            r.get("msg_devise"),
            r.get("msg_code_donneur_dordre"),
            r.get("msg_donneur_dordre"),
            r.get("msg_beneficiaire"),
            r.get("msg_pays_iso3"),
            r.get("msg_correspondant"),
            r.get("msg_commentaires"),
            r.get("msg_source_pdf"),
        ]
        rap_sheet.append(row_data)

        current_row = rap_sheet.max_row
        if date_ref and isinstance(date_ref, datetime):
            rap_sheet.cell(row=current_row, column=9).number_format = 'DD/MM/YYYY'

    _adjust_widths(rap_sheet, len(rap_headers))

    # ---- Sheet 2 : Msg_non_rapproches ----
    msg_nr_headers = [
        "correspondant",
        "date_reference",
        "reference",
        "reference_origine",
        "type_MT",
        "pays_iso3",
        "Code du donneur d'ordre",
        "donneur d'ordre",
        "Bénéficiaire",
        "montant",
        "devise",
        "commentaires",
        "source_pdf",
    ]

    msg_nr_sheet = wb.create_sheet(title="Msg_non_rapproches")
    msg_nr_sheet.append(msg_nr_headers)

    for r in non_rapproches_messages:
        date_ref = _convert_date_to_excel(r.get("date_reference"))
        row_data = [
            r.get("correspondant"),
            date_ref,
            r.get("reference"),
            r.get("reference_origine"),
            r.get("type_MT"),
            r.get("pays_iso3"),
            r.get("code_donneur_dordre"),
            r.get("donneur_dordre") or r.get("institution_name"),
            r.get("beneficiaire"),
            r.get("montant"),
            r.get("devise"),
            r.get("commentaires"),
            r.get("source_pdf"),
        ]
        msg_nr_sheet.append(row_data)
        current_row = msg_nr_sheet.max_row
        if date_ref and isinstance(date_ref, datetime):
            msg_nr_sheet.cell(row=current_row, column=2).number_format = 'DD/MM/YYYY'

    _adjust_widths(msg_nr_sheet, len(msg_nr_headers))

    # ---- Sheet 3 : F61_non_rapproches ----
    f61_nr_headers = [
        "N° F61",
        "C/D",
        "Réf. Propriétaire",
        "Réf. Institution",
        "Code Id.",
        "Montant",
        "Date Valeur",
        "Détails",
    ]

    f61_nr_sheet = wb.create_sheet(title="F61_non_rapproches")
    f61_nr_sheet.append(f61_nr_headers)

    for f in non_rapproches_f61:
        vd_raw = f.get("value_date") or ""
        if len(vd_raw) == 6:
            try:
                vd_display = f"{vd_raw[4:6]}/{vd_raw[2:4]}/20{vd_raw[0:2]}"
            except Exception:
                vd_display = vd_raw
        else:
            vd_display = vd_raw
        row_data = [
            f.get("f61_index"),
            f.get("cd"),
            f.get("ref_owner"),
            f.get("ref_serv"),
            f.get("identification_code"),
            f.get("amount"),
            vd_display,
            f.get("supplementary_details"),
        ]
        f61_nr_sheet.append(row_data)

    _adjust_widths(f61_nr_sheet, len(f61_nr_headers))

    # ---- Hyperlinks dans les en-têtes source_pdf ----
    # Rapproches: "Source PDF" header → lien vers Msg_non_rapproches
    source_pdf_col = len(rap_headers)  # dernière colonne
    rap_header_cell = rap_sheet.cell(row=1, column=source_pdf_col)
    try:
        rap_header_cell.hyperlink = "#Msg_non_rapproches!A1"
        rap_header_cell.style = "Hyperlink"
        rap_header_cell.value = "Source PDF ➡"
    except Exception:
        pass

    # Msg_non_rapproches: "source_pdf" header → lien retour vers Rapproches
    source_pdf_col_msg = len(msg_nr_headers)  # dernière colonne
    msg_nr_header_cell = msg_nr_sheet.cell(row=1, column=source_pdf_col_msg)
    try:
        msg_nr_header_cell.hyperlink = "#Rapproches!A1"
        msg_nr_header_cell.style = "Hyperlink"
        msg_nr_header_cell.value = "⬅ Rapproches"
    except Exception:
        pass

    # F61_non_rapproches: ajouter une colonne lien retour vers Rapproches
    f61_link_col = len(f61_nr_headers) + 1
    f61_link_cell = f61_nr_sheet.cell(row=1, column=f61_link_col)
    try:
        f61_link_cell.hyperlink = "#Rapproches!A1"
        f61_link_cell.style = "Hyperlink"
        f61_link_cell.value = "⬅ Rapproches"
    except Exception:
        pass

    wb.save(out_path)
    logger.info(
        "MT950 reconciliation workbook created: %s (rapproches: %d, msg_nr: %d, f61_nr: %d)",
        out_path, len(rapproches), len(non_rapproches_messages), len(non_rapproches_f61),
    )
    return out_path


# ================================================================
# MODE 4 v2 : Rapprochement MT950 — feuilles séparées par catégorie
# ================================================================

def create_mt950_reconciliation_workbook_v2(
    rapproches_by_cat: Dict[str, List[Dict]],
    non_rap_msg_by_cat: Dict[str, List[Dict]],
    non_rapproches_f61: List[Dict],
    out_dir: Path,
    sub_mode: str = "entrants",
    date_start: str = None,
    date_end: str = None,
) -> Path:
    """
    Créer un workbook Excel pour le rapprochement MT950 v2.
    
    Produit des feuilles séparées par catégorie, exactement comme les modes 1/2,
    avec des feuilles _rapprochés et _non_rapprochés pour chaque catégorie.

    Catégories possibles : summary, BEACCMCX091, Exceptions_323201, 
                           Autres_Exceptions, BANQUE DE FRANCE, forex

    Sheets générées (pour chaque catégorie non vide) :
    - <catégorie>_rap      — Messages de cette catégorie rapprochés avec F61
    - <catégorie>_non_rap  — Messages de cette catégorie sans correspondance F61
    
    Plus :
    - F61_non_rapproches   — Écritures F61 sans correspondance message
    """
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    date_suffix = ""
    if date_start and date_end:
        date_suffix = f"_{date_start}_to_{date_end}"
    elif date_start:
        date_suffix = f"_from_{date_start}"
    elif date_end:
        date_suffix = f"_to_{date_end}"

    out_path = out_dir / f"rapprochement_mt950_{sub_mode}{date_suffix}_{ts}.xlsx"

    wb = Workbook()
    # Supprimer la feuille par défaut, on va les créer nous-mêmes
    wb.remove(wb.active)

    def _convert_date_to_excel(date_str):
        if not date_str:
            return None
        try:
            if isinstance(date_str, str) and re.match(r'^\d{4}-\d{2}-\d{2}$', date_str):
                return datetime.strptime(date_str, '%Y-%m-%d')
            return date_str
        except Exception:
            return date_str

    def _adjust_widths(sheet, num_cols):
        try:
            for col_idx in range(1, num_cols + 1):
                vals = [cell.value for cell in sheet[get_column_letter(col_idx)] if cell.value is not None]
                max_len = max((len(str(v)) for v in vals), default=10)
                sheet.column_dimensions[get_column_letter(col_idx)].width = min(60, max(12, max_len + 2))
        except Exception:
            pass

    def _format_f61_value_date(vd_raw):
        """Format YYMMDD -> DD/MM/YYYY."""
        if not vd_raw or len(vd_raw) != 6:
            return vd_raw or ""
        try:
            return f"{vd_raw[4:6]}/{vd_raw[2:4]}/20{vd_raw[0:2]}"
        except Exception:
            return vd_raw

    # Headers pour les feuilles rapprochées (F61 + message côte-à-côte)
    rap_headers = [
        "N° F61",
        "MT950 C/D",
        "MT950 Réf. Propriétaire",
        "MT950 Réf. Institution",
        "MT950 Date Valeur",
        "MT950 Détails",
        "Type MT",
        "Référence",
        "Date Référence",
        "Montant",
        "Devise",
        "Code donneur d'ordre",
        "Donneur d'ordre",
        "Bénéficiaire",
        "Pays ISO3",
        "Correspondant",
        "Commentaires",
        "Source PDF",
    ]

    # Headers pour les feuilles non-rapprochées (identiques au mode standard)
    msg_headers = [
        "correspondant",
        "date_reference",
        "reference",
        "reference_origine",
        "type_MT",
        "pays_iso3",
        "Code du donneur d'ordre",
        "donneur d'ordre",
        "Bénéficiaire",
        "montant",
        "devise",
        "commentaires",
        "source_pdf",
    ]

    # Headers pour F61 non rapprochés
    f61_nr_headers = [
        "N° F61",
        "C/D",
        "Réf. Propriétaire",
        "Réf. Institution",
        "Code Id.",
        "Montant",
        "Date Valeur",
        "Détails",
    ]

    # Ordre des catégories pour les feuilles
    category_order = ["summary", "BEACCMCX091", "Exceptions_323201", 
                      "Autres_Exceptions", "BANQUE DE FRANCE", "forex"]

    sheet_names_created = []

    for cat_name in category_order:
        rap_rows = rapproches_by_cat.get(cat_name, [])
        non_rap_rows = non_rap_msg_by_cat.get(cat_name, [])
        
        if not rap_rows and not non_rap_rows:
            continue  # Skip catégorie vide
        
        # --- Feuille rapprochés ---
        if rap_rows:
            sheet_title_rap = _sanitize_sheet_title(f"{cat_name}_rap")
            ws_rap = wb.create_sheet(title=sheet_title_rap)
            ws_rap.append(rap_headers)
            sheet_names_created.append(sheet_title_rap)

            for r in rap_rows:
                vd_display = _format_f61_value_date(r.get("mt950_value_date"))
                date_ref = _convert_date_to_excel(r.get("msg_date_reference"))

                row_data = [
                    r.get("f61_index"),
                    r.get("mt950_cd"),
                    r.get("mt950_ref_owner"),
                    r.get("mt950_ref_serv"),
                    vd_display,
                    r.get("mt950_supplementary_details"),
                    r.get("msg_type_MT"),
                    r.get("msg_reference"),
                    date_ref,
                    r.get("msg_montant"),
                    r.get("msg_devise"),
                    r.get("msg_code_donneur_dordre"),
                    r.get("msg_donneur_dordre"),
                    r.get("msg_beneficiaire"),
                    r.get("msg_pays_iso3"),
                    r.get("msg_correspondant"),
                    r.get("msg_commentaires"),
                    r.get("msg_source_pdf"),
                ]
                ws_rap.append(row_data)

                current_row = ws_rap.max_row
                if date_ref and isinstance(date_ref, datetime):
                    ws_rap.cell(row=current_row, column=9).number_format = 'DD/MM/YYYY'

            _adjust_widths(ws_rap, len(rap_headers))

        # --- Feuille non-rapprochés ---
        if non_rap_rows:
            sheet_title_nonrap = _sanitize_sheet_title(f"{cat_name}_non_rap")
            ws_nonrap = wb.create_sheet(title=sheet_title_nonrap)
            ws_nonrap.append(msg_headers)
            sheet_names_created.append(sheet_title_nonrap)

            for r in non_rap_rows:
                date_ref = _convert_date_to_excel(r.get("date_reference"))
                row_data = [
                    r.get("correspondant"),
                    date_ref,
                    r.get("reference"),
                    r.get("reference_origine"),
                    r.get("type_MT"),
                    r.get("pays_iso3"),
                    r.get("code_donneur_dordre"),
                    r.get("donneur_dordre") or r.get("institution_name"),
                    r.get("beneficiaire"),
                    r.get("montant"),
                    r.get("devise"),
                    r.get("commentaires"),
                    r.get("source_pdf"),
                ]
                ws_nonrap.append(row_data)
                current_row = ws_nonrap.max_row
                if date_ref and isinstance(date_ref, datetime):
                    ws_nonrap.cell(row=current_row, column=2).number_format = 'DD/MM/YYYY'

            _adjust_widths(ws_nonrap, len(msg_headers))

    # --- Feuille F61_non_rapprochés ---
    if non_rapproches_f61:
        f61_nr_sheet = wb.create_sheet(title="F61_non_rapproches")
        f61_nr_sheet.append(f61_nr_headers)
        sheet_names_created.append("F61_non_rapproches")

        for f in non_rapproches_f61:
            vd_display = _format_f61_value_date(f.get("value_date"))
            row_data = [
                f.get("f61_index"),
                f.get("cd"),
                f.get("ref_owner"),
                f.get("ref_serv"),
                f.get("identification_code"),
                f.get("amount"),
                vd_display,
                f.get("supplementary_details"),
            ]
            f61_nr_sheet.append(row_data)

        _adjust_widths(f61_nr_sheet, len(f61_nr_headers))

    # Si aucune feuille n'a été créée, ajouter une feuille vide
    if not wb.sheetnames:
        ws_empty = wb.create_sheet(title="Aucun_resultat")
        ws_empty.append(["Aucun message extrait ou rapproché."])

    # Ajouter des liens de navigation entre les feuilles
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Ajouter un lien retour vers la première feuille en première ligne (insertion)
        first_sheet = wb.sheetnames[0]
        if sheet_name != first_sheet:
            ws.insert_rows(1, 1)
            back_cell = ws.cell(row=1, column=1)
            back_cell.value = f"⬅ Retour à {first_sheet}"
            try:
                escaped = first_sheet.replace("'", "''")
                back_cell.hyperlink = f"#'{escaped}'!A1"
                back_cell.style = "Hyperlink"
            except Exception:
                pass

    wb.save(out_path)
    
    total_rap = sum(len(v) for v in rapproches_by_cat.values())
    total_non_rap = sum(len(v) for v in non_rap_msg_by_cat.values())
    logger.info(
        "MT950 reconciliation workbook v2 created: %s (rapproches: %d, msg_non_rap: %d, f61_non_rap: %d, sheets: %s)",
        out_path, total_rap, total_non_rap, len(non_rapproches_f61), sheet_names_created,
    )
    return out_path
