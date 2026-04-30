# streamlit_app/app.py
# Interface Streamlit pour l'extracteur PDF SWIFT - Version Hugging Face Spaces
# Version: 2026-01-12 - BIC codes 81, MT900 support
import sys
from pathlib import Path
import tempfile
import shutil
import traceback

import streamlit as st
import pandas as pd

# --- Configuration des chemins pour HF Spaces ---
ROOT = Path(__file__).resolve().parent

# Import des modules backend (tous dans le même répertoire pour HF Spaces)
try:
    from extractor_manager import extract_single, create_workbook, extract_dispatch
    from extractors import bic_utils
    from extractors.mt950 import extract_mt950_entries, match_f61_with_messages
    from extractor_manager import create_mt950_reconciliation_workbook, create_mt950_reconciliation_workbook_v2
    from extractors.excel_extractor import extract_excel_files, detect_correspondant_from_file
    from extractors.eastnet_extractor import (
        extract_from_rje_files as eastnet_extract,
        extract_mt950_entries_from_rje_files,
    )
except Exception as e:
    st.error(f"Impossible d'importer l'extracteur backend: {e}")
    st.stop()

# Test du chargement de la base BIC
try:
    m = bic_utils.load_bic_mapping()
    # Utiliser le cache interne de bic_utils pour obtenir le compte
    # load_bic_mapping charge déjà le fichier, on récupère le nombre depuis le cache fullkey
    from extractors.bic_utils import _BIC_FULLKEY_MAP
    nb_codes = len(_BIC_FULLKEY_MAP) if _BIC_FULLKEY_MAP else len(m)
    st.sidebar.success(f"✅ {nb_codes} codes BIC chargés")
except Exception as e:
    st.sidebar.warning(f"⚠️ BIC mapping: {e}")

# UI configuration
st.set_page_config(page_title="PDF SWIFT Extractor", layout="wide", page_icon="📄")

# Header avec style
st.markdown("""
    <h1 style='text-align: center;'>📄 SWIFT Extractor</h1>
    <p style='text-align: center; color: gray;'>Extraction automatique de messages SWIFT depuis PDF et fichiers Excel</p>
""", unsafe_allow_html=True)

with st.expander("📖 Mode d'emploi", expanded=False):
    st.markdown("""
    **Comment utiliser l'application :**
    1. Sélectionnez le type de messages (entrants, sortants, analyse, rapprochement, ou Excel)
    2. Glissez-déposez un ou plusieurs fichiers PDF ou Excel
    3. Choisissez une date de référence (optionnel)
    4. Cliquez sur **Extraire**
    5. Téléchargez le fichier Excel généré
    
    **Types de messages supportés (PDF) :**
    - **MT202** : Virements interbancaires
    - **MT103** : Virements clients
    - **MT910** : Confirmations de crédit
    - **MT900/fin.900** : Analyse des transferts exécutés
    
    **Fichiers Excel supportés :**
    - **BDF (Banque de France)** : Relevés de compte avec colonnes Date/Libellé/Référence/Contrepartie/Débit/Crédit
    - **CITI (EUR et USD)** : Transactions avec colonnes Date/Devise/Montant/Bénéficiaire/Référence/Détails
    """)

# Direction selector
st.markdown("### 📨 Type de messages")
direction = st.radio(
    "Sélectionnez le type de messages à extraire",
    ("incoming", "outgoing", "transfer_analysis", "mt950_reconciliation", "excel_extraction", "eastnet_extraction"),
    format_func=lambda x: {"incoming": "📥 Messages Entrants (MT202, MT103, MT910)", "outgoing": "📤 Messages Sortants (MT202, MT103)", "transfer_analysis": "🔄 Analyse Transferts Exécutés (MT202/MT103 + fin.900)", "mt950_reconciliation": "📊 Rapprochement MT950", "excel_extraction": "📊 Extraction Fichiers Excel (BDF, CITI, Standard)", "eastnet_extraction": "📦 Archive EastNet (RJE)"}.get(x, x),
    horizontal=False
)

if direction == "incoming":
    st.info("**Messages entrants** : Pour les MT202, le bénéficiaire sera vide. Pour les MT910, le bénéficiaire sera identique au donneur d'ordre.")
elif direction == "outgoing":
    st.info("**Messages sortants** : Pour les MT202, le bénéficiaire sera extrait depuis F58A.")
elif direction == "mt950_reconciliation":
    st.info("**Rapprochement MT950** : Charge le(s) fichier(s) MT950 et les fichiers de messages (MT202, MT103, MT910). Rapproche les écritures F61 du MT950 avec les messages par référence + montant.")
elif direction == "excel_extraction":
    st.info("**Extraction Excel** : Charge les fichiers Excel de correspondants (BDF, CITI, Standard ou CITI USD Relevé de compte). Applique les règles de routage (BEACCMCX091, T2PI/T2RM/T2PL, nivellement, forex, exception BC, intérêts). Génère deux fichiers (entrants et sortants).")
elif direction == "eastnet_extraction":
    st.info("**Archive EastNet (RJE)** : Pour l'instant, ce mode propose la logique du mode 1 (messages entrants) et du mode 4 (rapprochement MT950), directement à partir de fichiers RJE.")
else:
    st.info("**Analyse Transferts Exécutés** : Charge les fichiers MT900 d'un côté et les fichiers MT103/MT202 de l'autre. Le système match les références (F21 du MT900 = F20 du MT103/MT202) pour compléter les infos.")

# File uploader(s) - différent selon le mode
if direction == "excel_extraction":
    st.markdown("### 📁 Fichiers Excel")
else:
    st.markdown("### 📁 Fichiers PDF")

# Gestion du Clear All : utiliser un compteur pour régénérer les clés des uploaders
if 'uploader_key_counter' not in st.session_state:
    st.session_state.uploader_key_counter = 0

def clear_all_files():
    """Réinitialiser tous les fichiers chargés en incrémentant le compteur de clés."""
    st.session_state.uploader_key_counter += 1
    # Réinitialiser aussi les résultats d'extraction
    st.session_state.extraction_results = None
    st.session_state.excel_data = None
    st.session_state.excel_filename = None
    st.session_state.excel_data_entrant = None
    st.session_state.excel_filename_entrant = None
    st.session_state.excel_data_sortant = None
    st.session_state.excel_filename_sortant = None

# Bouton Clear All
col_clear1, col_clear2 = st.columns([4, 1])
with col_clear2:
    if st.button("🗑️ Clear All", help="Supprimer tous les fichiers chargés", use_container_width=True):
        clear_all_files()
        st.rerun()

uploader_suffix = f"_{st.session_state.uploader_key_counter}"

if direction == "transfer_analysis":
    # Mode analyse transferts: deux zones d'upload séparées (disposition verticale)
    st.markdown("#### 📋 Fichiers MT900 (confirmations)")
    uploaded_mt900_files = st.file_uploader(
        "Déposez les fichiers contenant les MT900",
        type="pdf",
        accept_multiple_files=True,
        help="Fichiers PDF contenant les messages MT900/fin.900",
        key=f"mt900_uploader{uploader_suffix}"
    )
    
    st.markdown("#### 📤 Fichiers MT103/MT202 (transferts sortants)")
    uploaded_mt103_202_files = st.file_uploader(
        "Déposez les fichiers contenant les MT103/MT202",
        type="pdf",
        accept_multiple_files=True,
        help="Fichiers PDF contenant les messages MT103 et MT202 sortants",
        key=f"mt103_202_uploader{uploader_suffix}"
    )
    
    # Pour compatibilité avec le reste du code
    uploaded_files = None
    uploaded_rje_files = None
    uploaded_mt950_files = None
    uploaded_mt950_msg_files = None
    uploaded_excel_files = None
    uploaded_rje_mt900_files = None
    uploaded_rje_sortants_files = None
    uploaded_eastnet_csv_files = None
elif direction == "mt950_reconciliation":
    # Mode Rapprochement MT950: deux zones + sous-mode
    mt950_sub_mode = st.radio(
        "Sous-mode de rapprochement",
        ("entrants", "sortants"),
        format_func=lambda x: "📥 Entrants (Crédit C)" if x == "entrants" else "📤 Sortants (Débit D)",
        horizontal=True
    )
    if mt950_sub_mode == "entrants":
        st.caption("Les écritures **Crédit (C)** du MT950 sont rapprochées avec les messages entrants (MT202, MT103, MT910).")
    else:
        st.caption("Les écritures **Débit (D)** du MT950 sont rapprochées avec les messages sortants (MT202, MT103).")
    
    st.markdown("#### 📊 Fichiers MT950")
    uploaded_mt950_files = st.file_uploader(
        "Déposez les fichiers contenant les MT950",
        type="pdf",
        accept_multiple_files=True,
        help="Fichiers PDF contenant les relevés MT950/fin.950",
        key=f"mt950_uploader{uploader_suffix}"
    )
    
    msg_dir_label = "entrants (MT202, MT103, MT910)" if mt950_sub_mode == "entrants" else "sortants (MT202, MT103)"
    st.markdown(f"#### 📨 Fichiers de messages {msg_dir_label}")
    uploaded_mt950_msg_files = st.file_uploader(
        f"Déposez les fichiers contenant les messages {msg_dir_label}",
        type="pdf",
        accept_multiple_files=True,
        help=f"Fichiers PDF contenant les messages {msg_dir_label}",
        key=f"mt950_msg_uploader{uploader_suffix}"
    )
    
    uploaded_files = None
    uploaded_rje_files = None
    uploaded_mt900_files = None
    uploaded_mt103_202_files = None
    uploaded_excel_files = None
    uploaded_rje_mt900_files = None
    uploaded_rje_sortants_files = None
elif direction == "eastnet_extraction":
    # Mode EastNet RJE: sélecteur de correspondant + uploader .rje
    eastnet_sub_mode = st.radio(
        "Sous-mode EastNet",
        ("rje_incoming_mode1", "rje_outgoing_mode2", "rje_transfer_analysis_mode3", "rje_mt950_mode4"),
        format_func=lambda x: {
            "rje_incoming_mode1": "📥 Mode 1 RJE — Messages entrants (MT202, MT103, MT910)",
            "rje_outgoing_mode2": "📤 Mode 2 RJE — Messages sortants (MT202, MT103)",
            "rje_transfer_analysis_mode3": "🔄 Mode 3 RJE — Analyse Transferts Exécutés (MT202/MT103 + fin.900)",
            "rje_mt950_mode4": "📊 Mode 4 RJE — Rapprochement MT950",
        }.get(x, x),
        horizontal=False
    )

    rje_correspondant = st.radio(
        "Correspondant",
        ("CITIUS33XXX", "CITIGB2LXXX", "BDFEFRPPXXX", "SCBLGB2LXXX", "FRNYUS33XXX"),
        format_func=lambda x: {
            "CITIUS33XXX": "🏦 CITI USD (CITIUS33XXX)",
            "CITIGB2LXXX": "🏦 CITI EUR (CITIGB2LXXX)",
            "BDFEFRPPXXX": "🏦 Banque de France (BDFEFRPPXXX)",
            "SCBLGB2LXXX": "🏦 Standard (SCBLGB2LXXX)",
            "FRNYUS33XXX": "🏦 FED",
        }.get(x, x),
        horizontal=True
    )
    st.caption("La direction est détectée automatiquement depuis le bloc 2 SWIFT: I=sortant, O=entrant.")

    if eastnet_sub_mode in ("rje_incoming_mode1", "rje_outgoing_mode2"):
        _label = (
            "Déposez vos fichiers RJE de messages entrants ici"
            if eastnet_sub_mode == "rje_incoming_mode1"
            else "Déposez vos fichiers RJE de messages sortants ici"
        )
        uploaded_rje_files = st.file_uploader(
            _label,
            type=["rje"],
            accept_multiple_files=True,
            help="Archives EastNet au format RJE (messages SWIFT bruts)",
            key=f"rje_uploader{uploader_suffix}"
        )
        uploaded_rje_mt950_files = None
        uploaded_rje_msg_files = None
        uploaded_rje_mt900_files = None
        uploaded_rje_sortants_files = None
        mt950_sub_mode = "entrants"

        # Mode 2 (sortants initiés) — pour CITI USD et BDF, demander en plus
        # le CSV companion exporté depuis EastNets afin de filtrer les messages
        # qui n'ont pas reçu d'acquittement réseau (Nt.Status != 'Network Ack').
        uploaded_eastnet_csv_files = None
        if eastnet_sub_mode == "rje_outgoing_mode2" and rje_correspondant in ("CITIUS33XXX", "BDFEFRPPXXX"):
            st.markdown("#### 📎 CSV companion EastNets (filtre Nt.Status)")
            st.caption(
                "Pour CITI USD et BDF en mode sortants initiés, joignez le CSV "
                "exporté depuis EastNets (mêmes dates que le RJE). Seuls les "
                "messages avec **Nt.Status = 'Network Ack'** seront conservés. "
                "Les autres (Nack, '-', absents du CSV) iront dans la feuille "
                "« Sortants_Rejetes_NtStatus »."
            )
            uploaded_eastnet_csv_files = st.file_uploader(
                "Déposez le(s) CSV EastNets companion",
                type=["csv"],
                accept_multiple_files=True,
                help="Export EastNets : colonnes IO, Reference, Identifier, Status, Nt.Status, ...",
                key=f"eastnet_csv_uploader{uploader_suffix}",
            )
    elif eastnet_sub_mode == "rje_transfer_analysis_mode3":
        st.markdown("#### 📤 Fichiers RJE des messages sortants (MT202/MT103)")
        uploaded_rje_sortants_files = st.file_uploader(
            "Déposez les fichiers RJE des messages sortants",
            type=["rje"],
            accept_multiple_files=True,
            help="Archives EastNet contenant les MT202/MT103 sortants",
            key=f"rje_sortants_uploader{uploader_suffix}"
        )
        st.markdown("#### 🔄 Fichiers RJE des confirmations MT900")
        uploaded_rje_mt900_files = st.file_uploader(
            "Déposez les fichiers RJE contenant les MT900",
            type=["rje"],
            accept_multiple_files=True,
            help="Archives EastNet contenant les confirmations MT900 (fin.900)",
            key=f"rje_mt900_uploader{uploader_suffix}"
        )
        uploaded_rje_files = None
        uploaded_rje_mt950_files = None
        uploaded_rje_msg_files = None
        mt950_sub_mode = "entrants"
    else:
        mt950_sub_mode = st.radio(
            "Sous-mode de rapprochement",
            ("entrants", "sortants"),
            format_func=lambda x: "📥 Entrants (Crédit C)" if x == "entrants" else "📤 Sortants (Débit D)",
            horizontal=True
        )
        st.markdown("#### 📊 Fichiers RJE contenant les MT950")
        uploaded_rje_mt950_files = st.file_uploader(
            "Déposez les fichiers RJE contenant les MT950",
            type=["rje"],
            accept_multiple_files=True,
            help="Fichiers RJE avec messages MT950 (tags :61:)",
            key=f"rje_mt950_uploader{uploader_suffix}"
        )
        msg_dir_label = "entrants (MT202, MT103, MT910)" if mt950_sub_mode == "entrants" else "sortants (MT202, MT103)"
        st.markdown(f"#### 📨 Fichiers RJE de messages {msg_dir_label}")
        uploaded_rje_msg_files = st.file_uploader(
            f"Déposez les fichiers RJE de messages {msg_dir_label}",
            type=["rje"],
            accept_multiple_files=True,
            help=f"Fichiers RJE contenant les messages {msg_dir_label}",
            key=f"rje_msg_uploader{uploader_suffix}"
        )
        uploaded_rje_files = None

    # Pour compatibilité avec le reste du code
    uploaded_files = None
    uploaded_mt900_files = None
    uploaded_mt103_202_files = None
    uploaded_mt950_files = None
    uploaded_mt950_msg_files = None
    uploaded_excel_files = None
    # CSV companion EastNets : initialisé à None par défaut (uniquement utilisé
    # en mode 2 sortants pour CITI USD / BDF, voir UI plus haut).
    try:
        uploaded_eastnet_csv_files
    except NameError:
        uploaded_eastnet_csv_files = None

elif direction == "excel_extraction":
    # Mode Excel: sélecteur de correspondant + uploader xlsx/xls
    excel_correspondant = st.radio(
        "Correspondant Excel",
        ("BDF", "CITI", "Standard", "CITI_USD_Releve"),
        format_func=lambda x: {
            "BDF": "🏦 BDF (Banque de France)",
            "CITI": "🏦 CITI (Citibank EUR / USD)",
            "Standard": "🏦 Standard (SCBL)",
            "CITI_USD_Releve": "🏦 CITI USD (Relevé de compte)",
        }.get(x, x),
        horizontal=True
    )
    if excel_correspondant == "BDF":
        st.caption("Relevés BDF : colonnes Date opération / Date valeur / Libellé mouvement / Référence demandeur / Référence client / Nom contrepartie / Débit / Crédit")
    elif excel_correspondant == "CITI":
        st.caption("CITI : colonnes Date de valeur / Date du relevé / Devise / Montant / Bénéficiaire-Remettant / Type / Référence bancaire / Description / Détails du paiement")
    elif excel_correspondant == "CITI_USD_Releve":
        st.caption("Relevé de compte CITI USD brut (.xls) : sections journalières avec Entry Date / Value Date / Customer Reference / Bank Reference / Transaction Description / By Order Of / Transaction Amount")
    else:
        st.caption("Standard (SCBL) : colonnes Account / Currency / Date / Description / Withdrawal / Deposit / Balance")

    uploaded_excel_files = st.file_uploader(
        "Déposez vos fichiers Excel (.xlsx ou .xls) ici",
        type=["xlsx", "xls"],
        accept_multiple_files=True,
        help="Fichiers Excel du correspondant sélectionné",
        key=f"excel_uploader{uploader_suffix}"
    )

    # Pour compatibilité avec le reste du code
    uploaded_files = None
    uploaded_mt900_files = None
    uploaded_mt103_202_files = None
    uploaded_mt950_files = None
    uploaded_mt950_msg_files = None
    uploaded_rje_mt900_files = None
    uploaded_rje_sortants_files = None
    uploaded_eastnet_csv_files = None
else:
    # Mode standard: un seul uploader
    uploaded_files = st.file_uploader(
        "Déposez vos fichiers PDF ici",
        type="pdf",
        accept_multiple_files=True,
        help="Vous pouvez sélectionner plusieurs fichiers à la fois",
        key=f"main_uploader{uploader_suffix}"
    )
    uploaded_rje_files = None
    uploaded_mt900_files = None
    uploaded_mt103_202_files = None
    uploaded_mt950_files = None
    uploaded_mt950_msg_files = None
    uploaded_excel_files = None
    uploaded_rje_mt900_files = None
    uploaded_rje_sortants_files = None
    uploaded_eastnet_csv_files = None

# Date filter - PLAGE DE DATES
st.markdown("### 📅 Filtre par plage de dates")
from datetime import date as date_type, datetime, timedelta
default_date = date_type.today()
col_date1, col_date2 = st.columns(2)
with col_date1:
    date_debut = st.date_input("Date de début", value=default_date, help="Date de début de la plage (incluse)")
with col_date2:
    date_fin = st.date_input("Date de fin", value=default_date, help="Date de fin de la plage (incluse)")

if date_debut > date_fin:
    st.warning("⚠️ La date de début doit être antérieure ou égale à la date de fin.")

# Extract button
col1, col2, col3 = st.columns([1, 2, 1])
with col2:
    run_button = st.button("🚀 Extraire les données", use_container_width=True, type="primary")

# Initialiser session_state pour persister les résultats après téléchargement
if 'extraction_results' not in st.session_state:
    st.session_state.extraction_results = None
if 'excel_data' not in st.session_state:
    st.session_state.excel_data = None
if 'excel_filename' not in st.session_state:
    st.session_state.excel_filename = None

# Helper function
def save_uploaded_to_temp(uploaded) -> Path:
    tmpdir = Path(tempfile.mkdtemp(prefix="pdf_extr_"))
    dest = tmpdir / uploaded.name
    with open(dest, "wb") as f:
        f.write(uploaded.getbuffer())
    return dest

# Main extraction logic
if run_button:
    # Vérification des fichiers selon le mode
    if direction == "transfer_analysis":
        if not uploaded_mt900_files and not uploaded_mt103_202_files:
            st.warning("⚠️ Aucun fichier sélectionné. Veuillez charger au moins des fichiers MT900 ou MT103/MT202.")
            has_files = False
        elif not uploaded_mt900_files:
            st.warning("⚠️ Aucun fichier MT900 sélectionné. Veuillez charger les fichiers contenant les confirmations MT900.")
            has_files = False
        elif not uploaded_mt103_202_files:
            st.warning("⚠️ Aucun fichier MT103/MT202 sélectionné. Veuillez charger les fichiers contenant les transferts sortants.")
            has_files = False
        else:
            has_files = True
    elif direction == "mt950_reconciliation":
        if not uploaded_mt950_files and not uploaded_mt950_msg_files:
            st.warning("⚠️ Aucun fichier sélectionné. Veuillez charger les fichiers MT950 et les fichiers de messages.")
            has_files = False
        elif not uploaded_mt950_files:
            st.warning("⚠️ Aucun fichier MT950 sélectionné.")
            has_files = False
        elif not uploaded_mt950_msg_files:
            st.warning("⚠️ Aucun fichier de messages sélectionné.")
            has_files = False
        else:
            has_files = True
    elif direction == "excel_extraction":
        if not uploaded_excel_files:
            st.warning("⚠️ Aucun fichier Excel sélectionné.")
            has_files = False
        else:
            has_files = True
    elif direction == "eastnet_extraction":
        if eastnet_sub_mode in ("rje_incoming_mode1", "rje_outgoing_mode2"):
            if not uploaded_rje_files:
                st.warning("⚠️ Aucun fichier RJE sélectionné.")
                has_files = False
            else:
                has_files = True
        elif eastnet_sub_mode == "rje_transfer_analysis_mode3":
            if not uploaded_rje_sortants_files and not uploaded_rje_mt900_files:
                st.warning("⚠️ Aucun fichier RJE sélectionné. Chargez les fichiers sortants et les fichiers MT900.")
                has_files = False
            elif not uploaded_rje_sortants_files:
                st.warning("⚠️ Aucun fichier RJE sortants (MT202/MT103) sélectionné.")
                has_files = False
            elif not uploaded_rje_mt900_files:
                st.warning("⚠️ Aucun fichier RJE MT900 sélectionné.")
                has_files = False
            else:
                has_files = True
        else:
            if not uploaded_rje_mt950_files and not uploaded_rje_msg_files:
                st.warning("⚠️ Aucun fichier RJE sélectionné. Chargez les fichiers MT950 et les fichiers messages.")
                has_files = False
            elif not uploaded_rje_mt950_files:
                st.warning("⚠️ Aucun fichier RJE MT950 sélectionné.")
                has_files = False
            elif not uploaded_rje_msg_files:
                st.warning("⚠️ Aucun fichier RJE de messages sélectionné.")
                has_files = False
            else:
                has_files = True
    else:
        if not uploaded_files:
            st.warning("⚠️ Aucun fichier sélectionné.")
            has_files = False
        else:
            has_files = True
    
    if has_files:
        # Réinitialiser les résultats précédents (y compris les paires entrant/sortant
        # pour éviter qu'un ancien run laisse afficher des fichiers obsolètes)
        st.session_state.extraction_results = None
        st.session_state.excel_data = None
        st.session_state.excel_filename = None
        st.session_state.excel_data_entrant = None
        st.session_state.excel_filename_entrant = None
        st.session_state.excel_data_sortant = None
        st.session_state.excel_filename_sortant = None
        # Réinitialiser la liste des rejets Nt.Status pour ce run.
        st.session_state['nt_status_rejected_rows'] = []
        
        rows = []
        beaccmcx091_rows = []  # Liste séparée pour les BEACCMCX091
        errors = []
        tmp_dirs = []
        
        all_missing_codes = {"unmapped": set(), "empty": set()}
        exception_323201_rows = []  # Liste pour les exceptions 323201
        other_exceptions_rows = []  # Liste pour les autres exceptions (EUR/nivellement)
        banque_de_france_rows = []  # MT103 USD avec BANQUE DE FRANCE / FW021083459
        forex_rows = []  # MT910 entrants avec code donneur forex
        bdf_corr_exception_rows = []  # MT202 sortants exceptions correspondants BdF
        
        if direction == "transfer_analysis":
            # ======== MODE ANALYSE DES TRANSFERTS ========
            # Logique: extraire MT900 séparément, puis les MT103/MT202 comme sortants, puis matcher
            from extractor_manager import extract_mt900_only, match_mt900_with_transfers, create_transfer_analysis_workbook
            
            total_files = len(uploaded_mt900_files) + len(uploaded_mt103_202_files)
            progress = st.progress(0, text="Initialisation...")
            idx = 0
            
            st.info(f"🔄 Traitement de {len(uploaded_mt900_files)} fichier(s) MT900 et {len(uploaded_mt103_202_files)} fichier(s) MT103/MT202...")
            
            # Étape 1: Extraire tous les MT900
            all_mt900_rows = []
            for uf in uploaded_mt900_files:
                idx += 1
                progress.progress(idx / total_files, text=f"MT900 : {uf.name} ({idx}/{total_files})")
                
                try:
                    tmp_path = save_uploaded_to_temp(uf)
                    tmp_dirs.append(tmp_path.parent)
                    
                    mt900_rows, missing = extract_mt900_only(tmp_path)
                    all_missing_codes["unmapped"].update(missing.get("unmapped", set()))
                    all_missing_codes["empty"].update(missing.get("empty", set()))
                    
                    all_mt900_rows.extend(mt900_rows)
                    st.success(f"✅ {uf.name}: {len(mt900_rows)} MT900 extraits")
                    
                except Exception as e:
                    tb = traceback.format_exc()
                    errors.append((uf.name, str(e)))
                    st.error(f"❌ Erreur: {uf.name}")
                    with st.expander(f"Détails de l'erreur pour {uf.name}"):
                        st.code(tb)
            
            # Étape 2: Extraire tous les MT103/MT202 (traités comme sortants)
            all_transfer_rows = []
            for uf in uploaded_mt103_202_files:
                idx += 1
                progress.progress(idx / total_files, text=f"MT103/MT202 : {uf.name} ({idx}/{total_files})")
                
                try:
                    tmp_path = save_uploaded_to_temp(uf)
                    tmp_dirs.append(tmp_path.parent)
                    
                    # Utiliser extract_dispatch avec direction="outgoing" pour bénéficier de toute la logique sortante
                    # (F58A bénéficiaire, codes Trésor/CCF, etc.)
                    new_rows, new_beac_rows, new_exc_rows, new_other_exc_rows, new_bdf_rows, new_forex_rows, new_bdf_corr_exc_rows, missing = extract_dispatch(tmp_path, direction="outgoing")
                    all_missing_codes["unmapped"].update(missing.get("unmapped", set()))
                    all_missing_codes["empty"].update(missing.get("empty", set()))
                    
                    # Filtrer pour ne garder que MT103 et MT202 (exclure MT910 et autres)
                    transfer_rows = [r for r in new_rows if r.get("type_MT") and ("103" in r.get("type_MT") or "202" in r.get("type_MT"))]
                    
                    all_transfer_rows.extend(transfer_rows)
                    st.success(f"✅ {uf.name}: {len(transfer_rows)} MT103/MT202 extraits")
                    
                except Exception as e:
                    tb = traceback.format_exc()
                    errors.append((uf.name, str(e)))
                    st.error(f"❌ Erreur: {uf.name}")
                    with st.expander(f"Détails de l'erreur pour {uf.name}"):
                        st.code(tb)
            
            progress.progress(1.0, text="Matching des références...")
            
            # Étape 3: Matcher les MT900 avec les MT103/MT202
            # Retourne maintenant 4 éléments: matched, suspens, exceptions, unmatched_mt900
            matched_rows, suspens_rows, exception_mt900_rows, unmatched_mt900_rows = match_mt900_with_transfers(all_mt900_rows, all_transfer_rows)
            
            rows = matched_rows
            st.session_state.suspens_rows = suspens_rows
            st.session_state.exception_mt900_rows = exception_mt900_rows
            st.session_state.unmatched_mt900_rows = unmatched_mt900_rows
            
            progress.empty()
            st.info(f"📊 **Résumé** : {len(matched_rows)} MT900 matchés, {len(unmatched_mt900_rows)} MT900 sans correspondant, {len(suspens_rows)} MT103/MT202 en suspens, {len(exception_mt900_rows)} MT900 en exception")
            
        elif direction == "mt950_reconciliation":
            # ======== MODE RAPPROCHEMENT MT950 (v2 — feuilles séparées par catégorie) ========
            total_files = len(uploaded_mt950_files) + len(uploaded_mt950_msg_files)
            progress = st.progress(0, text="Initialisation...")
            idx = 0
            
            msg_direction = "incoming" if mt950_sub_mode == "entrants" else "outgoing"
            st.info(f"🔄 Traitement de {len(uploaded_mt950_files)} fichier(s) MT950 et {len(uploaded_mt950_msg_files)} fichier(s) de messages ({mt950_sub_mode})...")
            
            # Étape 1: Extraire les F61 de tous les MT950
            all_f61_entries = []
            for uf in uploaded_mt950_files:
                idx += 1
                progress.progress(idx / total_files, text=f"MT950 : {uf.name} ({idx}/{total_files})")
                try:
                    tmp_path = save_uploaded_to_temp(uf)
                    tmp_dirs.append(tmp_path.parent)
                    f61_entries = extract_mt950_entries(tmp_path)
                    all_f61_entries.extend(f61_entries)
                    target_cd = "C" if mt950_sub_mode == "entrants" else "D"
                    n_target = sum(1 for e in f61_entries if e.get("cd") == target_cd)
                    st.success(f"✅ {uf.name}: {len(f61_entries)} F61 extraits ({n_target} {target_cd})")
                except Exception as e:
                    tb = traceback.format_exc()
                    errors.append((uf.name, str(e)))
                    st.error(f"❌ Erreur: {uf.name}")
                    with st.expander(f"Détails de l'erreur pour {uf.name}"):
                        st.code(tb)
            
            # Étape 2: Extraire les messages NORMALEMENT (comme mode 1/2),
            # en conservant les catégories séparées
            all_summary_rows = []
            all_beac_rows = []
            all_exc_323201_rows = []
            all_other_exc_rows = []
            all_bdf_rows = []
            all_forex_rows_mt950 = []
            all_bdf_corr_exc_rows_mt950 = []
            
            for uf in uploaded_mt950_msg_files:
                idx += 1
                progress.progress(idx / total_files, text=f"Messages : {uf.name} ({idx}/{total_files})")
                try:
                    tmp_path = save_uploaded_to_temp(uf)
                    tmp_dirs.append(tmp_path.parent)
                    new_rows, new_beac_rows, new_exc_rows, new_other_exc_rows, new_bdf_rows, new_forex_rows, new_bdf_corr_exc_rows, missing = extract_dispatch(tmp_path, direction=msg_direction)
                    all_missing_codes["unmapped"].update(missing.get("unmapped", set()))
                    all_missing_codes["empty"].update(missing.get("empty", set()))
                    
                    # Normaliser chaque catégorie (source_pdf, donneur_dordre, beneficiaire)
                    def _normalize_rows(row_list, pdf_name):
                        for r in row_list:
                            if "beneficiaire" not in r:
                                r["beneficiaire"] = None
                            if "donneur_dordre" not in r:
                                r["donneur_dordre"] = r.get("institution_name") or None
                            if not r.get("source_pdf"):
                                r["source_pdf"] = pdf_name
                            if not r.get("type_MT"):
                                r["type_MT"] = None
                    
                    # Filtrer pour ne garder que MT202, MT103, MT910
                    def _filter_valid(rlist):
                        return [r for r in rlist if r.get("type_MT") and any(t in r.get("type_MT") for t in ("202", "103", "910"))]
                    
                    _normalize_rows(new_rows, uf.name)
                    _normalize_rows(new_beac_rows, uf.name)
                    _normalize_rows(new_exc_rows, uf.name)
                    _normalize_rows(new_other_exc_rows, uf.name)
                    _normalize_rows(new_bdf_rows, uf.name)
                    _normalize_rows(new_forex_rows, uf.name)
                    _normalize_rows(new_bdf_corr_exc_rows, uf.name)
                    
                    all_summary_rows.extend(_filter_valid(new_rows))
                    all_beac_rows.extend(_filter_valid(new_beac_rows))
                    all_exc_323201_rows.extend(_filter_valid(new_exc_rows))
                    all_other_exc_rows.extend(_filter_valid(new_other_exc_rows))
                    all_bdf_rows.extend(_filter_valid(new_bdf_rows))
                    all_forex_rows_mt950.extend(_filter_valid(new_forex_rows))
                    all_bdf_corr_exc_rows_mt950.extend(_filter_valid(new_bdf_corr_exc_rows))
                    
                    total_extracted = len(new_rows) + len(new_beac_rows) + len(new_exc_rows) + len(new_other_exc_rows) + len(new_bdf_rows) + len(new_forex_rows) + len(new_bdf_corr_exc_rows)
                    st.success(f"✅ {uf.name}: {total_extracted} messages extraits")
                except Exception as e:
                    tb = traceback.format_exc()
                    errors.append((uf.name, str(e)))
                    st.error(f"❌ Erreur: {uf.name}")
                    with st.expander(f"Détails de l'erreur pour {uf.name}"):
                        st.code(tb)
            
            progress.progress(1.0, text="Rapprochement en cours...")
            
            # Étape 3: Rapprocher chaque catégorie séparément avec les F61
            from extractors.mt950 import match_f61_with_messages as match_f61
            
            # Combiner TOUTES les catégories pour le matching global
            # (un F61 peut matcher un msg de n'importe quelle catégorie)
            categories = {
                "summary": all_summary_rows,
                "BEACCMCX091": all_beac_rows,
                "Exceptions_323201": all_exc_323201_rows,
                "Autres_Exceptions": all_other_exc_rows,
                "BANQUE DE FRANCE": all_bdf_rows,
                "forex": all_forex_rows_mt950,
                "Exceptions_Correspondants": all_bdf_corr_exc_rows_mt950,
            }
            
            # Pool de tous les messages pour le matching
            all_msg_pool = []
            msg_to_category = {}  # id(msg) -> category_name
            for cat_name, cat_rows in categories.items():
                for r in cat_rows:
                    all_msg_pool.append(r)
                    msg_to_category[id(r)] = cat_name
            
            # Matching global
            rapproches_raw, non_rap_msg_raw, non_rap_f61 = match_f61(
                all_f61_entries, all_msg_pool, sub_mode=mt950_sub_mode
            )
            
            # Répartir les rapprochés par catégorie d'origine
            rapproches_by_cat = {cat: [] for cat in categories}
            for rap in rapproches_raw:
                # Retrouver le message d'origine via la référence et le source_pdf
                matched_msg_ref = rap.get("msg_reference")
                matched_msg_src = rap.get("msg_source_pdf")
                matched_msg_type = rap.get("msg_type_MT")
                matched_msg_montant = rap.get("msg_montant")
                
                # Chercher dans le pool pour identifier la catégorie
                found_cat = "summary"  # default
                for msg in all_msg_pool:
                    if (msg.get("reference") == matched_msg_ref and 
                        msg.get("source_pdf") == matched_msg_src and
                        msg.get("type_MT") == matched_msg_type and
                        msg.get("montant") == matched_msg_montant):
                        found_cat = msg_to_category.get(id(msg), "summary")
                        break
                rapproches_by_cat[found_cat].append(rap)
            
            # Répartir les non-rapprochés par catégorie d'origine
            non_rap_msg_by_cat = {cat: [] for cat in categories}
            for msg in non_rap_msg_raw:
                cat = msg_to_category.get(id(msg), "summary")
                non_rap_msg_by_cat[cat].append(msg)
            
            # Stocker pour la génération du workbook
            rows = rapproches_raw  # pour le flow de date-filtering et le check "if rows:"
            st.session_state.mt950_rapproches_by_cat = rapproches_by_cat
            st.session_state.mt950_non_rap_msg_by_cat = non_rap_msg_by_cat
            st.session_state.mt950_non_rap_f61 = non_rap_f61
            st.session_state.mt950_sub_mode = mt950_sub_mode
            st.session_state.mt950_categories = categories
            
            progress.empty()
            total_rap = sum(len(v) for v in rapproches_by_cat.values())
            total_non_rap = sum(len(v) for v in non_rap_msg_by_cat.values())
            st.info(f"📊 **Résumé** : {total_rap} rapprochés, {total_non_rap} messages non rapprochés, {len(non_rap_f61)} F61 non rapprochés")
            
            # Détail par catégorie
            for cat_name in categories:
                n_rap = len(rapproches_by_cat[cat_name])
                n_non = len(non_rap_msg_by_cat[cat_name])
                if n_rap > 0 or n_non > 0:
                    st.caption(f"  📂 **{cat_name}** : {n_rap} rapprochés, {n_non} non rapprochés")
            
        elif direction == "excel_extraction":
            # ======== MODE EXTRACTION EXCEL (BDF, CITI) ========
            progress = st.progress(0, text="Initialisation...")
            total = len(uploaded_excel_files)
            idx = 0

            st.info(f"🔄 Traitement de {total} fichier(s) Excel ({excel_correspondant})...")

            temp_excel_paths = []
            for uf in uploaded_excel_files:
                idx += 1
                progress.progress(idx / total, text=f"Enregistrement : {uf.name} ({idx}/{total})")
                try:
                    tmp_path = save_uploaded_to_temp(uf)
                    tmp_dirs.append(tmp_path.parent)
                    temp_excel_paths.append(str(tmp_path))
                except Exception as e:
                    errors.append((uf.name, f"Impossible d'enregistrer temporairement: {e}"))
                    st.error(f"❌ Erreur avec {uf.name}")

            if temp_excel_paths:
                try:
                    # Trouver le fichier bic_codes.xlsx
                    bic_file = ROOT / "data" / "bic_codes.xlsx"
                    if not bic_file.exists():
                        bic_file = Path("data/bic_codes.xlsx")
                    xlsx_bic_path = str(bic_file) if bic_file.exists() else None

                    new_rows, new_beac_rows, new_other_exc_rows, new_forex_rows, new_bdf_rows, missing_codes = extract_excel_files(
                        temp_excel_paths, excel_correspondant, xlsx_bic_path=xlsx_bic_path
                    )

                    all_missing_codes["unmapped"].update(missing_codes.get("unmapped", set()))
                    all_missing_codes["empty"].update(missing_codes.get("empty", set()))

                    # Normaliser les rows (même pattern que le mode standard)
                    def _normalize_excel_rows(row_list, target_list):
                        for r in row_list:
                            if "beneficiaire" not in r:
                                r["beneficiaire"] = None
                            if "donneur_dordre" not in r:
                                r["donneur_dordre"] = r.get("institution_name")
                            if not r.get("type_MT"):
                                r["type_MT"] = None
                            target_list.append(r)

                    _normalize_excel_rows(new_rows, rows)
                    _normalize_excel_rows(new_beac_rows, beaccmcx091_rows)
                    _normalize_excel_rows(new_other_exc_rows, other_exceptions_rows)
                    _normalize_excel_rows(new_forex_rows, forex_rows)
                    _normalize_excel_rows(new_bdf_rows, banque_de_france_rows)

                    beac_msg = f" + {len(new_beac_rows)} BEACCMCX091" if new_beac_rows else ""
                    other_exc_msg = f" + {len(new_other_exc_rows)} autres exceptions" if new_other_exc_rows else ""
                    forex_msg = f" + {len(new_forex_rows)} forex" if new_forex_rows else ""
                    bdf_msg = f" + {len(new_bdf_rows)} Banque de France" if new_bdf_rows else ""

                    st.success(f"✅ {total} fichier(s) Excel : {len(new_rows)} opération(s){beac_msg}{other_exc_msg}{forex_msg}{bdf_msg}")

                except Exception as e:
                    tb = traceback.format_exc()
                    errors.append(("Extraction Excel", str(e)))
                    st.error(f"❌ Erreur lors de l'extraction Excel")
                    with st.expander("Détails de l'erreur"):
                        st.code(tb)

            progress.empty()

        elif direction == "eastnet_extraction":
            # ======== MODE EASTNET RJE ========
            if eastnet_sub_mode in ("rje_incoming_mode1", "rje_outgoing_mode2"):
                _is_outgoing = (eastnet_sub_mode == "rje_outgoing_mode2")
                _mode_label = "mode 2" if _is_outgoing else "mode 1"
                _dir_label = "sortants" if _is_outgoing else "entrants"
                _dir_value = "outgoing" if _is_outgoing else "incoming"

                progress = st.progress(0, text="Initialisation...")
                total = len(uploaded_rje_files)
                idx = 0

                st.info(f"🔄 EastNet {_mode_label}: traitement de {total} fichier(s) RJE ({rje_correspondant})...")

                bic_file = ROOT / "data" / "bic_codes.xlsx"
                if not bic_file.exists():
                    bic_file = Path("data/bic_codes.xlsx")
                xlsx_bic_path = str(bic_file) if bic_file.exists() else None

                temp_rje_paths = []
                for uf in uploaded_rje_files:
                    idx += 1
                    progress.progress(idx / total, text=f"Enregistrement : {uf.name} ({idx}/{total})")
                    try:
                        tmp_path = save_uploaded_to_temp(uf)
                        tmp_dirs.append(tmp_path.parent)
                        temp_rje_paths.append(tmp_path)
                    except Exception as e:
                        errors.append((uf.name, f"Impossible d'enregistrer temporairement: {e}"))
                        st.error(f"❌ Erreur avec {uf.name}")

                if temp_rje_paths:
                    try:
                        # Mode 1 (entrants) ou Mode 2 (sortants): pré-filtrer pour
                        # n'analyser que la direction concernée + les types MT
                        # supportés par chaque direction. Accélère fortement le
                        # traitement des gros fichiers RJE multi-directions.
                        _allowed_dirs = {_dir_value}
                        _allowed_types = ({"202", "103"} if _is_outgoing
                                          else {"202", "103", "910"})
                        (new_incoming, new_outgoing, new_beac_rows, new_exc_323201,
                         new_other_exc, new_bdf_rows, new_forex_rows, _new_bdf_corr,
                         missing_codes) = eastnet_extract(
                            temp_rje_paths, correspondant=rje_correspondant,
                            xlsx_path=xlsx_bic_path,
                            allowed_directions=_allowed_dirs,
                            allowed_mt_types=_allowed_types,
                        )
                        all_missing_codes["unmapped"].update(missing_codes.get("unmapped", set()))
                        all_missing_codes["empty"].update(missing_codes.get("empty", set()))

                        def _norm_rje(r_list):
                            for r in r_list:
                                r.setdefault("beneficiaire", None)
                                r.setdefault("donneur_dordre", r.get("institution_name"))
                                r.setdefault("type_MT", None)
                                r["direction"] = _dir_value

                        _main_rows = new_outgoing if _is_outgoing else new_incoming
                        _norm_rje(_main_rows)
                        for lst in (new_beac_rows, new_exc_323201, new_other_exc, new_bdf_rows, new_forex_rows):
                            _norm_rje(lst)

                        # ── Filtre Nt.Status (CITI USD / BDF en mode 2 sortants) ─
                        # Uniquement si l'utilisateur a fourni le CSV companion EastNets.
                        nt_rejected_rows = []
                        if (
                            _is_outgoing
                            and rje_correspondant in ("CITIUS33XXX", "BDFEFRPPXXX")
                            and uploaded_eastnet_csv_files
                        ):
                            try:
                                from extractors.eastnet_extractor import (
                                    parse_eastnet_companion_csv,
                                    filter_outgoing_by_nt_status,
                                )
                                csv_index = {}
                                for cf in uploaded_eastnet_csv_files:
                                    try:
                                        tmp_csv = save_uploaded_to_temp(cf)
                                        tmp_dirs.append(tmp_csv.parent)
                                        csv_index.update(parse_eastnet_companion_csv(tmp_csv))
                                    except Exception as _e_csv:
                                        st.warning(f"⚠️ CSV {cf.name} ignoré : {_e_csv}")
                                if csv_index:
                                    _kept, _rej = filter_outgoing_by_nt_status(_main_rows, csv_index)
                                    _main_rows = _kept
                                    nt_rejected_rows = _rej
                                    if _rej:
                                        st.warning(
                                            f"🚫 Filtre Nt.Status : {len(_rej)} message(s) sortant(s) "
                                            f"écarté(s) (Nack / non acquittés / absents du CSV) — "
                                            f"voir feuille « Sortants_Rejetes_NtStatus »."
                                        )
                                    st.info(
                                        f"✅ Filtre Nt.Status appliqué : {len(_kept)} sortant(s) conservé(s) "
                                        f"sur {len(_kept) + len(_rej)}"
                                    )
                            except Exception as _e_filter:
                                tb = traceback.format_exc()
                                errors.append(("Filtre Nt.Status (CSV EastNets)", str(_e_filter)))
                                st.error(f"❌ Erreur filtre Nt.Status : {_e_filter}")
                                with st.expander("Détails de l'erreur (filtre Nt.Status)"):
                                    st.code(tb)
                        elif (
                            _is_outgoing
                            and rje_correspondant in ("CITIUS33XXX", "BDFEFRPPXXX")
                            and not uploaded_eastnet_csv_files
                        ):
                            st.info(
                                "ℹ️ Aucun CSV EastNets companion fourni : le filtre Nt.Status "
                                "n'est pas appliqué. Tous les sortants extraits du RJE sont conservés."
                            )

                        # Stocker les rejets pour le workbook
                        st.session_state['nt_status_rejected_rows'] = (
                            st.session_state.get('nt_status_rejected_rows', []) + nt_rejected_rows
                        )

                        rows.extend(_main_rows)
                        beaccmcx091_rows.extend(new_beac_rows)
                        # exception_323201 ne s'applique qu'aux entrants (MT202)
                        if not _is_outgoing:
                            exception_323201_rows.extend(new_exc_323201)
                        other_exceptions_rows.extend(new_other_exc)
                        banque_de_france_rows.extend(new_bdf_rows)
                        forex_rows.extend(new_forex_rows)

                        st.success(
                            f"✅ {total} fichier(s) RJE ({_mode_label}): "
                            f"{len(_main_rows)} {_dir_label}"
                            + (f" + {len(new_beac_rows)} BEACCMCX091" if new_beac_rows else "")
                            + (f" + {len(new_exc_323201)} exceptions 323201" if (new_exc_323201 and not _is_outgoing) else "")
                            + (f" + {len(new_other_exc)} autres exceptions" if new_other_exc else "")
                            + (f" + {len(new_bdf_rows)} Banque de France" if new_bdf_rows else "")
                            + (f" + {len(new_forex_rows)} forex" if new_forex_rows else "")
                        )

                    except Exception as e:
                        tb = traceback.format_exc()
                        errors.append(("Extraction EastNet RJE", str(e)))
                        st.error(f"❌ Erreur lors de l'extraction EastNet {_mode_label}")
                        with st.expander("Détails de l'erreur"):
                            st.code(tb)

                progress.empty()
            elif eastnet_sub_mode == "rje_transfer_analysis_mode3":
                # ======== MODE 3 EastNet RJE : Analyse Transferts Exécutés ========
                from extractor_manager import match_mt900_with_transfers, create_transfer_analysis_workbook
                from extractors.eastnet_extractor import (
                    extract_from_rje_files as _ee_extract,
                    extract_mt900_from_rje_files as _ee_mt900_extract,
                    dedupe_mt900_rows as _ee_mt900_dedupe,
                )

                total_files = len(uploaded_rje_sortants_files) + len(uploaded_rje_mt900_files)
                progress = st.progress(0, text="Initialisation...")
                idx = 0

                st.info(
                    f"🔄 EastNet mode 3 : {len(uploaded_rje_sortants_files)} fichier(s) sortants + "
                    f"{len(uploaded_rje_mt900_files)} fichier(s) MT900 ({rje_correspondant})..."
                )

                bic_file = ROOT / "data" / "bic_codes.xlsx"
                if not bic_file.exists():
                    bic_file = Path("data/bic_codes.xlsx")
                xlsx_bic_path = str(bic_file) if bic_file.exists() else None

                # Sauvegarder en temporaire
                tmp_sortants_paths = []
                for uf in uploaded_rje_sortants_files:
                    idx += 1
                    progress.progress(idx / total_files, text=f"Sortants : {uf.name} ({idx}/{total_files})")
                    try:
                        tmp_path = save_uploaded_to_temp(uf)
                        tmp_dirs.append(tmp_path.parent)
                        tmp_sortants_paths.append(tmp_path)
                    except Exception as e:
                        errors.append((uf.name, f"Impossible d'enregistrer temporairement: {e}"))
                        st.error(f"❌ Erreur avec {uf.name}")

                tmp_mt900_paths = []
                for uf in uploaded_rje_mt900_files:
                    idx += 1
                    progress.progress(idx / total_files, text=f"MT900 : {uf.name} ({idx}/{total_files})")
                    try:
                        tmp_path = save_uploaded_to_temp(uf)
                        tmp_dirs.append(tmp_path.parent)
                        tmp_mt900_paths.append(tmp_path)
                    except Exception as e:
                        errors.append((uf.name, f"Impossible d'enregistrer temporairement: {e}"))
                        st.error(f"❌ Erreur avec {uf.name}")

                if tmp_sortants_paths and tmp_mt900_paths:
                    try:
                        # 1) Extraire tous les MT202/MT103 sortants + catégories d'exception
                        # Mode 3: on n'a besoin que des sortants MT202/MT103 → pré-filtrer.
                        (new_incoming, new_outgoing, new_beac_rows, new_exc_323201,
                         new_other_exc, new_bdf_rows, new_forex_rows, _new_bdf_corr,
                         missing_codes) = _ee_extract(
                            tmp_sortants_paths, correspondant=rje_correspondant,
                            xlsx_path=xlsx_bic_path,
                            allowed_directions={"outgoing"},
                            allowed_mt_types={"202", "103"},
                        )
                        all_missing_codes["unmapped"].update(missing_codes.get("unmapped", set()))
                        all_missing_codes["empty"].update(missing_codes.get("empty", set()))

                        # Tous les transferts sortants (incl. BEAC, BDF, forex, exceptions)
                        # sont considérés comme candidats au matching MT900.
                        # Chaque ligne est taguée avec sa catégorie d'origine pour que
                        # create_transfer_analysis_workbook puisse les router vers
                        # la bonne feuille (main / BEACCMCX091 / Exceptions_323201 / ...)
                        def _tag(rows_list, cat):
                            for _r in rows_list:
                                _r["_category"] = cat
                                _r.setdefault("direction", "outgoing")
                            return list(rows_list)

                        all_transfers = (
                            _tag(new_outgoing, "main")
                            + _tag(new_beac_rows, "beaccmcx091")
                            + _tag(new_exc_323201, "exception_323201")
                            + _tag(new_other_exc, "other_exception")
                            + _tag(new_bdf_rows, "banque_de_france")
                            + _tag(new_forex_rows, "forex")
                            + _tag(_new_bdf_corr, "bdf_corr_exception")
                        )

                        # 2) Extraire tous les MT900
                        all_mt900 = _ee_mt900_extract(
                            tmp_mt900_paths, correspondant=rje_correspondant,
                            xlsx_path=xlsx_bic_path
                        )

                        # 2bis) Dédoublonner les MT900 retransmis par SWIFT FIN
                        # (trailer {DLM:}, même MIR). Les copies écartées sont
                        # listées dans la feuille MT900_Doublons du workbook.
                        all_mt900, duplicate_mt900 = _ee_mt900_dedupe(all_mt900)
                        st.session_state['duplicate_mt900_rows'] = duplicate_mt900
                        if duplicate_mt900:
                            st.warning(
                                f"⚠️ {len(duplicate_mt900)} MT900 doublon(s) réseau (trailer DLM / "
                                f"même MIR) écarté(s) avant matching — voir feuille MT900_Doublons."
                            )

                        # 3) Matcher MT900 ↔ transferts via F21 = F20
                        matched_mt900, suspens, exception_mt900, unmatched_mt900 = (
                            match_mt900_with_transfers(all_mt900, all_transfers)
                        )

                        rows.extend(matched_mt900)
                        st.session_state['suspens_rows'] = suspens
                        st.session_state['exception_mt900_rows'] = exception_mt900
                        st.session_state['unmatched_mt900_rows'] = unmatched_mt900

                        st.success(
                            f"✅ {len(all_mt900)} MT900 extraits, {len(all_transfers)} transferts sortants → "
                            f"{len(matched_mt900)} matchés, {len(unmatched_mt900)} MT900 non rapprochés, "
                            f"{len(suspens)} en suspens, {len(exception_mt900)} en exception"
                        )

                    except Exception as e:
                        tb = traceback.format_exc()
                        errors.append(("Extraction EastNet mode 3", str(e)))
                        st.error("❌ Erreur lors de l'extraction EastNet mode 3")
                        with st.expander("Détails de l'erreur"):
                            st.code(tb)

                progress.empty()
            else:
                total_files = len(uploaded_rje_mt950_files) + len(uploaded_rje_msg_files)
                progress = st.progress(0, text="Initialisation...")
                idx = 0

                msg_direction = "incoming" if mt950_sub_mode == "entrants" else "outgoing"
                st.info(
                    f"🔄 EastNet mode 4: {len(uploaded_rje_mt950_files)} fichier(s) MT950 RJE "
                    f"et {len(uploaded_rje_msg_files)} fichier(s) messages RJE ({mt950_sub_mode})..."
                )

                bic_file = ROOT / "data" / "bic_codes.xlsx"
                if not bic_file.exists():
                    bic_file = Path("data/bic_codes.xlsx")
                xlsx_bic_path = str(bic_file) if bic_file.exists() else None

                # Étape 1: extraire F61 depuis les RJE MT950
                temp_mt950_paths = []
                for uf in uploaded_rje_mt950_files:
                    idx += 1
                    progress.progress(idx / total_files, text=f"MT950 RJE : {uf.name} ({idx}/{total_files})")
                    try:
                        tmp_path = save_uploaded_to_temp(uf)
                        tmp_dirs.append(tmp_path.parent)
                        temp_mt950_paths.append(tmp_path)
                    except Exception as e:
                        errors.append((uf.name, f"Impossible d'enregistrer temporairement: {e}"))
                        st.error(f"❌ Erreur avec {uf.name}")

                all_f61_entries = extract_mt950_entries_from_rje_files(temp_mt950_paths)
                target_cd = "C" if mt950_sub_mode == "entrants" else "D"
                n_target = sum(1 for e in all_f61_entries if e.get("cd") == target_cd)
                st.success(f"✅ MT950 RJE: {len(all_f61_entries)} F61 extraits ({n_target} {target_cd})")

                # Étape 2: extraire les messages RJE et classer comme mode 4 existant
                all_summary_rows = []
                all_beac_rows = []
                all_exc_323201_rows = []
                all_other_exc_rows = []
                all_bdf_rows = []
                all_forex_rows_mt950 = []
                all_bdf_corr_exc_rows_mt950 = []

                for uf in uploaded_rje_msg_files:
                    idx += 1
                    progress.progress(idx / total_files, text=f"Messages RJE : {uf.name} ({idx}/{total_files})")
                    try:
                        tmp_path = save_uploaded_to_temp(uf)
                        tmp_dirs.append(tmp_path.parent)
                        (new_incoming, new_outgoing, new_beac_rows, new_exc_rows,
                         new_other_exc_rows, new_bdf_rows, new_forex_rows, new_bdf_corr_exc_rows,
                         missing) = eastnet_extract(
                            [tmp_path], correspondant=rje_correspondant, xlsx_path=xlsx_bic_path,
                            allowed_directions={msg_direction},
                            allowed_mt_types=({"202", "103"} if msg_direction == "outgoing"
                                              else {"202", "103", "910"}),
                        )
                        all_missing_codes["unmapped"].update(missing.get("unmapped", set()))
                        all_missing_codes["empty"].update(missing.get("empty", set()))

                        msg_rows = new_incoming if msg_direction == "incoming" else new_outgoing

                        def _normalize_rows(row_list, source_name):
                            for r in row_list:
                                r.setdefault("beneficiaire", None)
                                r.setdefault("donneur_dordre", r.get("institution_name") or None)
                                r.setdefault("type_MT", None)
                                if not r.get("source_pdf"):
                                    r["source_pdf"] = source_name

                        def _filter_valid(rlist):
                            return [r for r in rlist if r.get("type_MT") and any(t in r.get("type_MT") for t in ("202", "103", "910"))]

                        _normalize_rows(msg_rows, uf.name)
                        _normalize_rows(new_beac_rows, uf.name)
                        _normalize_rows(new_exc_rows, uf.name)
                        _normalize_rows(new_other_exc_rows, uf.name)
                        _normalize_rows(new_bdf_rows, uf.name)
                        _normalize_rows(new_forex_rows, uf.name)
                        _normalize_rows(new_bdf_corr_exc_rows, uf.name)

                        all_summary_rows.extend(_filter_valid(msg_rows))
                        all_beac_rows.extend(_filter_valid(new_beac_rows))
                        all_exc_323201_rows.extend(_filter_valid(new_exc_rows))
                        all_other_exc_rows.extend(_filter_valid(new_other_exc_rows))
                        all_bdf_rows.extend(_filter_valid(new_bdf_rows))
                        all_forex_rows_mt950.extend(_filter_valid(new_forex_rows))
                        all_bdf_corr_exc_rows_mt950.extend(_filter_valid(new_bdf_corr_exc_rows))
                    except Exception as e:
                        tb = traceback.format_exc()
                        errors.append((uf.name, str(e)))
                        st.error(f"❌ Erreur: {uf.name}")
                        with st.expander(f"Détails de l'erreur pour {uf.name}"):
                            st.code(tb)

                progress.progress(1.0, text="Rapprochement en cours...")
                from extractors.mt950 import match_f61_with_messages as match_f61

                categories = {
                    "summary": all_summary_rows,
                    "BEACCMCX091": all_beac_rows,
                    "Exceptions_323201": all_exc_323201_rows,
                    "Autres_Exceptions": all_other_exc_rows,
                    "BANQUE DE FRANCE": all_bdf_rows,
                    "forex": all_forex_rows_mt950,
                    "Exceptions_Correspondants": all_bdf_corr_exc_rows_mt950,
                }

                all_msg_pool = []
                msg_to_category = {}
                for cat_name, cat_rows in categories.items():
                    for r in cat_rows:
                        all_msg_pool.append(r)
                        msg_to_category[id(r)] = cat_name

                rapproches_raw, non_rap_msg_raw, non_rap_f61 = match_f61(
                    all_f61_entries, all_msg_pool, sub_mode=mt950_sub_mode
                )

                rapproches_by_cat = {cat: [] for cat in categories}
                for rap in rapproches_raw:
                    matched_msg_ref = rap.get("msg_reference")
                    matched_msg_src = rap.get("msg_source_pdf")
                    matched_msg_type = rap.get("msg_type_MT")
                    matched_msg_montant = rap.get("msg_montant")
                    found_cat = "summary"
                    for msg in all_msg_pool:
                        if (msg.get("reference") == matched_msg_ref and
                            msg.get("source_pdf") == matched_msg_src and
                            msg.get("type_MT") == matched_msg_type and
                            msg.get("montant") == matched_msg_montant):
                            found_cat = msg_to_category.get(id(msg), "summary")
                            break
                    rapproches_by_cat[found_cat].append(rap)

                non_rap_msg_by_cat = {cat: [] for cat in categories}
                for msg in non_rap_msg_raw:
                    cat = msg_to_category.get(id(msg), "summary")
                    non_rap_msg_by_cat[cat].append(msg)

                rows = rapproches_raw
                st.session_state.mt950_rapproches_by_cat = rapproches_by_cat
                st.session_state.mt950_non_rap_msg_by_cat = non_rap_msg_by_cat
                st.session_state.mt950_non_rap_f61 = non_rap_f61
                st.session_state.mt950_sub_mode = mt950_sub_mode
                st.session_state.mt950_categories = categories

                progress.empty()
                st.info(
                    f"📊 **Résumé** : {len(rapproches_raw)} rapprochés, "
                    f"{len(non_rap_msg_raw)} messages non rapprochés, {len(non_rap_f61)} F61 non rapprochés"
                )

        else:
            # ======== MODE STANDARD (incoming/outgoing) ========
            progress = st.progress(0, text="Initialisation...")
            total = len(uploaded_files)
            idx = 0
            
            st.info(f"🔄 Traitement de {total} fichier(s)...")
            
            for uf in uploaded_files:
                idx += 1
                progress.progress(idx / total, text=f"Traitement : {uf.name} ({idx}/{total})")
                
                try:
                    tmp_path = save_uploaded_to_temp(uf)
                    tmp_dirs.append(tmp_path.parent)
                except Exception as e:
                    errors.append((uf.name, f"Impossible d'enregistrer temporairement: {e}"))
                    st.error(f"❌ Erreur avec {uf.name}")
                    continue
                
                try:
                    # Mode standard (incoming/outgoing)
                    # extract_dispatch retourne (rows, beaccmcx091_rows, exception_323201_rows, other_exceptions_rows, banque_de_france_rows, forex_rows, bdf_corr_exception_rows, missing_codes)
                    new_rows, new_beac_rows, new_exc_rows, new_other_exc_rows, new_bdf_rows, new_forex_rows, new_bdf_corr_exc_rows, missing_codes = extract_dispatch(tmp_path, direction=direction)
                    
                    # Accumulate missing codes
                    all_missing_codes["unmapped"].update(missing_codes.get("unmapped", set()))
                    all_missing_codes["empty"].update(missing_codes.get("empty", set()))
                    
                    # Normalisations pour les rows normales
                    for r in new_rows:
                        if "beneficiaire" not in r:
                            r["beneficiaire"] = None
                        
                        if "donneur_dordre" not in r:
                            if "institution_name" in r and r["institution_name"]:
                                r["donneur_dordre"] = r.get("institution_name")
                            else:
                                r["donneur_dordre"] = None
                        
                        if not r.get("source_pdf"):
                            r["source_pdf"] = uf.name
                        
                        if not r.get("type_MT"):
                            r["type_MT"] = None
                        
                        rows.append(r)
                    
                    # Normalisations pour les BEACCMCX091
                    for r in new_beac_rows:
                        if "beneficiaire" not in r:
                            r["beneficiaire"] = None
                        
                        if "donneur_dordre" not in r:
                            if "institution_name" in r and r["institution_name"]:
                                r["donneur_dordre"] = r.get("institution_name")
                            else:
                                r["donneur_dordre"] = None
                        
                        if not r.get("source_pdf"):
                            r["source_pdf"] = uf.name
                        
                        if not r.get("type_MT"):
                            r["type_MT"] = None
                        
                        beaccmcx091_rows.append(r)
                    
                    # Normalisations pour les exceptions 323201
                    for r in new_exc_rows:
                        if "beneficiaire" not in r:
                            r["beneficiaire"] = None
                        
                        if "donneur_dordre" not in r:
                            if "institution_name" in r and r["institution_name"]:
                                r["donneur_dordre"] = r.get("institution_name")
                            else:
                                r["donneur_dordre"] = None
                        
                        if not r.get("source_pdf"):
                            r["source_pdf"] = uf.name
                        
                        if not r.get("type_MT"):
                            r["type_MT"] = None
                        
                        exception_323201_rows.append(r)
                    
                    # Normalisations pour les autres exceptions (EUR/nivellement)
                    for r in new_other_exc_rows:
                        if "beneficiaire" not in r:
                            r["beneficiaire"] = None
                        
                        if "donneur_dordre" not in r:
                            if "institution_name" in r and r["institution_name"]:
                                r["donneur_dordre"] = r.get("institution_name")
                            else:
                                r["donneur_dordre"] = None
                        
                        if not r.get("source_pdf"):
                            r["source_pdf"] = uf.name
                        
                        if not r.get("type_MT"):
                            r["type_MT"] = None
                        
                        other_exceptions_rows.append(r)
                    
                    # Normalisations pour les BANQUE DE FRANCE
                    for r in new_bdf_rows:
                        if "beneficiaire" not in r:
                            r["beneficiaire"] = None
                        
                        if "donneur_dordre" not in r:
                            if "institution_name" in r and r["institution_name"]:
                                r["donneur_dordre"] = r.get("institution_name")
                            else:
                                r["donneur_dordre"] = None
                        
                        if not r.get("source_pdf"):
                            r["source_pdf"] = uf.name
                        
                        if not r.get("type_MT"):
                            r["type_MT"] = None
                        
                        banque_de_france_rows.append(r)
                    
                    # Normalisations pour les forex rows
                    for r in new_forex_rows:
                        if "beneficiaire" not in r:
                            r["beneficiaire"] = None
                        
                        if "donneur_dordre" not in r:
                            if "institution_name" in r and r["institution_name"]:
                                r["donneur_dordre"] = r.get("institution_name")
                            else:
                                r["donneur_dordre"] = None
                        
                        if not r.get("source_pdf"):
                            r["source_pdf"] = uf.name
                        
                        if not r.get("type_MT"):
                            r["type_MT"] = None
                        
                        forex_rows.append(r)
                    
                    # Normalisations pour les exceptions correspondants BdF
                    for r in new_bdf_corr_exc_rows:
                        if "beneficiaire" not in r:
                            r["beneficiaire"] = None
                        
                        if "donneur_dordre" not in r:
                            if "institution_name" in r and r["institution_name"]:
                                r["donneur_dordre"] = r.get("institution_name")
                            else:
                                r["donneur_dordre"] = None
                        
                        if not r.get("source_pdf"):
                            r["source_pdf"] = uf.name
                        
                        if not r.get("type_MT"):
                            r["type_MT"] = None
                        
                        bdf_corr_exception_rows.append(r)
                    
                    types = sorted({rr.get("type_MT") or "inconnu" for rr in new_rows})
                    beac_msg = f" + {len(new_beac_rows)} BEACCMCX091" if new_beac_rows else ""
                    exc_msg = f" + {len(new_exc_rows)} exceptions 323201" if new_exc_rows else ""
                    other_exc_msg = f" + {len(new_other_exc_rows)} autres exceptions" if new_other_exc_rows else ""
                    bdf_msg = f" + {len(new_bdf_rows)} BANQUE DE FRANCE" if new_bdf_rows else ""
                    forex_msg = f" + {len(new_forex_rows)} forex" if new_forex_rows else ""
                    
                    st.success(f"✅ {uf.name}: {len(new_rows)} message(s){beac_msg}{exc_msg}{other_exc_msg}{bdf_msg}{forex_msg} — Types: {', '.join(types)}")
                    
                except Exception as e:
                    tb = traceback.format_exc()
                    errors.append((uf.name, str(e)))
                    st.error(f"❌ Erreur: {uf.name}")
                    with st.expander(f"Détails de l'erreur pour {uf.name}"):
                        st.code(tb)
            
            progress.empty()
        
        # Cleanup
        for d in tmp_dirs:
            try:
                shutil.rmtree(d, ignore_errors=True)
            except Exception:
                pass
        
        # Filter by date range (plage de dates) — sauf mode MT950 reconciliation et mode 3 (transfer_analysis)
        if not (
            direction == "mt950_reconciliation"
            or direction == "transfer_analysis"
            or (direction == "eastnet_extraction" and eastnet_sub_mode == "rje_mt950_mode4")
            or (direction == "eastnet_extraction" and eastnet_sub_mode == "rje_transfer_analysis_mode3")
        ) and date_debut and date_fin and rows and date_debut <= date_fin:
            date_debut_str = date_debut.strftime("%Y-%m-%d")
            date_fin_str = date_fin.strftime("%Y-%m-%d")
            rows_filtered = [r for r in rows if r.get("date_reference") and date_debut_str <= r.get("date_reference") <= date_fin_str]
            if len(rows_filtered) < len(rows):
                if date_debut_str == date_fin_str:
                    st.info(f"📅 Filtrage: {len(rows_filtered)} message(s) pour {date_debut_str} (sur {len(rows)})")
                else:
                    st.info(f"📅 Filtrage: {len(rows_filtered)} message(s) du {date_debut_str} au {date_fin_str} (sur {len(rows)})")
            rows = rows_filtered
        
        # Stocker les résultats si disponibles
        # On considère qu'il y a un résultat dès qu'au moins une des listes
        # (summary ou exceptions) contient des messages — sinon, le cas
        # "tout en BEACCMCX091" (ou tout en exception) ne déclencherait
        # jamais la création du workbook.
        _has_any_result = bool(
            rows
            or beaccmcx091_rows
            or exception_323201_rows
            or other_exceptions_rows
            or banque_de_france_rows
            or forex_rows
            or bdf_corr_exception_rows
        )
        if _has_any_result:
            # Ensure backward-compatibility (sauf mode MT950 où rows = rapprochés)
            if not (
                direction == "mt950_reconciliation"
                or (direction == "eastnet_extraction" and eastnet_sub_mode == "rje_mt950_mode4")
            ):
                for r in rows:
                    if not r.get("institution_name"):
                        r["institution_name"] = r.get("donneur_dordre")
            
            # Create and download workbook
            
            temp_outdir = Path(tempfile.mkdtemp(prefix="swift_out_"))
            # ── Indicateur visuel : la construction du workbook peut être longue
            # sur les gros volumes (~1-2 min pour 30k+ lignes). On affiche un
            # placeholder informatif qui sera effacé après création.
            _wb_total_rows = (
                len(rows or []) + len(beaccmcx091_rows or [])
                + len(exception_323201_rows or [])
                + len(other_exceptions_rows or [])
                + len(banque_de_france_rows or [])
                + len(forex_rows or [])
                + len(bdf_corr_exception_rows or [])
                + len(st.session_state.get('nt_status_rejected_rows', []) or [])
            )
            _wb_placeholder = st.empty()
            if _wb_total_rows >= 5000:
                _wb_placeholder.info(
                    f"📊 Construction du fichier Excel en cours ({_wb_total_rows:,} lignes)… "
                    "cette étape peut prendre 1 à 2 minutes pour les gros volumes. Merci de patienter."
                )
            else:
                _wb_placeholder.info("📊 Construction du fichier Excel en cours…")
            try:
                if direction == "transfer_analysis" or (
                    direction == "eastnet_extraction"
                    and eastnet_sub_mode == "rje_transfer_analysis_mode3"
                ):
                    # Mode analyse des transferts - utiliser create_transfer_analysis_workbook
                    from extractor_manager import create_transfer_analysis_workbook
                    suspens_rows = st.session_state.get('suspens_rows', [])
                    exception_mt900_rows = st.session_state.get('exception_mt900_rows', [])
                    unmatched_mt900_rows = st.session_state.get('unmatched_mt900_rows', [])
                    duplicate_mt900_rows = st.session_state.get('duplicate_mt900_rows', [])
                    
                    # Convertir les dates en string pour le workbook
                    date_start_str = date_debut.strftime("%Y-%m-%d") if date_debut else None
                    date_end_str = date_fin.strftime("%Y-%m-%d") if date_fin else None
                    
                    # Localiser bic_codes.xlsx pour enrichir les MT900 via la
                    # colonne Reglement (4 premiers caractères de related_reference).
                    bic_file = ROOT / "data" / "bic_codes.xlsx"
                    if not bic_file.exists():
                        bic_file = Path("data/bic_codes.xlsx")

                    out_path = create_transfer_analysis_workbook(
                        rows, suspens_rows, exception_mt900_rows, temp_outdir, 
                        date_start=date_start_str, date_end=date_end_str,
                        unmatched_mt900_rows=unmatched_mt900_rows,
                        xlsx_path=str(bic_file) if bic_file.exists() else None,
                        duplicate_mt900_rows=duplicate_mt900_rows,
                    )
                    
                    # Afficher le résumé
                    st.info(f"📊 **Résumé** : {len(rows)} transfert(s) exécuté(s), {len(unmatched_mt900_rows)} MT900 non matchés, {len(suspens_rows)} en suspens, {len(exception_mt900_rows)} en exception")
                elif direction == "mt950_reconciliation" or (direction == "eastnet_extraction" and eastnet_sub_mode == "rje_mt950_mode4"):
                    # Mode rapprochement MT950 v2 — feuilles séparées par catégorie
                    rapproches_by_cat = st.session_state.get('mt950_rapproches_by_cat', {})
                    non_rap_msg_by_cat = st.session_state.get('mt950_non_rap_msg_by_cat', {})
                    non_rap_f61 = st.session_state.get('mt950_non_rap_f61', [])
                    sub_mode_val = st.session_state.get('mt950_sub_mode', 'entrants')
                    
                    date_start_str = date_debut.strftime("%Y-%m-%d") if date_debut else None
                    date_end_str = date_fin.strftime("%Y-%m-%d") if date_fin else None
                    
                    out_path = create_mt950_reconciliation_workbook_v2(
                        rapproches_by_cat, non_rap_msg_by_cat, non_rap_f61, temp_outdir,
                        sub_mode=sub_mode_val,
                        date_start=date_start_str, date_end=date_end_str,
                    )
                else:
                    # Tous les correspondants: deux fichiers séparés (entrants + sortants)
                    _is_eastnet_single = (
                        direction == "eastnet_extraction"
                        and eastnet_sub_mode in ("rje_incoming_mode1", "rje_outgoing_mode2")
                    )
                    if direction in ("excel_extraction", "eastnet_extraction"):
                        if _is_eastnet_single:
                            _single_dir = (
                                "outgoing" if eastnet_sub_mode == "rje_outgoing_mode2" else "incoming"
                            )
                            out_path = create_workbook(
                                rows, temp_outdir, direction=_single_dir,
                                beaccmcx091_rows=beaccmcx091_rows,
                                exception_323201_rows=exception_323201_rows if _single_dir == "incoming" else None,
                                other_exceptions_rows=other_exceptions_rows,
                                banque_de_france_rows=banque_de_france_rows,
                                forex_rows=forex_rows,
                                bdf_corr_exception_rows=[],
                                nt_status_rejected_rows=st.session_state.get('nt_status_rejected_rows', []) if _single_dir == "outgoing" else None,
                            )
                        elif direction == "eastnet_extraction":
                            # Récupérer les listes entrants/sortants stockées en session
                            rows_in = st.session_state.get("rje_incoming_rows", [])
                            rows_out = st.session_state.get("rje_outgoing_rows", [])
                            # Les catégories d'exception ne sont pas séparées par direction dans eastnet
                            # → on les met toutes côté entrant pour la cohérence
                            beac_in, beac_out = beaccmcx091_rows, []
                            exc_in, exc_out = other_exceptions_rows, []
                            forex_in, forex_out = forex_rows, []
                            bdf_in, bdf_out = banque_de_france_rows, []
                            bdf_corr_in, bdf_corr_out = bdf_corr_exception_rows, []
                            exc_323201_in = exception_323201_rows
                        else:
                            def _split_by_direction(row_list):
                                incoming = [r for r in row_list if r.get("direction") == "incoming"]
                                outgoing = [r for r in row_list if r.get("direction") == "outgoing"]
                                return incoming, outgoing
                            rows_in, rows_out = _split_by_direction(rows)
                            beac_in, beac_out = _split_by_direction(beaccmcx091_rows)
                            exc_in, exc_out = _split_by_direction(other_exceptions_rows)
                            forex_in, forex_out = _split_by_direction(forex_rows)
                            bdf_in, bdf_out = _split_by_direction(banque_de_france_rows)
                            bdf_corr_in, bdf_corr_out = _split_by_direction(bdf_corr_exception_rows)
                            exc_323201_in = None
                        
                        if _is_eastnet_single:
                            dir_label_in = None
                            dir_label_out = None
                        else:
                            dir_label_in = "eastnet_entrant" if direction == "eastnet_extraction" else "excel_extraction_entrant"
                            dir_label_out = "eastnet_sortant" if direction == "eastnet_extraction" else "excel_extraction_sortant"
                        
                        if not _is_eastnet_single:
                            # Fichier entrants
                            out_path_in = create_workbook(
                                rows_in, temp_outdir, direction=dir_label_in,
                                beaccmcx091_rows=beac_in,
                                exception_323201_rows=exc_323201_in,
                                other_exceptions_rows=exc_in,
                                banque_de_france_rows=bdf_in, forex_rows=forex_in,
                                bdf_corr_exception_rows=bdf_corr_in
                            )
                            # Fichier sortants
                            out_path_out = create_workbook(
                                rows_out, temp_outdir, direction=dir_label_out,
                                beaccmcx091_rows=beac_out, exception_323201_rows=None,
                                other_exceptions_rows=exc_out,
                                banque_de_france_rows=bdf_out, forex_rows=forex_out,
                                bdf_corr_exception_rows=bdf_corr_out
                            )
                        
                        if not _is_eastnet_single:
                            # Lire les deux fichiers
                            import zipfile, io
                            zip_buffer = io.BytesIO()
                            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
                                zf.write(out_path_in, out_path_in.name)
                                zf.write(out_path_out, out_path_out.name)
                            zip_buffer.seek(0)
                        
                        if not _is_eastnet_single:
                            # Stocker les deux fichiers individuellement pour le téléchargement
                            with open(out_path_in, "rb") as f:
                                excel_data_in = f.read()
                            with open(out_path_out, "rb") as f:
                                excel_data_out = f.read()
                        
                        if not _is_eastnet_single:
                            _corr_label = rje_correspondant if direction == "eastnet_extraction" else excel_correspondant
                            st.session_state.excel_data = zip_buffer.getvalue()
                            st.session_state.excel_filename = f"swift_extraction_{_corr_label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
                            st.session_state.excel_data_entrant = excel_data_in
                            st.session_state.excel_filename_entrant = out_path_in.name
                            st.session_state.excel_data_sortant = excel_data_out
                            st.session_state.excel_filename_sortant = out_path_out.name
                            out_path = out_path_in  # Pour la compatibilité
                    else:
                        out_path = create_workbook(rows, temp_outdir, direction=direction, beaccmcx091_rows=beaccmcx091_rows, exception_323201_rows=exception_323201_rows, other_exceptions_rows=other_exceptions_rows, banque_de_france_rows=banque_de_france_rows, forex_rows=forex_rows, bdf_corr_exception_rows=bdf_corr_exception_rows)
                
                # Lecture du fichier output (si pas déjà fait pour le cas CITI 2 fichiers)
                if not st.session_state.get('excel_data'):
                    with open(out_path, "rb") as f:
                        excel_data = f.read()
                    
                    # Stocker les données dans session_state pour persister après téléchargement
                    st.session_state.excel_data = excel_data
                    st.session_state.excel_filename = out_path.name
                
                st.session_state.extraction_results = {
                    'rows': rows,
                    'beaccmcx091_rows': beaccmcx091_rows,
                    'exception_323201_rows': exception_323201_rows,
                    'all_missing_codes': all_missing_codes,
                    'direction': direction
                }
                
                # Construction terminée : effacer le message d'attente
                try:
                    _wb_placeholder.empty()
                except Exception:
                    pass

            except Exception as e:
                try:
                    _wb_placeholder.empty()
                except Exception:
                    pass
                st.error(f"❌ Impossible de créer le workbook: {e}")
            finally:
                try:
                    shutil.rmtree(temp_outdir, ignore_errors=True)
                except Exception:
                    pass
        
        else:
            st.warning("⚠️ Aucun résultat extrait. Vérifiez le format des fichiers.")
        
        # Show errors
        if errors:
            st.markdown("---")
            st.markdown("### ❌ Erreurs rencontrées")
            for name, msg in errors:
                st.error(f"**{name}** : {msg}")
        
        if st.session_state.excel_data:
            st.success("✅ Extraction terminée ! Les résultats sont affichés ci-dessous.")

# Afficher les résultats stockés (persiste après clic sur téléchargement)
if st.session_state.extraction_results is not None and st.session_state.excel_data is not None:
    st.markdown("---")
    st.markdown("### 📊 Résultats de l'extraction")
    
    results = st.session_state.extraction_results
    rows = results['rows']
    beaccmcx091_rows = results.get('beaccmcx091_rows', [])
    exception_323201_rows = results.get('exception_323201_rows', [])
    all_missing_codes = results.get('all_missing_codes', {"unmapped": set(), "empty": set()})
    direction = results.get('direction', 'incoming')
    
    # Afficher le tableau des résultats
    display_rows = []
    for r in rows:
        display_rows.append({
            "Correspondant": r.get("correspondant"),
            "Date Référence": r.get("date_reference"),
            "Référence": r.get("reference"),
            "Réf. Origine": r.get("reference_origine"),
            "Type MT": r.get("type_MT"),
            "Pays ISO3": r.get("pays_iso3"),
            "Code Donneur": r.get("code_donneur_dordre"),
            "Donneur d'ordre": r.get("donneur_dordre"),
            "Bénéficiaire": r.get("beneficiaire"),
            "Montant": r.get("montant"),
            "Devise": r.get("devise"),
            "Commentaires": r.get("commentaires"),
            "Source PDF": r.get("source_pdf")
        })
    
    df = pd.DataFrame(display_rows)
    
    # Format montant
    if "Montant" in df.columns:
        try:
            df["Montant"] = df["Montant"].apply(
                lambda x: ("{:,}".format(x)).replace(",", " ") if pd.notnull(x) else x
            )
        except Exception:
            pass
    
    st.dataframe(df, use_container_width=True, height=400)
    
    # Afficher BEACCMCX091 si présents
    if beaccmcx091_rows:
        st.markdown("---")
        st.markdown("### 🔴 Messages BEACCMCX091 (stockés séparément)")
        st.info(f"📌 {len(beaccmcx091_rows)} message(s) BEACCMCX091 détecté(s) et stocké(s) dans une feuille séparée de l'Excel")
        
        beac_display_rows = []
        for r in beaccmcx091_rows:
            beac_display_rows.append({
                "Correspondant": r.get("correspondant"),
                "Date Référence": r.get("date_reference"),
                "Référence": r.get("reference"),
                "Type MT": r.get("type_MT"),
                "Montant": r.get("montant"),
                "Devise": r.get("devise"),
                "Source PDF": r.get("source_pdf")
            })
        df_beac = pd.DataFrame(beac_display_rows)
        st.dataframe(df_beac, use_container_width=True, height=200)
    
    # Afficher exceptions 323201 si présentes
    if exception_323201_rows:
        st.markdown("---")
        st.markdown("### 🟠 Exceptions 323201 (stockées séparément)")
        st.info(f"📌 {len(exception_323201_rows)} message(s) avec 323201 dans F58A détecté(s)")
        
        exc_display_rows = []
        for r in exception_323201_rows:
            exc_display_rows.append({
                "Correspondant": r.get("correspondant"),
                "Date Référence": r.get("date_reference"),
                "Référence": r.get("reference"),
                "Type MT": r.get("type_MT"),
                "Montant": r.get("montant"),
                "Devise": r.get("devise"),
                "Source PDF": r.get("source_pdf")
            })
        df_exc = pd.DataFrame(exc_display_rows)
        st.dataframe(df_exc, use_container_width=True, height=200)
    
    # Statistics
    st.markdown("---")
    st.markdown("### 📈 Statistiques")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Messages extraits", len(rows))
    with col2:
        types_list = [r.get("type_MT") for r in rows if r.get("type_MT")]
        types_count = len(set(types_list))
        st.metric("Types de messages", types_count)
    with col3:
        sources_count = len(set(r.get("source_pdf") for r in rows if r.get("source_pdf")))
        st.metric("Fichiers sources", sources_count)
    
    # Afficher les types de messages détectés
    if types_list:
        types_summary = {}
        for t in types_list:
            types_summary[t] = types_summary.get(t, 0) + 1
        st.caption("**Types détectés:** " + ", ".join([f"{k} ({v})" for k, v in sorted(types_summary.items())]))
    
    # Missing codes display
    if all_missing_codes.get("unmapped") or all_missing_codes.get("empty"):
        st.markdown("---")
        st.markdown("### 📋 Codes BIC manquants")
        
        col1, col2 = st.columns(2)
        
        if all_missing_codes.get("unmapped"):
            with col1:
                st.warning(f"**{len(all_missing_codes['unmapped'])} codes non mappés**")
                st.caption("Codes trouvés dans le PDF mais sans nom de banque mapping :")
                for code in sorted(all_missing_codes["unmapped"]):
                    st.code(code)
        
        if all_missing_codes.get("empty"):
            with col2:
                st.error(f"**{len(all_missing_codes['empty'])} codes vides**")
                st.caption("Champs BIC complètement vides dans le PDF :")
                for code in sorted(all_missing_codes["empty"]):
                    st.code(code)
        
        # Form to add missing BIC codes
        st.markdown("#### ➕ Ajouter un nouveau code BIC")
        
        st.warning("""
            ⚠️ **Important** : Sur Hugging Face Spaces gratuit, les modifications sont **temporaires** 
            et seront perdues au redémarrage de l'application.
            
            **Pour ajouter des codes de manière permanente :**
            - Remplissez le formulaire ci-dessous pour tester immédiatement
            - Ou cliquez sur le bouton "📧 Envoyer par email" pour nous transmettre les codes
        """)
        
        with st.form("add_bic_form"):
            col1, col2 = st.columns(2)
            with col1:
                new_code = st.text_input("Code BIC (8-11 caractères)", max_chars=11, placeholder="Ex: ECOCMLCX")
                new_country = st.text_input("Code ISO3 du pays (3 lettres)", max_chars=3, placeholder="Ex: CMR")
            with col2:
                new_name = st.text_input("Nom de la banque", max_chars=100, placeholder="Ex: ECOBANK CAMEROUN")
                new_reglement = st.text_input("Code de règlement (4 chiffres, optionnel)", max_chars=4, placeholder="Ex: 8101")
            
            col_btn1, col_btn2 = st.columns([3, 2])
            with col_btn1:
                submitted = st.form_submit_button("✅ Ajouter temporairement", use_container_width=True, type="primary")
            with col_btn2:
                if new_code and new_name and new_country:
                    email_subject = "Codes BIC manquants - SWIFT Extractor"
                    email_body = f"""Bonjour,

Je souhaite ajouter un nouveau code BIC à la base de données :

Code BIC : {new_code.upper().strip()}
Nom de la banque : {new_name.strip()}
Pays (ISO3) : {new_country.upper().strip()}
Code de règlement : {new_reglement.strip() if new_reglement else 'Non renseigné'}

Merci de l'ajouter de manière permanente.

Cordialement"""
                    
                    import urllib.parse
                    mailto_link = f"mailto:alannmatistiabou@gmail.com?subject={urllib.parse.quote(email_subject)}&body={urllib.parse.quote(email_body)}"
                    st.markdown(f'<a href="{mailto_link}" target="_blank"><button style="width:100%;padding:0.5rem;background-color:#4CAF50;color:white;border:none;border-radius:0.3rem;cursor:pointer;">📧 Envoyer par email</button></a>', unsafe_allow_html=True)
                else:
                    st.form_submit_button("📧 Envoyer par email", use_container_width=True, disabled=True, help="Remplissez d'abord tous les champs obligatoires")
            
            if submitted:
                if not new_code or not new_name or not new_country:
                    st.error("⚠️ Code BIC, Nom et Pays sont obligatoires")
                elif len(new_code) < 8 or len(new_code) > 11:
                    st.error("⚠️ Le code BIC doit contenir entre 8 et 11 caractères")
                elif len(new_country) != 3:
                    st.error("⚠️ Le code pays doit contenir exactement 3 lettres (ISO3)")
                else:
                    try:
                        from extractors.bic_utils import add_bic_code_to_xlsx
                        
                        bic_file = Path("data/bic_codes.xlsx")
                        if not bic_file.exists():
                            bic_file = ROOT / "data" / "bic_codes.xlsx"
                        
                        add_bic_code_to_xlsx(new_code.upper().strip(), new_name.strip(), new_country.upper().strip(), str(bic_file), new_reglement.strip() if new_reglement else None)
                        st.success(f"✅ Code **{new_code.upper()}** ajouté temporairement !")
                        st.info("💡 Rechargez la page pour utiliser ce code. ⚠️ Il sera perdu au redémarrage du Space.")
                    except Exception as e:
                        st.error(f"❌ Erreur lors de l'ajout du code : {e}")
    
    # Bouton de téléchargement persistant
    st.markdown("---")
    st.markdown("### 💾 Téléchargement")
    
    # Si on a deux fichiers (entrants + sortants), proposer deux boutons
    if st.session_state.get('excel_data_entrant') and st.session_state.get('excel_data_sortant'):
        col_dl1, col_dl2 = st.columns(2)
        with col_dl1:
            st.download_button(
                label="⬇️ Télécharger ENTRANTS",
                data=st.session_state.excel_data_entrant,
                file_name=st.session_state.excel_filename_entrant,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary"
            )
        with col_dl2:
            st.download_button(
                label="⬇️ Télécharger SORTANTS",
                data=st.session_state.excel_data_sortant,
                file_name=st.session_state.excel_filename_sortant,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary"
            )
        st.success(f"✅ Deux fichiers prêts: entrants + sortants")
    else:
        st.download_button(
            label="⬇️ Télécharger le fichier Excel",
            data=st.session_state.excel_data,
            file_name=st.session_state.excel_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )
        st.success(f"✅ Workbook prêt: {st.session_state.excel_filename}")
    
    # Aperçu de toutes les feuilles du fichier Excel
    st.markdown("---")
    st.markdown("### 📑 Aperçu de toutes les feuilles du fichier Excel")
    
    # Déterminer quels fichiers afficher (2 fichiers entrant/sortant ou fichier unique)
    preview_files = []
    if st.session_state.get('excel_data_entrant') and st.session_state.get('excel_data_sortant'):
        preview_files.append(("📥 Entrants", st.session_state.excel_data_entrant))
        preview_files.append(("📤 Sortants", st.session_state.excel_data_sortant))
    else:
        preview_files.append(("Résultat", st.session_state.excel_data))
    
    for preview_label, preview_data in preview_files:
        if len(preview_files) > 1:
            st.markdown(f"#### {preview_label}")
        try:
            import io
            from openpyxl import load_workbook
            wb_preview = load_workbook(io.BytesIO(preview_data), read_only=True, data_only=True)
            sheet_names = wb_preview.sheetnames
            
            # Onglets pour chaque feuille
            if len(sheet_names) > 1:
                tabs = st.tabs(sheet_names)
                for tab, sheet_name in zip(tabs, sheet_names):
                    with tab:
                        ws = wb_preview[sheet_name]
                        sheet_data = []
                        headers = None
                        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                            # Ignorer la ligne "⬅ Retour au summary" et les lignes vides
                            row_str = " ".join(str(c) for c in row if c is not None)
                            if "Retour au summary" in row_str:
                                continue
                            if all(c is None for c in row):
                                continue
                            if headers is None:
                                raw_headers = [str(c) if c is not None else "" for c in row]
                                # Dé-dupliquer les en-têtes (ajouter _1, _2... si nécessaire)
                                seen = {}
                                deduped = []
                                for h in raw_headers:
                                    if h in seen:
                                        seen[h] += 1
                                        deduped.append(f"{h}_{seen[h]}" if h else f"col_{seen[h]}")
                                    else:
                                        seen[h] = 0
                                        deduped.append(h)
                                headers = deduped
                            else:
                                sheet_data.append(row)
                        
                        if headers and sheet_data:
                            df_sheet = pd.DataFrame(sheet_data, columns=headers)
                            # Formater les montants si la colonne existe
                            if "montant" in df_sheet.columns:
                                try:
                                    df_sheet["montant"] = df_sheet["montant"].apply(
                                        lambda x: ("{:,}".format(x)).replace(",", " ") if pd.notnull(x) and isinstance(x, (int, float)) else x
                                    )
                                except Exception:
                                    pass
                            # Forcer toutes les colonnes en string pour éviter les ArrowTypeError
                            # de pyarrow sur les colonnes à types mixtes (None / float / int / str)
                            for _col in df_sheet.columns:
                                df_sheet[_col] = df_sheet[_col].apply(
                                    lambda v: "" if v is None else (str(v) if not isinstance(v, str) else v)
                                )
                            st.dataframe(df_sheet, use_container_width=True, height=300)
                            st.caption(f"📊 {len(sheet_data)} ligne(s)")
                        elif headers:
                            st.info("Feuille vide (en-têtes uniquement)")
                        else:
                            st.info("Feuille vide")
            wb_preview.close()
        except Exception as e:
            st.warning(f"⚠️ Impossible de charger l'aperçu multi-feuilles: {e}")

# Footer
st.markdown("---")
st.markdown("""
    <div style='text-align: center; color: gray; font-size: 0.8em;'>
        PDF & Excel SWIFT Extractor | Déployé sur Hugging Face Spaces
    </div>
""", unsafe_allow_html=True)
