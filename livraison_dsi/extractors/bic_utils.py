# backend/app/extractors/bic_utils.py
"""
Utilities to extract the IdentifierCode from F52A and to map the first-8-chars
to a human-readable bank name using an Excel table (data/bic_codes.xlsx).

Provides:
 - load_bic_mapping(xlsx_path: Optional[str]) -> Dict[str, str]
 - map_code_to_name(code: str, xlsx_path: Optional[str]) -> Optional[str]
 - get_name_for_code(code: str, xlsx_path: Optional[str]) -> Optional[str]  # alias for compatibility
 - get_donneur_from_f52(f52_text, message_text=None, xlsx_path=None) -> Optional[str]
"""
from pathlib import Path
import re
from functools import lru_cache
from typing import Optional, Dict

try:
    import pandas as pd
except Exception as e:
    raise RuntimeError("pandas is required by bic_utils.py (pip install pandas)") from e

# exact label to find (case-insensitive)
_LABEL_TEXT = "IdentifierCode: Code d'identifiant:"
_LABEL_RE = re.compile(re.escape(_LABEL_TEXT), re.I)

# strict code rule: only uppercase letters, length 8..11
_STRICT_TOKEN_RE = re.compile(r'^[A-Z]{8,11}$')

# fallback token pattern (alphanumeric 8..11) - used only if strict not found
_FALLBACK_TOKEN_RE = re.compile(r'\b([A-Z0-9]{8,11})\b', re.I)

# module-level cache
_BIC_MAP_CACHE: Optional[Dict[str, str]] = None
_BIC_FULLKEY_MAP: Optional[Dict[str, str]] = None
_BIC_COUNTRY_MAP: Optional[Dict[str, str]] = None  # Code BIC 8-char -> Pays (ISO3)
_BIC_COUNTRY_FULLKEY_MAP: Optional[Dict[str, str]] = None  # Code BIC full -> Pays (ISO3)
_REGLEMENT_MAP: Optional[Dict[str, Dict[str, str]]] = None  # Code règlement 4 chiffres -> {name, country, bic}
_CCF_MAP: Optional[Dict[str, Dict[str, str]]] = None  # Code CCF simplifié -> {name, country, bic}
_FOREX_CODES: Optional[set] = None  # Codes BIC forex (depuis feuille 'forex' de bic_codes.xlsx)

# Codes Trésor valides à chercher dans F50F (les infos seront récupérées depuis Excel)
TRESOR_CODES_VALIDES = frozenset(["1001", "2001", "3001", "4001", "5001", "6001"])


def _find_strict_identifier_in_f52(f52_text: str) -> Optional[str]:
    """
    Apply the strict rule:
      - find exact label (case-insensitive) inside f52_text
      - inspect: same line after label, next line, next-next line
      - return the first token matching ^[A-Z]{8,11}$
    """
    if not f52_text:
        return None

    lines = [ln.rstrip() for ln in f52_text.splitlines()]
    label_idx = None
    label_span_end = None
    for i, ln in enumerate(lines):
        m = _LABEL_RE.search(ln)
        if m:
            label_idx = i
            label_span_end = m.end()
            break

    if label_idx is None:
        return None

    candidates = []

    # same line after label
    same_line_after = lines[label_idx][label_span_end:].strip()
    if same_line_after:
        candidates.append(same_line_after)

    # next two lines (allow blank lines)
    for j in (label_idx + 1, label_idx + 2):
        if j < len(lines):
            candidates.append(lines[j].strip())

    # evaluate candidates strictly: only A-Z length 8..11
    for cand in candidates:
        if not cand:
            continue
        # split candidate into tokens of letters only (strip punctuation)
        toks = re.findall(r'[A-Z]+', cand.upper())
        for t in toks:
            if _STRICT_TOKEN_RE.match(t):
                return t
    return None


@lru_cache(maxsize=1)
def load_bic_mapping(xlsx_path: Optional[str] = None) -> Dict[str, str]:
    """
    Load the BIC mapping Excel file and return a dict mapping 8-char key -> bank name.
    Also populate a full-key map for 11-char exact matches, a country map, a reglement map, and a CCF map.
    Default path: data/bic_codes.xlsx

    Expected columns: try to detect columns for code and name (flexible).
    """
    global _BIC_MAP_CACHE, _BIC_FULLKEY_MAP, _BIC_COUNTRY_MAP, _BIC_COUNTRY_FULLKEY_MAP, _REGLEMENT_MAP, _CCF_MAP
    if _BIC_MAP_CACHE is not None and _BIC_FULLKEY_MAP is not None:
        return _BIC_MAP_CACHE

    fp = Path(xlsx_path) if xlsx_path else Path("data/bic_codes.xlsx")
    mapping: Dict[str, str] = {}
    mapping_full: Dict[str, str] = {}
    country_map: Dict[str, str] = {}
    country_fullkey_map: Dict[str, str] = {}
    reglement_map: Dict[str, Dict[str, str]] = {}
    ccf_map: Dict[str, Dict[str, str]] = {}

    if not fp.exists():
        _BIC_MAP_CACHE = {}
        _BIC_FULLKEY_MAP = {}
        _BIC_COUNTRY_MAP = {}
        _BIC_COUNTRY_FULLKEY_MAP = {}
        _REGLEMENT_MAP = {}
        _CCF_MAP = {}
        return _BIC_MAP_CACHE

    df = pd.read_excel(fp, dtype=str)
    # normalize column names
    cols = [c for c in df.columns]

    # heuristics to find code and name columns
    code_col = None
    name_col = None
    country_col = None
    reglement_col = None
    ccf_col = None
    col_upper = [c.strip().upper() for c in cols]

    # common name candidates
    for i, cu in enumerate(col_upper):
        if cu in ("CODE", "BIC", "CODE BIC", "BIC_CODE", "BIC8", "CODE8", "CODE_BIC", "CODEBIC"):
            code_col = cols[i]
            break
    if not code_col:
        # fallback: pick first column whose values look like BICs/8..11 alnum
        for c in cols:
            sample = df[c].dropna().astype(str)
            if not sample.empty and sample.str.match(r'^[A-Z0-9]{6,12}$', case=False).sum() > 0:
                code_col = c
                break

    for i, cu in enumerate(col_upper):
        if cu in ("NOMS", "NOM", "NAME", "BANK", "INSTITUTION", "NOMINSTITUTION"):
            name_col = cols[i]
            break
    if not name_col:
        # fallback: choose the first non-code column
        for c in cols:
            if c != code_col:
                name_col = c
                break

    # look for country column
    for i, cu in enumerate(col_upper):
        if cu in ("PAYS", "COUNTRY", "COUNTRY_CODE", "ISO3", "ISO_COUNTRY"):
            country_col = cols[i]
            break

    # look for reglement column (compte de règlement / code 4 chiffres)
    for i, cu in enumerate(col_upper):
        if 'REGLEMENT' in cu or 'RÈGLEMENT' in cu or cu in ("COMPTE", "SETTLEMENT", "COMPTE_REGLEMENT"):
            reglement_col = cols[i]
            break

    # look for CCF column
    for i, cu in enumerate(col_upper):
        if cu == "CCF" or "CCF" in cu:
            ccf_col = cols[i]
            break

    if not code_col:
        _BIC_MAP_CACHE = {}
        _BIC_FULLKEY_MAP = {}
        _BIC_COUNTRY_MAP = {}
        _BIC_COUNTRY_FULLKEY_MAP = {}
        _REGLEMENT_MAP = {}
        _CCF_MAP = {}
        return _BIC_MAP_CACHE

    # build mapping
    for _, row in df.iterrows():
        raw_code = str(row.get(code_col) or "").strip()
        if not raw_code:
            continue
        raw_code = raw_code.replace(" ", "").upper()
        raw_name = ""
        if name_col:
            raw_name = str(row.get(name_col) or "").strip()
        raw_country = ""
        if country_col:
            raw_country = str(row.get(country_col) or "").strip().upper()
        raw_reglement = ""
        if reglement_col:
            raw_reglement = str(row.get(reglement_col) or "").strip()
        raw_ccf = ""
        if ccf_col:
            raw_ccf = str(row.get(ccf_col) or "").strip()
        
        # map first 8 chars -> name
        key8 = raw_code[:8]
        if key8:
            if raw_name:
                mapping[key8] = raw_name
            else:
                # if no name column found, use the raw_code as fallback value (rare)
                mapping.setdefault(key8, raw_code)
            # also store country for this 8-char key
            if raw_country:
                country_map[key8] = raw_country
        
        # also keep full mapping for exact 11-char keys (useful for BEACCMCX100)
        if len(raw_code) >= 8:
            mapping_full[raw_code] = raw_name or mapping.get(raw_code[:8], "")
            # also store full-key country
            if raw_country:
                country_fullkey_map[raw_code] = raw_country
        
        # Store reglement code mapping (4 digit codes)
        if raw_reglement and re.match(r'^\d{4,}$', raw_reglement):
            # Extraire les 4 derniers chiffres si plus long
            code_4 = raw_reglement[-4:] if len(raw_reglement) > 4 else raw_reglement
            reglement_map[code_4] = {
                "name": raw_name,
                "country": raw_country,
                "bic": raw_code
            }
        
        # Store CCF code mapping
        # CCF format: XX.XXXXXX.X.XXXX.X.X.X.X.X (ex: 10.311301.0.1401.0.0.0.0.0)
        # Simplified format: XX.XXXXXX.XXXX (ex: 10.311301.1401) - remove middle 0 and trailing zeros
        if raw_ccf and '.' in raw_ccf:
            # Simplifier le code CCF : garder les 2 premières parties + la 4ème partie
            parts = raw_ccf.split('.')
            if len(parts) >= 4:
                # Format simplifié: partie1.partie2.partie4
                simplified_ccf = f"{parts[0]}.{parts[1]}.{parts[3]}"
                ccf_map[simplified_ccf] = {
                    "name": raw_name,
                    "country": raw_country,
                    "bic": raw_code,
                    "full_ccf": raw_ccf
                }

    _BIC_MAP_CACHE = mapping
    _BIC_FULLKEY_MAP = mapping_full
    _BIC_COUNTRY_MAP = country_map
    _BIC_COUNTRY_FULLKEY_MAP = country_fullkey_map
    _REGLEMENT_MAP = reglement_map
    _CCF_MAP = ccf_map
    return _BIC_MAP_CACHE


def load_forex_codes(xlsx_path: Optional[str] = None) -> set:
    """
    Charger les codes BIC forex depuis la feuille 'forex' de bic_codes.xlsx.
    Les codes sont mis en cache après le premier chargement.
    
    Returns:
        Set de codes BIC (8 premiers caractères, uppercase) considérés comme forex.
    """
    global _FOREX_CODES
    if _FOREX_CODES is not None:
        return _FOREX_CODES
    
    fp = Path(xlsx_path) if xlsx_path else Path("data/bic_codes.xlsx")
    if not fp.exists():
        _FOREX_CODES = set()
        return _FOREX_CODES
    
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(fp), read_only=True, data_only=True)
        if 'forex' not in wb.sheetnames:
            _FOREX_CODES = set()
            wb.close()
            return _FOREX_CODES
        
        ws = wb['forex']
        codes = set()
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0]:
                code = str(row[0]).strip().upper()
                if code:
                    codes.add(code[:8])  # Normaliser à 8 caractères
        wb.close()
        _FOREX_CODES = codes
        return _FOREX_CODES
    except Exception as e:
        import logging
        logging.getLogger(__name__).warning("load_forex_codes: error loading forex sheet: %s", e)
        _FOREX_CODES = set()
        return _FOREX_CODES


def map_code_to_name(code: str, xlsx_path: Optional[str] = None) -> Optional[str]:
    """
    Map a raw code (8..11 chars) to bank name using loaded mapping.
    Rules:
      - If code exactly matches an 11-char full key present in the sheet (e.g. BEACCMCX100),
        then prefer that exact mapping.
      - Otherwise map using the first 8 characters.
    Returns the bank name or None.
    """
    if not code:
        return None
    _ = load_bic_mapping(xlsx_path=xlsx_path)  # populate caches
    global _BIC_MAP_CACHE, _BIC_FULLKEY_MAP
    code_u = code.strip().upper()
    # prefer exact full key if present
    if _BIC_FULLKEY_MAP and code_u in _BIC_FULLKEY_MAP and _BIC_FULLKEY_MAP[code_u]:
        return _BIC_FULLKEY_MAP[code_u]
    key8 = code_u[:8]
    if _BIC_MAP_CACHE and key8 in _BIC_MAP_CACHE:
        return _BIC_MAP_CACHE[key8]
    return None


# Backwards-compatibility alias expected by older code
def get_name_for_code(code: str, xlsx_path: Optional[str] = None) -> Optional[str]:
    """
    Compatibility wrapper used by some older extractors.
    Returns same as map_code_to_name.
    """
    return map_code_to_name(code, xlsx_path=xlsx_path)


def get_donneur_from_f52(f52_text: Optional[str], message_text: Optional[str] = None, xlsx_path: Optional[str] = None) -> Optional[str]:
    """
    Public helper used by extractors:
     - find strict IdentifierCode token inside f52_text (or within message_text if absent)
     - map to bank name (first 8 chars) and return "CODE/Bank Name"
     - if no mapping, return CODE (the extracted token)
     - return None if no token found

    NOTE: token is expected to be letters A-Z only (8..11) based on your rule.
    """
    code = None
    if f52_text:
        code = _find_strict_identifier_in_f52(f52_text)
    if not code and message_text:
        # try in the full message (cross-page)
        code = _find_strict_identifier_in_f52(message_text)
    # final fallback: search for an 8..11 alpha token inside f52 block
    if not code and f52_text:
        m = _FALLBACK_TOKEN_RE.search(f52_text)
        if m:
            cand = m.group(1).upper()
            # accept only all-letters candidate (user requested only letters as valid code)
            if re.fullmatch(r'[A-Z]{8,11}', cand):
                code = cand

    if not code:
        return None

    name = map_code_to_name(code, xlsx_path=xlsx_path)
    if name:
        return f"{code}/{name}"
    return code


def map_reglement_code(code_4: str, xlsx_path: Optional[str] = None) -> Optional[Dict[str, str]]:
    """
    Map a 4-digit reglement code to bank info using the Reglement column.
    
    Args:
        code_4: 4-digit reglement code (e.g., "8428")
        xlsx_path: Optional path to bic_codes.xlsx
        
    Returns:
        Dict with keys 'name', 'country', 'bic' if found, None otherwise
    """
    if not code_4:
        return None
    
    _ = load_bic_mapping(xlsx_path=xlsx_path)  # populate caches
    global _REGLEMENT_MAP
    
    if not _REGLEMENT_MAP:
        return None
    
    # Normaliser le code (prendre les 4 derniers chiffres)
    code_clean = code_4.strip()[-4:]
    
    if code_clean in _REGLEMENT_MAP:
        return _REGLEMENT_MAP[code_clean]
    
    return None


def extract_4digit_code_from_f52d(f52d_text: str) -> Optional[str]:
    """
    Extract a 4-digit reglement code from F52D field.
    The code is typically found after "PartyIdentifier: Identifiant de partie:"
    
    Args:
        f52d_text: The F52D block text
        
    Returns:
        4-digit code if found, None otherwise
    """
    if not f52d_text:
        return None
    
    # D'abord chercher après "PartyIdentifier: Identifiant de partie:"
    party_id_pattern = r'(?:PartyIdentifier|Identifiant de partie)\s*[:\s]*\s*(\d{4})\b'
    m = re.search(party_id_pattern, f52d_text, re.IGNORECASE)
    if m:
        return m.group(1)
    
    # Chercher le label exact puis le code sur la même ligne ou ligne suivante
    label_match = re.search(r'PartyIdentifier\s*:\s*Identifiant de partie\s*:', f52d_text, re.IGNORECASE)
    if label_match:
        # Chercher le code 4 chiffres après ce label
        after_label = f52d_text[label_match.end():]
        code_match = re.search(r'\s*(\d{4})\b', after_label)
        if code_match:
            return code_match.group(1)
    
    # Fallback: chercher un code de 4 chiffres dans le texte
    # Il peut être dans un format comme /8428 ou juste 8428
    patterns = [
        r'/(\d{4})\b',           # /8428
        r'\b(\d{4})\b',          # 8428 seul
        r'[A-Z]{2}\d{2}(\d{4})', # Dans un IBAN-like
    ]
    
    for pattern in patterns:
        m = re.search(pattern, f52d_text)
        if m:
            return m.group(1)
    
    return None


def extract_ccf_4digit_code_from_f50f(f50f_text: str, xlsx_path: Optional[str] = None) -> Optional[Dict[str, str]]:
    """
    Extract CCF 4-digit codes from F50F field by searching for codes found in the CCF column of bic_codes.xlsx.
    The CCF format is like: 10.311301.0.1401.0.0.0.0.0
    We extract the 4th part (1401, 1428, 1073, 9165, 1304, 6401) from all CCF entries,
    then search for any of these 4-digit codes in the F50F text.
    
    Args:
        f50f_text: The F50F block text
        xlsx_path: Optional path to bic_codes.xlsx
        
    Returns:
        Dict with keys 'code', 'name', 'country', 'bic' if found, None otherwise
    """
    if not f50f_text:
        return None
    
    # Load the BIC mapping to populate the CCF map
    _ = load_bic_mapping(xlsx_path=xlsx_path)
    global _CCF_MAP
    
    if not _CCF_MAP:
        return None
    
    # Extract all 4-digit codes from CCF entries (4th part of the CCF format)
    ccf_4digit_codes = {}
    for simplified_ccf, info in _CCF_MAP.items():
        # simplified_ccf format: 10.311301.1401
        parts = simplified_ccf.split('.')
        if len(parts) >= 3:
            code_4digit = parts[2]  # The 3rd part is the 4-digit code
            if code_4digit and code_4digit.isdigit() and len(code_4digit) == 4:
                ccf_4digit_codes[code_4digit] = {
                    "ccf": simplified_ccf,
                    "name": info["name"],
                    "country": info["country"],
                    "bic": info.get("bic", ""),
                    "full_ccf": info.get("full_ccf", "")
                }
    
    # Search for any of these 4-digit codes in F50F text
    for code_4digit, info in ccf_4digit_codes.items():
        # Pattern: the 4-digit code should be a standalone number or part of a longer sequence
        # We search for it with word boundaries or within a longer numeric sequence
        pattern = r'\b' + re.escape(code_4digit) + r'\b'
        if re.search(pattern, f50f_text):
            return {
                "code": code_4digit,
                "name": info["name"],
                "country": info["country"],
                "bic": info.get("bic", ""),
                "ccf": info["ccf"]
            }
    
    return None
    for pattern in patterns:
        m = re.search(pattern, f52d_text)
        if m:
            return m.group(1)
    
    return None


def map_code_to_country(code: str, xlsx_path: Optional[str] = None) -> Optional[str]:
    """
    Map a raw code (8..11 chars) to country ISO3 code using loaded mapping.
    Checks exact full key first (11 chars), then falls back to 8-char prefix.
    Returns the country ISO3 code (e.g., "CMR") or None if not found.
    """
    if not code:
        return None
    _ = load_bic_mapping(xlsx_path=xlsx_path)  # populate caches
    global _BIC_COUNTRY_MAP, _BIC_COUNTRY_FULLKEY_MAP
    code_u = code.strip().upper()
    # Try exact full key first (resolves e.g. BEACCMCX090 vs BEACCMCX091)
    if _BIC_COUNTRY_FULLKEY_MAP and code_u in _BIC_COUNTRY_FULLKEY_MAP:
        return _BIC_COUNTRY_FULLKEY_MAP[code_u]
    key8 = code_u[:8]
    if _BIC_COUNTRY_MAP and key8 in _BIC_COUNTRY_MAP:
        return _BIC_COUNTRY_MAP[key8]
    return None


def add_bic_code_to_xlsx(code: str, name: str, country: str, xlsx_path: Optional[str] = None) -> bool:
    """
    Add a new BIC code entry to the Excel file.
    
    Args:
        code: BIC code (8-11 chars)
        name: Bank name
        country: Country ISO3 code
        xlsx_path: Optional custom path to bic_codes.xlsx
        
    Returns:
        True if successful, False otherwise
    """
    try:
        from openpyxl import load_workbook
    except Exception:
        logger.error("openpyxl required for add_bic_code_to_xlsx")
        return False
    
    fp = Path(xlsx_path) if xlsx_path else Path("data/bic_codes.xlsx")
    
    if not fp.exists():
        logger.error("File not found: %s", fp)
        return False
    
    try:
        # Load workbook
        wb = load_workbook(fp)
        ws = wb.active
        
        # Append new row (Noms, Nom abrégé, Code BIC, Pays, ...)
        ws.append([
            name,                           # Noms
            "",                             # Nom abrégé (optional)
            code.upper(),                   # Code BIC
            country.upper(),                # Pays
            "",                             # Compte de règlement
            "Actif",                        # Statut
            "",                             # Type de participation
            "",                             # Type d'institution
            "",                             # Mode de connexion
            ""                              # Heure de modification
        ])
        
        # Save workbook
        wb.save(fp)
        
        # Clear caches so next load picks up new data
        global _BIC_MAP_CACHE, _BIC_FULLKEY_MAP, _BIC_COUNTRY_MAP, _BIC_COUNTRY_FULLKEY_MAP
        _BIC_MAP_CACHE = None
        _BIC_FULLKEY_MAP = None
        _BIC_COUNTRY_MAP = None
        _BIC_COUNTRY_FULLKEY_MAP = None
        load_bic_mapping.cache_clear()
        
        logger.info("Added BIC code: %s (%s) - %s", code, name, country)
        return True
    
    except Exception as e:
        logger.exception("Failed to add BIC code to %s: %s", fp, e)
        return False


def extract_tresor_code_from_f50f(f50f_text: str, xlsx_path: Optional[str] = None) -> Optional[Dict[str, str]]:
    """
    Extraire le code Trésor depuis le champ F50F pour les MT103 sortants.
    
    Logique en 2 étapes:
    1) Chercher un code Trésor à 4 chiffres (1001, 2001, 3001, 4001, 5001, 6001)
       Ces codes peuvent être adjacents à des /
       Les infos (nom, pays, BIC) sont récupérées depuis le fichier Excel (colonne Reglement)
    2) Si non trouvé, chercher un code CCF:
       - Format long: XX.XXXXXX.X.XXXX.X.X.X.X.X (ex: 10.312501.0.1073.0.0.0.0.0)
       - Format court: XX.XXXXXX.XXXX (ex: 10.312501.1073)
       Simplifier en XX.XXXXXX.XXXX et matcher avec la colonne CCF
    
    Args:
        f50f_text: Le contenu du champ F50F
        xlsx_path: Chemin optionnel vers le fichier Excel BIC
        
    Returns:
        Dict avec 'code', 'name', 'country', 'bic' si trouvé, None sinon
    """
    if not f50f_text:
        return None
    
    # Charger le mapping si nécessaire
    _ = load_bic_mapping(xlsx_path=xlsx_path)
    global _CCF_MAP, _REGLEMENT_MAP
    
    # ÉTAPE 1: Chercher un code Trésor spécifique (1001, 2001, 3001, 4001, 5001, 6001)
    # Ces codes peuvent être adjacents à des / comme /1001 ou 1001/
    # Les infos sont récupérées dynamiquement depuis le fichier Excel
    for tresor_code in TRESOR_CODES_VALIDES:
        # Pattern pour trouver le code exact, possiblement adjacent à /
        pattern = r'(?:^|/)?' + re.escape(tresor_code) + r'(?:/|$|\s|[^\d])'
        if re.search(pattern, f50f_text):
            # Chercher les infos dans le mapping Reglement
            if _REGLEMENT_MAP and tresor_code in _REGLEMENT_MAP:
                info = _REGLEMENT_MAP[tresor_code]
                return {
                    "code": tresor_code,
                    "name": info.get("name", ""),
                    "country": info.get("country", ""),
                    "bic": info.get("bic", "")
                }
            else:
                # Code trouvé mais pas dans Excel - retourner au moins le code
                return {
                    "code": tresor_code,
                    "name": f"Code Trésor {tresor_code}",
                    "country": "",
                    "bic": ""
                }
    
    # ÉTAPE 2: Chercher un code CCF
    simplified_ccf = None
    
    # D'abord essayer le format long: XX.XXXXXX.X.XXXX.X.X.X.X.X
    ccf_long_pattern = r'(\d{2}\.\d{6}\.\d\.\d{4}(?:\.\d){0,5})'
    ccf_long_match = re.search(ccf_long_pattern, f50f_text)
    
    if ccf_long_match:
        full_ccf = ccf_long_match.group(1)
        parts = full_ccf.split('.')
        if len(parts) >= 4:
            # Simplifier: partie1.partie2.partie4 (sans le 0 du milieu et les 0 de la fin)
            simplified_ccf = f"{parts[0]}.{parts[1]}.{parts[3]}"
    
    # Si pas trouvé, essayer le format court: XX.XXXXXX.XXXX (directement 3 parties)
    if not simplified_ccf:
        ccf_short_pattern = r'(\d{2}\.\d{6}\.\d{4})(?:\.|$|\s|[^\d])'
        ccf_short_match = re.search(ccf_short_pattern, f50f_text)
        if ccf_short_match:
            simplified_ccf = ccf_short_match.group(1)
    
    if simplified_ccf:
        # Chercher dans le mapping CCF
        if _CCF_MAP and simplified_ccf in _CCF_MAP:
            info = _CCF_MAP[simplified_ccf]
            return {
                "code": simplified_ccf,
                "name": info["name"],
                "country": info["country"],
                "bic": info.get("bic", "")
            }
        
        # Si pas dans le mapping CCF, extraire la 3ème partie comme code règlement
        parts = simplified_ccf.split('.')
        if len(parts) >= 3:
            reglement_code = parts[2]  # La 3ème partie (ex: 1073)
            reglement_info = map_reglement_code(reglement_code, xlsx_path=xlsx_path)
            if reglement_info:
                return {
                    "code": simplified_ccf,
                    "name": reglement_info["name"],
                    "country": reglement_info["country"],
                    "bic": reglement_info.get("bic", "")
                }
    
    return None


def map_ccf_code(ccf_simplified: str, xlsx_path: Optional[str] = None) -> Optional[Dict[str, str]]:
    """
    Map a simplified CCF code to bank info using the CCF column.
    
    Args:
        ccf_simplified: Simplified CCF code (e.g., "10.311301.1401")
        xlsx_path: Optional path to bic_codes.xlsx
        
    Returns:
        Dict with keys 'name', 'country', 'bic' if found, None otherwise
    """
    if not ccf_simplified:
        return None
    
    _ = load_bic_mapping(xlsx_path=xlsx_path)  # populate caches
    global _CCF_MAP
    
    if not _CCF_MAP:
        return None
    
    if ccf_simplified in _CCF_MAP:
        return _CCF_MAP[ccf_simplified]
    
    return None