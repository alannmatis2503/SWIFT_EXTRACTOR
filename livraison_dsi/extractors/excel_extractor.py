"""Extracteur de données depuis les fichiers Excel des correspondants (BDF, CITI, Standard, CITI USD Relevé).
Applique les règles de routage et d'exception adaptées à chaque correspondant.
"""

import re
import logging
from collections import defaultdict
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from datetime import datetime

import openpyxl

logger = logging.getLogger(__name__)

# Import BIC utilities
try:
    from extractors import bic_utils
    from extractors.bic_utils import load_bic_mapping, map_code_to_name, map_code_to_country, load_forex_codes, map_reglement_code
    HAS_BIC_UTILS = True
except Exception:
    bic_utils = None
    load_bic_mapping = None
    map_code_to_name = None
    map_code_to_country = None
    load_forex_codes = None
    map_reglement_code = None
    HAS_BIC_UTILS = False

# --- Pays CEMAC pour détection fuzzy ---
CEMAC_COUNTRIES = {
    "CAM": ["CAMEROUN", "CAMEROON", "CAMERO", "CAMEROU",
            "REPUBLIQUE DU CAM"],
    "GAB": ["GABON", "GABO"],
    "CGO": ["CONGO", "CONG", "REPUBLIQUE DU CONE"],
    "TCH": ["TCHAD", "CHAD", "TCHA"],
    "RCA": ["CENTRAFRIQUE", "CENTRAFRICAINE", "CENTRAFRICAIN", "CENTRAL AFRICAN", "RCA"],
    "GEQ": ["GUINEE EQUATORIALE", "GUINEE EQUATORIAL", "GUINEA EQUATORIAL", "EQUATORIAL GUINEA",
            "GUINEE EQUAT", "EQUATORIA", "GUINEA ECUATORIA", "GUINEA ECUATORIAL",
            "GUINE ECUATORIAL", "AEQUATORIAL GUINEA", "GUINEE EQUATO"],
}

# --- Villes CEMAC → pays ---
CEMAC_CITIES = {
    "DOUALA": "CAM", "YAOUNDE": "CAM", "BAFOUSSAM": "CAM", "BAMENDA": "CAM",
    "GAROUA": "CAM", "BERTOUA": "CAM", "KRIBI": "CAM", "LIMBE": "CAM",
    "LIBREVILLE": "GAB", "PORT-GENTIL": "GAB", "PORT GENTIL": "GAB",
    "FRANCEVILLE": "GAB", "OYEM": "GAB",
    "BRAZZAVILLE": "CGO", "POINTE-NOIRE": "CGO", "POINTE NOIRE": "CGO",
    "NDJAMENA": "TCH", "N'DJAMENA": "TCH", "N DJAMENA": "TCH", "MOUNDOU": "TCH",
    "BANGUI": "RCA", "BERBERATI": "RCA",
    "MALABO": "GEQ", "BATA": "GEQ", "EBEBIYIN": "GEQ",
}

# --- Codes ISO2 CEMAC → pays ---
CEMAC_ISO2 = {
    "CM": "CAM", "GA": "GAB", "TD": "TCH",
    "CG": "CGO", "GQ": "GEQ", "CF": "RCA",
}

# Pattern BIC standard : 4 lettres + 2 lettres + 2-5 alphanum
_BIC_PATTERN = re.compile(r'\b([A-Z]{4}[A-Z]{2}[A-Z0-9]{2,5})\b')

# Mots qui ressemblent à des BIC mais n'en sont pas
_FALSE_BIC_WORDS = frozenset([
    'CAMEROON', 'CAMEROUN', 'GABON', 'CONGO', 'TCHAD', 'CENTRAFRICAINE',
    'GUINEE', 'EQUATORIALE', 'DOUALA', 'YAOUNDE', 'MALABO', 'LIBREVILLE',
    'BRAZZAVILLE', 'BANGUI', 'NDJAMENA', 'INSTITUTION', 'IDENTIFIANT',
    'IDENTIFIER', 'BENEFICIAIRE', 'DONNEUR', 'ORDRE', 'PARTIE',
    'TRANSFER', 'PAYMENT', 'CREDIT', 'DEBIT', 'COMPTE', 'ACCOUNT',
    'FINANCES', 'MINISTERE', 'TRESOR', 'BANQUE', 'FRANCE', 'AFRIQUE',
    'STANDARD', 'CHARTERED', 'INTEREST', 'INTERNAL', 'VIREMENT',
    'INTERBANCAIRE', 'CLIENTELE', 'REIMBURSEMENT',
    'NOTPROVIDED', 'ATTIJARI', 'ATLANTIQUE', 'NATIONALE', 'NACIONAL',
    'CONGOLAISE', 'ISLAMIQUE', 'CONSULADO', 'CONSULAT', 'CONSULATE',
    'EMBAJADA', 'EMBAIXADA', 'AMBASSADE', 'PERMANENTE', 'PERCEPTION',
    'REPUBLIQUE', 'REPUBLICA', 'GOBIERNO', 'DELEGACION', 'COMMISSION',
    'BOTSCHAFT', 'ECUATORIA', 'ECUATORIAL', 'ECUATORIALE',
])

# Pattern ISO2 isolé (mot entier) pour les pays CEMAC
_ISO2_PATTERN = re.compile(r'(?<![A-Z])\b(CM|GA|TD|CG|GQ|CF)\b(?![A-Z])')

# Pattern pour extraire le code sous-participant après /BNF/
# Matche /BNF/ suivi optionnellement d'un espace, puis exactement 4 chiffres
_BNF_CODE_RE = re.compile(r'/BNF/\s?(\d{4})\b')

# Codes de règlement spéciaux BDF (4 premiers chars de Référence client) → (nom pays, ISO3)
# Utilisés comme fallback quand le code n'est pas dans bic_codes.xlsx
_BDF_REGLEMENT_PAYS = {
    "9910": ("Cameroun", "CAM"),
    "9920": ("Tchad", "TCH"),
    "9930": ("RCA", "RCA"),
    "9940": ("Congo", "CGO"),
    "9950": ("Gabon", "GAB"),
    "9960": ("Guinée Equatoriale", "GEQ"),
}


def _detect_cemac_country_fuzzy(text: str) -> Optional[str]:
    """
    Détection fuzzy d'un pays CEMAC dans un texte.
    Tolère la casse et les petites omissions (1-2 lettres).
    """
    if not text:
        return None
    text_upper = text.upper().strip()
    for pays_code, variantes in CEMAC_COUNTRIES.items():
        for var in variantes:
            # Match exact (case-insensitive)
            if var in text_upper:
                return pays_code
            # Match avec tolérance (supprimer les espaces et comparer)
            var_nospace = var.replace(" ", "")
            text_nospace = text_upper.replace(" ", "")
            if var_nospace in text_nospace:
                return pays_code
            # Match préfixe (au moins len-1 caractères, minimum 4)
            if len(var) >= 5:
                min_match = max(4, len(var) - 1)
                if var[:min_match] in text_upper:
                    return pays_code
    return None


def _detect_cemac_country_from_contrepartie(text: str) -> Optional[str]:
    """
    Détection du pays CEMAC depuis le texte 'Nom contrepartie'.
    Fallbacks successifs :
      1. Nom de pays CEMAC dans le texte
      2. Nom de ville CEMAC dans le texte
      3. Code ISO2 isolé (CM, GA, GQ, CG, TD, CF)
      4. Code ISO2 CEMAC en positions 4-5 d'un pattern BIC dans le texte
    """
    if not text:
        return None
    text_upper = text.upper().strip()

    # Fallback 1 : nom de pays (réutilise la logique existante)
    pays = _detect_cemac_country_fuzzy(text_upper)
    if pays:
        return pays

    # Fallback 2 : nom de ville CEMAC
    for ville, code_pays in CEMAC_CITIES.items():
        if ville in text_upper:
            return code_pays

    # Fallback 3 : code ISO2 isolé (ex: "CM DOUALA", "GA LIBREVILLE")
    m = _ISO2_PATTERN.search(text_upper)
    if m:
        return CEMAC_ISO2[m.group(1)]

    # Fallback 4 : pattern BIC avec code ISO2 CEMAC en positions 4-5
    for bic_match in _BIC_PATTERN.finditer(text_upper):
        candidate = bic_match.group(1)
        if len(candidate) >= 8 and candidate not in _FALSE_BIC_WORDS:
            iso2_in_bic = candidate[4:6]
            if iso2_in_bic in CEMAC_ISO2:
                return CEMAC_ISO2[iso2_in_bic]

    return None


def _extract_bic_from_text(text: str) -> Optional[str]:
    """
    Extraire un code BIC d'un texte.
    Gère le format 'BIC:description' (code avant le ':').
    """
    if not text:
        return None
    text = text.strip()

    # Cas 1 : code BIC avant ':'  (ex: "BCMACGCGXXX:CREDIT DU CO")
    if ':' in text:
        before_colon = text.split(':')[0].strip()
        m = _BIC_PATTERN.search(before_colon.upper())
        if m and m.group(1) not in _FALSE_BIC_WORDS:
            return m.group(1)

    # Cas 2 : code BIC pur dans le texte
    text_upper = text.upper()
    # Nettoyer les retours chariot CR/LF
    text_clean = text_upper.replace('\r', ' ').replace('\n', ' ').replace('_x000D_', ' ')
    m = _BIC_PATTERN.search(text_clean)
    if m and m.group(1) not in _FALSE_BIC_WORDS:
        return m.group(1)

    return None


def _extract_all_bic_candidates(text: str) -> List[str]:
    """Extraire tous les codes BIC candidats d'un texte (filtrés des faux positifs)."""
    if not text:
        return []
    text_clean = text.upper().replace('\r', ' ').replace('\n', ' ').replace('_x000D_', ' ')
    # Cas spécial : BIC avant ':'
    if ':' in text_clean:
        before_colon = text_clean.split(':')[0].strip()
        m = _BIC_PATTERN.search(before_colon)
        if m and m.group(1) not in _FALSE_BIC_WORDS:
            return [m.group(1)]
    candidates = []
    for m in _BIC_PATTERN.finditer(text_clean):
        c = m.group(1)
        if c not in _FALSE_BIC_WORDS:
            candidates.append(c)
    return candidates


def _parse_amount_string(amount_str: str) -> Optional[float]:
    """
    Parser un montant au format string avec virgules/points.
    Ex: '10,967.56', '-85,650.71', '193,275.54'
    """
    if not amount_str:
        return None
    try:
        if isinstance(amount_str, (int, float)):
            return float(amount_str)
        # Supprimer espaces
        s = str(amount_str).strip()
        # Supprimer les virgules qui sont des séparateurs de milliers (format anglais)
        s = s.replace(',', '')
        return float(s)
    except (ValueError, TypeError):
        return None


def _parse_date(date_val, format_hint: str = 'auto') -> Optional[str]:
    """
    Convertir une date (datetime, string) en format ISO YYYY-MM-DD.

    format_hint:
      'auto' (défaut) : DD/MM/YYYY sauf si groupe 2 > 12 → détecte MM/DD/YYYY
      'dmy'           : forcer DD/MM/YYYY (BDF)
      'mdy'           : forcer MM/DD/YYYY (CITI)
    """
    if not date_val:
        return None
    if isinstance(date_val, datetime):
        return date_val.strftime('%Y-%m-%d')
    s = str(date_val).strip()
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', s)
    if m:
        g1, g2, g3 = int(m.group(1)), int(m.group(2)), m.group(3)
        if format_hint == 'mdy' or (format_hint == 'auto' and g1 <= 12 and g2 > 12):
            # MM/DD/YYYY : g1=mois, g2=jour
            return f"{g3}-{g1:02d}-{g2:02d}"
        else:
            # DD/MM/YYYY : g1=jour, g2=mois
            return f"{g3}-{g2:02d}-{g1:02d}"
    # Format YYYY-MM-DD déjà correct
    if re.match(r'^\d{4}-\d{2}-\d{2}$', s):
        return s
    return s


def _find_header_row(ws) -> Optional[int]:
    """
    Trouver la ligne d'en-tête dans une feuille Excel.
    Retourne le numéro de ligne (1-based) ou None.
    """
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
        if row and row[0]:
            first_val = str(row[0]).strip().lower()
            # BDF : "Date opération"
            if 'date' in first_val and ('opération' in first_val or 'operation' in first_val):
                return row_idx
            # CITI : "Date de valeur"
            if 'date de valeur' in first_val or 'date valeur' in first_val:
                return row_idx
    return None


# ========================================================================================
# EXTRACTEUR BDF (Banque de France)
# ========================================================================================

def extract_bdf(file_path: str, xlsx_bic_path: Optional[str] = None) -> Tuple[List[Dict], List[Dict], List[Dict], List[Dict]]:
    """
    Extraire les données d'un fichier relevé BDF.

    Structure BDF (headers à la ligne ~19) :
    Col A: Date opération, Col B: Date valeur, Col C: Libellé mouvement,
    Col D: Référence demandeur (=F20), Col E: Référence client (=F21),
    Col F: Nom contrepartie, Col G: Débit, Col H: Crédit

    Returns:
        (rows, beaccmcx091_rows, other_exceptions_rows, forex_rows)
    """
    fp = Path(file_path)
    if not fp.exists():
        raise FileNotFoundError(f"Fichier BDF introuvable : {file_path}")

    # Charger BIC mapping
    if HAS_BIC_UTILS:
        try:
            load_bic_mapping(xlsx_bic_path)
        except Exception as e:
            logger.debug("excel_extractor: BIC mapping load failed: %s", e)

    # Charger les codes forex
    forex_codes = set()
    if HAS_BIC_UTILS and load_forex_codes:
        try:
            forex_codes = load_forex_codes(xlsx_bic_path)
        except Exception:
            pass

    wb = openpyxl.load_workbook(str(fp), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Trouver la devise depuis les métadonnées
    devise = "EUR"
    for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
        if row and row[0] and 'devise' in str(row[0]).lower():
            if row[1]:
                devise = str(row[1]).strip().upper()
            break

    # Trouver la ligne d'en-tête
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=25, values_only=True), start=1):
        if row and row[0] and 'date' in str(row[0]).strip().lower():
            vals = [str(v or '').lower() for v in row]
            if any('opération' in v or 'operation' in v for v in vals):
                header_row = row_idx
                break
            if any('valeur' in v for v in vals):
                header_row = row_idx
                break

    if not header_row:
        wb.close()
        raise ValueError(f"Impossible de trouver la ligne d'en-tête dans {fp.name}")

    rows = []
    beaccmcx091_rows = []
    other_exceptions_rows = []
    forex_rows = []

    data_start = header_row + 1
    file_name = fp.name

    for row_idx, row in enumerate(ws.iter_rows(min_row=data_start, values_only=True), start=data_start):
        # Ignorer les lignes vides
        if not row or (not row[0] and not row[6] and not row[7]):
            continue

        # Colonnes: A=Date opération, B=Date valeur, C=Libellé mouvement,
        #           D=Référence demandeur(F20), E=Référence client(F21),
        #           F=Nom contrepartie, G=Débit, H=Crédit
        date_val = _parse_date(row[1]) if len(row) > 1 else None
        libelle = str(row[2] or '') if len(row) > 2 else ''
        ref_demandeur = str(row[3] or '') if len(row) > 3 else ''  # F20
        ref_client = str(row[4] or '') if len(row) > 4 else ''    # F21

        # FILTRE: Exclure les lignes AOPD / AOPOD (Référence demandeur)
        _ref_dem_upper = ref_demandeur.strip().upper()
        if 'AOPOD' in _ref_dem_upper or 'AOPD' in _ref_dem_upper:
            continue
        contrepartie_raw = str(row[5] or '') if len(row) > 5 else ''
        debit = row[6] if len(row) > 6 else None
        credit = row[7] if len(row) > 7 else None

        # Ignorer les lignes sans montant
        if debit is None and credit is None:
            continue

        # Déterminer direction
        if debit is not None and debit != '' and debit != 0:
            try:
                montant = abs(float(debit))
            except (ValueError, TypeError):
                continue
            direction = "outgoing"
        elif credit is not None and credit != '' and credit != 0:
            try:
                montant = abs(float(credit))
            except (ValueError, TypeError):
                continue
            direction = "incoming"
        else:
            continue

        # Nettoyer la contrepartie
        contrepartie_clean = contrepartie_raw.replace('_x000D_', ' ').replace('\r', ' ').replace('\n', ' ')
        contrepartie_clean = re.sub(r'\s+', ' ', contrepartie_clean).strip()

        # Extraire le code BIC de la contrepartie
        bic_code = _extract_bic_from_text(contrepartie_raw)
        banque_name = None
        pays_iso3 = None

        if bic_code and HAS_BIC_UTILS:
            banque_name = map_code_to_name(bic_code, xlsx_path=xlsx_bic_path)
            pays_iso3 = map_code_to_country(bic_code, xlsx_path=xlsx_bic_path)

        # Si le premier BIC ne résout pas le pays, essayer les candidats suivants
        if not pays_iso3 and HAS_BIC_UTILS:
            for candidate in _extract_all_bic_candidates(contrepartie_raw):
                if candidate == bic_code:
                    continue
                c_pays = map_code_to_country(candidate, xlsx_path=xlsx_bic_path)
                if c_pays:
                    if not bic_code:
                        bic_code = candidate
                    if not banque_name:
                        banque_name = map_code_to_name(candidate, xlsx_path=xlsx_bic_path)
                    pays_iso3 = c_pays
                    break

        # Fallback : détection pays CEMAC depuis le texte contrepartie
        if not pays_iso3 and contrepartie_clean:
            pays_iso3 = _detect_cemac_country_from_contrepartie(contrepartie_clean)

        # Fallback BDF sortants : si BIC non résolu, chercher dans Référence client
        # Les 4 premiers caractères avant "/" sont le code de règlement
        if direction == "outgoing" and (not bic_code or not pays_iso3) and ref_client:
            _code_ref = ref_client.split('/')[0].strip()[:4] if '/' in ref_client else ref_client.strip()[:4]
            if _code_ref:
                # 1. Chercher dans bic_codes.xlsx (codes règlement) pour bic et nom banque
                if HAS_BIC_UTILS and map_reglement_code:
                    _reg_info = map_reglement_code(_code_ref, xlsx_path=xlsx_bic_path)
                    if _reg_info:
                        if not bic_code:
                            bic_code = _reg_info.get("bic") or bic_code
                        if not banque_name:
                            banque_name = _reg_info.get("name") or banque_name
                        if not pays_iso3:
                            # Accepter le pays de bic_codes.xlsx sauf les pseudo-codes internes
                            # (BEAC = ancien libellé, SCX = Services Centraux BEAC — pas des ISO3 CEMAC)
                            _country_from_map = _reg_info.get("country") or ""
                            if _country_from_map and _country_from_map not in ("BEAC", "SCX", ""):
                                pays_iso3 = _country_from_map
                            elif _country_from_map in ("BEAC", "SCX"):
                                # Conserver provisoirement pour que l'override ci-dessous puisse agir
                                pays_iso3 = _country_from_map
                # 2. Pour les codes pays spéciaux 9910-9960 : override avec le vrai ISO3 CEMAC
                # (ces codes pointent vers les directions nationales BEAC, pas le pays réel)
                if _code_ref in _BDF_REGLEMENT_PAYS:
                    _country_name, _country_iso3 = _BDF_REGLEMENT_PAYS[_code_ref]
                    # Remplacer BEAC/SCX (pseudo-codes internes) par le vrai code pays CEMAC
                    if not pays_iso3 or pays_iso3 in ("BEAC", "SCX"):
                        pays_iso3 = _country_iso3
            # 3. Fallback ISO2 CEMAC dans la référence client (ex: 9135/0201/CG/1 → CG → CGO)
            if not pays_iso3:
                _iso2_in_ref = _ISO2_PATTERN.search(ref_client.upper())
                if _iso2_in_ref:
                    pays_iso3 = CEMAC_ISO2[_iso2_in_ref.group(1)]

        # Si pas de BIC trouvé, garder le texte brut comme donneur/bénéficiaire
        if not banque_name:
            banque_name = contrepartie_clean if contrepartie_clean else None

        # Source tracking
        source = f"ligne N°{row_idx} du fichier {file_name}"

        entry = {
            "correspondant": "BDFEFRPPCCT",
            "date_reference": date_val,
            "reference": ref_demandeur,       # F20
            "reference_origine": ref_client,   # F21
            "type_MT": f"Excel BDF {'entrant' if direction == 'incoming' else 'sortant'}",
            "pays_iso3": pays_iso3,
            "code_donneur_dordre": bic_code,
            "donneur_dordre": banque_name,
            "institution_name": banque_name,
            "beneficiaire": None,
            "montant": montant,
            "devise": devise,
            "commentaires": libelle if libelle else None,
            "source_pdf": source,
            "direction": direction,
        }

        # ===== Application des règles de routage =====

        # RÈGLE: BEACCMCX091 (présence dans le texte Nom contrepartie)
        is_beaccmcx091 = False
        if 'BEACCMCX091' in contrepartie_clean.upper():
            is_beaccmcx091 = True

        # RÈGLE: T2PI dans la référence → exception "intérêts" (entrant ET sortant)
        is_exception = False
        ref_upper = ref_demandeur.upper()
        ref_client_upper = ref_client.upper()

        if devise.upper() == "EUR":
            if "T2PI" in ref_upper:
                entry["commentaires"] = "intérêts"
                is_exception = True
            elif direction == "incoming" and "T2RM" in ref_upper:
                entry["commentaires"] = "remboursement"
                is_exception = True
            elif direction == "outgoing" and "T2PL" in ref_upper:
                entry["commentaires"] = "placement"
                is_exception = True

        # RÈGLE: Nivellement (NIVLT dans F20 ou F21)
        is_nivellement = False
        if "NIVLT" in ref_upper or "NIVLT" in ref_client_upper:
            entry["commentaires"] = "nivellement"
            is_nivellement = True

        # RÈGLE: Forex BDF entrant
        # Critère : BEACCMCX091 dans Nom contrepartie ET /YYMM dans référence client
        # (ex: /2602 = février 2026). Remplace l'ancienne règle basée sur codes BIC forex.
        is_forex = False
        if direction == "incoming" and is_beaccmcx091:
            if re.search(r'/\d{4}', ref_client):
                is_forex = True
                is_beaccmcx091 = False  # Les forex ne vont PAS dans BEACCMCX091
                entry["commentaires"] = (
                    (entry.get("commentaires") or "") + " forex"
                ).strip()

        # Routage vers la bonne liste (forex prioritaire sur BEACCMCX091)
        if is_forex:
            forex_rows.append(entry)
        elif is_beaccmcx091:
            beaccmcx091_rows.append(entry)
        elif is_exception or is_nivellement:
            other_exceptions_rows.append(entry)
        else:
            rows.append(entry)

    wb.close()

    # ---- Dédoublonnage des lignes BDF sortantes ----
    # Une même transaction peut générer plusieurs lignes (principal + commissions) avec
    # la même Référence demandeur. On garde la ligne au montant le plus élevé dans la
    # feuille principale ; les autres sont envoyées en exceptions avec "frais et autres".
    outgoing_main = [r for r in rows if r.get("direction") == "outgoing"]
    incoming_main = [r for r in rows if r.get("direction") != "outgoing"]

    ref_groups: dict = defaultdict(list)
    no_ref_outgoing = []
    for r in outgoing_main:
        ref = (r.get("reference") or "").strip()
        if ref:
            ref_groups[ref].append(r)
        else:
            no_ref_outgoing.append(r)

    deduped_outgoing = []
    for ref, group in ref_groups.items():
        if len(group) == 1:
            deduped_outgoing.extend(group)
        else:
            group_sorted = sorted(group, key=lambda x: x.get("montant") or 0, reverse=True)
            deduped_outgoing.append(group_sorted[0])
            for dup in group_sorted[1:]:
                dup["commentaires"] = "frais et autres"
                other_exceptions_rows.append(dup)

    rows = incoming_main + deduped_outgoing + no_ref_outgoing

    return rows, beaccmcx091_rows, other_exceptions_rows, forex_rows


# ========================================================================================
# EXTRACTEUR CITI (CITI EUR et CITI USD)
# ========================================================================================

def extract_citi(file_path: str, xlsx_bic_path: Optional[str] = None) -> Tuple[List[Dict], List[Dict], List[Dict], List[Dict], List[Dict]]:
    """
    Extraire les données d'un fichier CITI (EUR ou USD).

    Structure CITI (headers en ligne 1) :
    Col A: Date de valeur, Col B: Date du relevé, Col C: Devise, Col D: Montant,
    Col E: Bénéficiaire/Remettant, Col F: Type, Col G: Référence bancaire (=F20),
    Col H: Description, Col I: Détails du paiement

    Returns:
        (rows, beaccmcx091_rows, other_exceptions_rows, forex_rows, banque_de_france_rows)
    """
    fp = Path(file_path)
    if not fp.exists():
        raise FileNotFoundError(f"Fichier CITI introuvable : {file_path}")

    # Charger BIC mapping
    if HAS_BIC_UTILS:
        try:
            load_bic_mapping(xlsx_bic_path)
        except Exception as e:
            logger.debug("excel_extractor: BIC mapping load failed: %s", e)

    # Charger les codes forex
    forex_codes = set()
    if HAS_BIC_UTILS and load_forex_codes:
        try:
            forex_codes = load_forex_codes(xlsx_bic_path)
        except Exception:
            pass

    wb = openpyxl.load_workbook(str(fp), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = []
    beaccmcx091_rows = []
    other_exceptions_rows = []
    forex_rows = []
    banque_de_france_rows = []

    file_name = fp.name
    data_start = 2  # La ligne 1 contient les en-têtes

    for row_idx, row in enumerate(ws.iter_rows(min_row=data_start, values_only=True), start=data_start):
        if not row or not row[0]:
            continue

        # Colonnes: A=Date de valeur, B=Date du relevé, C=Devise, D=Montant,
        #           E=Bénéficiaire/Remettant, F=Type, G=Référence bancaire(F20),
        #           H=Description, I=Détails du paiement
        date_val = _parse_date(row[0], format_hint='mdy') if len(row) > 0 else None
        devise = str(row[2] or '').strip().upper() if len(row) > 2 else ''
        montant_raw = row[3] if len(row) > 3 else None
        benef_remettant = str(row[4] or '') if len(row) > 4 else ''
        type_col = str(row[5] or '') if len(row) > 5 else ''
        ref_bancaire = str(row[6] or '') if len(row) > 6 else ''  # F20
        description = str(row[7] or '') if len(row) > 7 else ''
        details_paiement = str(row[8] or '') if len(row) > 8 else ''

        # Parser le montant
        montant = _parse_amount_string(montant_raw)
        if montant is None:
            continue

        # Déterminer direction
        if montant < 0:
            direction = "outgoing"
            montant = abs(montant)
        else:
            direction = "incoming"

        # Extraire le code BIC du champ Bénéficiaire/Remettant
        bic_code = _extract_bic_from_text(benef_remettant)
        banque_name = None
        pays_iso3 = None

        if bic_code and HAS_BIC_UTILS:
            banque_name = map_code_to_name(bic_code, xlsx_path=xlsx_bic_path)
            pays_iso3 = map_code_to_country(bic_code, xlsx_path=xlsx_bic_path)

        # Pour les sortants : extraire le code sous-participant depuis /BNF/ dans les détails
        code_sous_participant = None
        if direction == "outgoing" and details_paiement and HAS_BIC_UTILS and map_reglement_code:
            bnf_match = _BNF_CODE_RE.search(details_paiement)
            if bnf_match:
                code_4 = bnf_match.group(1)
                reglement_info = map_reglement_code(code_4, xlsx_path=xlsx_bic_path)
                if reglement_info:
                    code_sous_participant = code_4
                    # Remplir les infos depuis bic_codes.xlsx via le code sous-participant
                    if not bic_code or not banque_name:
                        bic_code = reglement_info.get("bic") or bic_code
                        banque_name = reglement_info.get("name") or banque_name
                    if not pays_iso3:
                        pays_iso3 = reglement_info.get("country") or pays_iso3

        # Si pas de BIC ou BIC non mappé : tenter détection pays CEMAC + garder le texte
        if not banque_name:
            benef_clean = benef_remettant.strip()
            # Si le BIC a été extrait mais pas mappé, garder le BIC comme code
            if not bic_code:
                # Pas de BIC trouvé → utiliser le texte brut
                banque_name = benef_clean if benef_clean else None
            else:
                # BIC trouvé mais pas mappé
                banque_name = benef_clean if benef_clean else bic_code

        # Tenter détection fuzzy pays CEMAC si pas de pays trouvé
        if not pays_iso3:
            pays_iso3 = _detect_cemac_country_fuzzy(benef_remettant)

        # Construire commentaires à partir des détails du paiement
        commentaire = details_paiement.strip() if details_paiement.strip() else (description.strip() if description.strip() else None)

        # Source tracking
        source = f"ligne N°{row_idx} du fichier {file_name}"

        entry = {
            "correspondant": "CITIUS33XXX" if devise == "USD" else "CITIGB2LXXX",
            "date_reference": date_val,
            "reference": ref_bancaire,           # F20
            "reference_origine": None,            # Pas de F21 distinct chez CITI
            "type_MT": f"Excel CITI {devise} {'entrant' if direction == 'incoming' else 'sortant'}",
            "pays_iso3": pays_iso3,
            "code_donneur_dordre": bic_code,
            "donneur_dordre": banque_name,
            "institution_name": banque_name,
            "beneficiaire": None,
            "montant": montant,
            "devise": devise,
            "commentaires": commentaire,
            "source_pdf": source,
            "direction": direction,
            "code_sous_participant": code_sous_participant,
        }

        # ===== Application des règles de routage =====

        # RÈGLE: BEACCMCX091 (seulement via le code BIC ou le texte bénéficiaire/remettant)
        is_beaccmcx091 = False
        if bic_code and 'BEACCMCX091' in bic_code.upper():
            is_beaccmcx091 = True
        # Vérifier aussi dans le texte brut du bénéficiaire/remettant
        if not is_beaccmcx091 and 'BEACCMCX091' in benef_remettant.upper():
            is_beaccmcx091 = True

        # RÈGLE: BANQUE DE FRANCE (sortants EUR uniquement)
        is_bdf = False
        if direction == "outgoing" and devise == "EUR":
            benef_upper = benef_remettant.upper()
            if "BANQUE DE FRANCE" in benef_upper:
                is_bdf = True

        # RÈGLE: DE-Data Entry (toutes directions → autres exceptions)
        # Couvre aussi CR.INT.CURR.PERIOD / INTEREST entrants (intérêts de compte, hors périmètre SWIFT)
        is_de_data_entry = False
        type_upper = type_col.strip().upper()
        if type_upper.startswith("DE") and "DATA ENTRY" in type_upper:
            entry["commentaires"] = "DE-Data Entry"
            is_de_data_entry = True
        if not is_de_data_entry and direction == "incoming":
            desc_upper = description.strip().upper()
            det_upper = details_paiement.strip().upper()
            if ("CR.INT.CURR.PERIOD" in desc_upper
                    or "INTEREST ON DEMAND DEPOSIT" in desc_upper
                    or "INTEREST ON DEMAND DEPOSIT" in det_upper
                    or desc_upper == "INTEREST"):
                entry["commentaires"] = "intérêts"
                is_de_data_entry = True

        # RÈGLE: Exceptions EUR (T2PI, T2RM, T2PL)
        is_exception = False
        ref_upper = ref_bancaire.upper()

        if devise == "EUR":
            if "T2PI" in ref_upper:
                entry["commentaires"] = "intérêts"
                is_exception = True
            elif direction == "incoming" and "T2RM" in ref_upper:
                entry["commentaires"] = "remboursement"
                is_exception = True
            elif direction == "outgoing" and "T2PL" in ref_upper:
                entry["commentaires"] = "placement"
                is_exception = True

        # RÈGLE: Nivellement (NIVLT dans la référence ou dans les détails SORTANTS uniquement)
        # Pour les entrants, "NIVELLEMENT" dans les détails désigne la nature des fonds reçus,
        # pas une opération de nivellement BEAC — ces transactions restent dans la feuille principale.
        is_nivellement = False
        if "NIVLT" in ref_upper:
            entry["commentaires"] = "nivellement"
            is_nivellement = True
        # Vérifier dans les détails uniquement pour les sortants
        if not is_nivellement and direction == "outgoing" and "NIVELLEMENT" in details_paiement.upper():
            entry["commentaires"] = "nivellement"
            is_nivellement = True

        # RÈGLE: Forex (entrants seulement)
        # - Via code BIC dans la table forex (EUR/USD générique)
        # - Via nom de bénéficiaire pour les entrants USD CITI (banques étrangères non-CEMAC)
        _CITI_USD_FOREX_NAMES = {
            "CITIBANK LONDON", "LANDESBANK", "STANDARD CHARTERED",
            "CREDIT AGRICOLE CORPORATE", "CRÉDIT AGRICOLE CORPORATE",
            "NATIXIS", "NATXFRPPMAR",
        }
        is_forex = False
        if direction == "incoming":
            if forex_codes and bic_code:
                bic_8 = bic_code[:8].upper()
                if bic_8 in forex_codes:
                    entry["commentaires"] = "forex"
                    is_forex = True
            if not is_forex and devise == "USD":
                benef_upper_chk = benef_remettant.upper()
                for _kw in _CITI_USD_FOREX_NAMES:
                    if _kw.upper() in benef_upper_chk:
                        entry["commentaires"] = "forex"
                        is_forex = True
                        break

        # Routage (priorité : BEACCMCX091 > BDF > DE-Data Entry > exceptions > nivellement > forex > normal)
        if is_beaccmcx091:
            beaccmcx091_rows.append(entry)
        elif is_bdf:
            banque_de_france_rows.append(entry)
        elif is_de_data_entry:
            other_exceptions_rows.append(entry)
        elif is_exception or is_nivellement:
            other_exceptions_rows.append(entry)
        elif is_forex:
            forex_rows.append(entry)
        else:
            rows.append(entry)

    wb.close()
    return rows, beaccmcx091_rows, other_exceptions_rows, forex_rows, banque_de_france_rows


# ========================================================================================
# EXTRACTEUR STANDARD (SCBL - Standard Chartered Bank)
# ========================================================================================

def _load_known_codes(xlsx_bic_path: Optional[str] = None) -> Tuple[dict, dict]:
    """
    Charger les codes BIC (8 chars) et codes règlement (4 chiffres) depuis bic_codes.xlsx.
    Retourne (bic_dict, reglement_dict) :
      - bic_dict: {code_8char_upper: {'name': ..., 'country': ..., 'bic_full': ...}}
      - reglement_dict: {code_4digit: {'name': ..., 'country': ..., 'bic': ...}}
    """
    if not HAS_BIC_UTILS:
        return {}, {}
    try:
        bic_map = load_bic_mapping(xlsx_bic_path)
    except Exception:
        return {}, {}

    # Construire le dictionnaire BIC → info complète
    bic_dict = {}
    if bic_map:
        from extractors.bic_utils import _BIC_MAP_CACHE, _BIC_COUNTRY_MAP
        for code8, name in (_BIC_MAP_CACHE or {}).items():
            country = (_BIC_COUNTRY_MAP or {}).get(code8, '')
            bic_dict[code8] = {'name': name, 'country': country, 'bic': code8}

    # Construire le dictionnaire règlement
    reglement_dict = {}
    from extractors.bic_utils import _REGLEMENT_MAP
    if _REGLEMENT_MAP:
        reglement_dict = dict(_REGLEMENT_MAP)

    return bic_dict, reglement_dict


def _search_bic_in_description(description: str, bic_dict: dict, reglement_dict: dict, is_outgoing: bool) -> Optional[Dict]:
    """
    Chercher un code BIC ou un code sous-participant dans la description.
    On itère sur les codes connus et on les cherche dans le texte (pas l'inverse).
    
    Priorité:
    1. Code BIC (8 chars, lettres majuscules) trouvé dans le texte
    2. Code sous-participant (4 chiffres) trouvé dans le texte (sortants seulement)
    3. Fallback: pattern /BNF/ suivi de 4 chiffres
    
    Returns:
        Dict avec 'bic', 'name', 'country', 'matched_code' ou None
    """
    if not description:
        return None

    desc_upper = description.upper()

    # 1. Chercher un code BIC connu dans le texte
    # On cherche les tokens de 8+ caractères alphanumériques (BIC = 8 lettres ou 11 alphanum)
    bic_tokens = re.findall(r'\b([A-Z][A-Z0-9]{7,10})\b', desc_upper)
    for token in bic_tokens:
        key8 = token[:8]
        if key8 in bic_dict and key8 not in _FALSE_BIC_WORDS:
            info = bic_dict[key8]
            return {
                'bic': info.get('bic', key8),
                'name': info.get('name', ''),
                'country': info.get('country', ''),
                'matched_code': token
            }

    # 2. Pour les sortants, chercher les codes sous-participant (4 chiffres)
    if is_outgoing and reglement_dict:
        # Extraire tous les tokens de 4 chiffres avec limites
        digit_tokens = re.findall(r'\b(\d{4})\b', description)
        for token in digit_tokens:
            if token in reglement_dict:
                info = reglement_dict[token]
                return {
                    'bic': info.get('bic', ''),
                    'name': info.get('name', ''),
                    'country': info.get('country', ''),
                    'matched_code': token
                }

    # 3. Fallback: pattern /BNF/ suivi de 4 chiffres
    bnf_match = _BNF_CODE_RE.search(description)
    if bnf_match:
        code_4 = bnf_match.group(1)
        if reglement_dict and code_4 in reglement_dict:
            info = reglement_dict[code_4]
            return {
                'bic': info.get('bic', ''),
                'name': info.get('name', ''),
                'country': info.get('country', ''),
                'matched_code': code_4
            }

    return None


def extract_standard(file_path: str, xlsx_bic_path: Optional[str] = None) -> Tuple[List[Dict], List[Dict], List[Dict], List[Dict]]:
    """
    Extraire les données d'un fichier Standard Chartered (SCBL).

    Structure Standard (headers en ligne 5) :
    Col B(1): Account Number, Col C(2): Account Name,
    Col H(7): Currency, Col I(8): Date,
    Col J(9): Description, Col L(11): Withdrawal (sortant),
    Col M(12): Deposit (entrant), Col O(14): Balance

    Returns:
        (rows, beaccmcx091_rows, other_exceptions_rows, forex_rows)
    """
    fp = Path(file_path)
    if not fp.exists():
        raise FileNotFoundError(f"Fichier Standard introuvable : {file_path}")

    # Charger BIC mapping et codes connus
    bic_dict, reglement_dict = _load_known_codes(xlsx_bic_path)

    # Charger les codes forex
    forex_codes = set()
    if HAS_BIC_UTILS and load_forex_codes:
        try:
            forex_codes = load_forex_codes(xlsx_bic_path)
        except Exception:
            pass

    # NOTE: read_only=True ne fonctionne pas avec certains fichiers Standard (cellules vides)
    wb = openpyxl.load_workbook(str(fp), read_only=False, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = []
    beaccmcx091_rows = []
    other_exceptions_rows = []
    forex_rows = []

    file_name = fp.name

    # Trouver la ligne d'en-tête (cherche "Description" dans la ligne)
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
        if row:
            vals = [str(v or '').strip().lower() for v in row]
            if 'description' in vals:
                header_row = row_idx
                break
    if not header_row:
        # Fallback : ligne 5 par défaut
        header_row = 5

    data_start = header_row + 1

    for row_idx, row in enumerate(ws.iter_rows(min_row=data_start, values_only=True), start=data_start):
        if not row:
            continue

        # Colonnes Standard :
        # B(1)=Account Number, C(2)=Account Name, F(5)=Address,
        # H(7)=Currency, I(8)=Date, J(9)=Description,
        # L(11)=Withdrawal, M(12)=Deposit, O(14)=Balance
        devise = str(row[7] or '').strip().upper() if len(row) > 7 else ''
        date_raw = row[8] if len(row) > 8 else None
        description = str(row[9] or '').strip() if len(row) > 9 else ''
        withdrawal_raw = row[11] if len(row) > 11 else None
        deposit_raw = row[12] if len(row) > 12 else None

        if not description or description.lower() == 'balance brought forward':
            continue

        # Parser le montant et la direction
        withdrawal = _parse_amount_string(withdrawal_raw)
        deposit = _parse_amount_string(deposit_raw)

        if withdrawal and withdrawal > 0:
            montant = withdrawal
            direction = "outgoing"
        elif deposit and deposit > 0:
            montant = deposit
            direction = "incoming"
        else:
            continue

        # Parser la date
        date_val = _parse_date_standard(date_raw)

        # Chercher un code BIC ou sous-participant dans la description
        match_info = _search_bic_in_description(description, bic_dict, reglement_dict, direction == "outgoing")

        bic_code = None
        banque_name = None
        pays_iso3 = None
        commentaire = description  # Tout le reste va en commentaires

        if match_info:
            bic_code = match_info.get('bic')
            banque_name = match_info.get('name')
            pays_iso3 = match_info.get('country')

        # Source tracking
        source = f"ligne N°{row_idx} du fichier {file_name}"

        entry = {
            "correspondant": "SCBLGB2LXXX",
            "date_reference": date_val,
            "reference": None,
            "reference_origine": None,
            "type_MT": f"Excel Standard {devise} {'entrant' if direction == 'incoming' else 'sortant'}",
            "pays_iso3": pays_iso3,
            "code_donneur_dordre": bic_code,
            "donneur_dordre": banque_name,
            "institution_name": banque_name,
            "beneficiaire": None,
            "montant": montant,
            "devise": devise,
            "commentaires": commentaire,
            "source_pdf": source,
            "direction": direction,
        }

        # ===== Application des règles de routage =====

        # RÈGLE: BEACCMCX091
        is_beaccmcx091 = False
        if bic_code and 'BEACCMCX091' in bic_code.upper():
            is_beaccmcx091 = True
        if not is_beaccmcx091 and 'BEACCMCX091' in description.upper():
            is_beaccmcx091 = True

        # RÈGLE: Nivellement
        is_nivellement = False
        if "NIVELLEMENT" in description.upper() or "NIVLT" in description.upper():
            entry["commentaires"] = "nivellement"
            is_nivellement = True

        # RÈGLE: Forex (entrants seulement)
        is_forex = False
        if direction == "incoming" and forex_codes and bic_code:
            bic_8 = bic_code[:8].upper()
            if bic_8 in forex_codes:
                entry["commentaires"] = (entry.get("commentaires") or "") + " forex" if entry.get("commentaires") else "forex"
                is_forex = True

        # Routage
        if is_beaccmcx091:
            beaccmcx091_rows.append(entry)
        elif is_nivellement:
            other_exceptions_rows.append(entry)
        elif is_forex:
            forex_rows.append(entry)
        else:
            rows.append(entry)

    wb.close()
    return rows, beaccmcx091_rows, other_exceptions_rows, forex_rows


def _parse_date_standard(date_val) -> Optional[str]:
    """Parser une date au format Standard '02 Feb 2026' ou datetime."""
    if not date_val:
        return None
    if isinstance(date_val, datetime):
        return date_val.strftime('%Y-%m-%d')
    s = str(date_val).strip()
    # Format "DD Mon YYYY" (ex: "02 Feb 2026")
    try:
        from datetime import datetime as dt
        parsed = dt.strptime(s, '%d %b %Y')
        return parsed.strftime('%Y-%m-%d')
    except ValueError:
        pass
    # Format DD/MM/YYYY
    m = re.match(r'^(\d{2})/(\d{2})/(\d{4})$', s)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    # Format YYYY-MM-DD
    if re.match(r'^\d{4}-\d{2}-\d{2}$', s):
        return s
    return s


# ========================================================================================
# EXTRACTEUR CITI USD RELEVÉ DE COMPTE (format .xls brut)
# ========================================================================================

# Descriptions de transaction valides dans le relevé CITI USD brut
_CITI_USD_RELEVE_VALID_DESCRIPTIONS = frozenset([
    'SAME DAY CR TRANSFER',
    'SAME DAY DR TRANSFER',
    'INTEREST',
    'BILLING DIRECT DEBIT',
    'ACH CREDIT',
    'ACH DEBIT',
])

# Forex pattern dans Customer Reference : 8 chiffres + point + 6 chiffres (YYMMDD)
_FOREX_CUST_REF_RE = re.compile(r'^\d{8}\.\d{6}$')

# Pattern exception BC : chiffres suivis de "BC" en fin de chaîne
_BC_SUFFIX_RE = re.compile(r'\d+BC$')


def extract_citi_usd_releve(file_path: str, xlsx_bic_path: Optional[str] = None) -> Tuple[List[Dict], List[Dict], List[Dict], List[Dict]]:
    """
    Extraire les données d'un relevé de compte brut CITI USD (format .xls).

    Structure : fichier .xls avec sections journalières répétées.
    Chaque section a un en-tête :
      Entry Date | Value Date | Customer Reference | Bank Reference |
      Transaction Description | By Order Of / Beneficiary | Transaction Amount

    Direction déterminée par le signe du montant : négatif = sortant, positif = entrant.

    Règles d'exception (par ordre de priorité) :
      1. BEACCMCX091 dans By Order Of → beaccmcx091_rows
      2. INTEREST ou BILLING DIRECT DEBIT → other_exceptions (commentaire "intérêts")
      3. Customer Reference finissant par "BC" → other_exceptions (commentaire "exception BC")
      4. Forex : pattern XXXXXXXX.YYMMDD dans Customer Reference → forex_rows (entrants ET sortants)
      5. Tout le reste → rows (opérations normales)

    Identification de la contrepartie :
      - Entrants : BIC ou nom de banque depuis "By Order Of" (matching bic_codes.xlsx)
      - Sortants : code de règlement (4 premiers chiffres) depuis Customer Reference (priorité),
                   puis "By Order Of" en complément

    Returns:
        (rows, beaccmcx091_rows, other_exceptions_rows, forex_rows)
    """
    import pandas as pd

    fp = Path(file_path)
    if not fp.exists():
        raise FileNotFoundError(f"Fichier CITI USD Relevé introuvable : {file_path}")

    # Charger BIC mapping
    if HAS_BIC_UTILS:
        try:
            load_bic_mapping(xlsx_bic_path)
        except Exception as e:
            logger.debug("extract_citi_usd_releve: BIC mapping load failed: %s", e)

    # Lire le fichier (pandas détecte automatiquement .xls vs .xlsx)
    df = pd.read_excel(str(fp), header=None)

    rows = []
    beaccmcx091_rows = []
    other_exceptions_rows = []
    forex_rows = []
    file_name = fp.name

    for idx, raw_row in df.iterrows():
        # Identifier les lignes de transaction via la description (colonne 4)
        desc_raw = raw_row.iloc[4] if len(raw_row) > 4 else None
        if pd.isna(desc_raw):
            continue
        desc = str(desc_raw).strip()
        if desc not in _CITI_USD_RELEVE_VALID_DESCRIPTIONS:
            continue

        # Parser le montant (colonne 6)
        amt_raw = raw_row.iloc[6] if len(raw_row) > 6 else None
        if pd.isna(amt_raw):
            continue
        try:
            montant_raw = float(amt_raw)
        except (ValueError, TypeError):
            continue
        if montant_raw == 0:
            continue

        # Direction depuis le signe du montant
        if montant_raw < 0:
            direction = "outgoing"
            montant = abs(montant_raw)
        else:
            direction = "incoming"
            montant = montant_raw

        # Extraire les champs
        value_date_raw = raw_row.iloc[1] if len(raw_row) > 1 and pd.notna(raw_row.iloc[1]) else None
        customer_ref = str(raw_row.iloc[2]).strip() if len(raw_row) > 2 and pd.notna(raw_row.iloc[2]) else ''
        bank_ref = str(raw_row.iloc[3]).strip() if len(raw_row) > 3 and pd.notna(raw_row.iloc[3]) else ''
        by_order_of = str(raw_row.iloc[5]).strip() if len(raw_row) > 5 and pd.notna(raw_row.iloc[5]) else ''

        # Parser la date de valeur (format MM/DD/YYYY pour CITI US)
        date_val = _parse_date(value_date_raw, format_hint='mdy')

        # ===== Résolution de la contrepartie =====
        bic_code = None
        banque_name = None
        pays_iso3 = None
        beneficiary = None

        if direction == "incoming":
            # ENTRANTS : identification via By Order Of
            if by_order_of:
                bic_code = _extract_bic_from_text(by_order_of)
                if bic_code and HAS_BIC_UTILS:
                    banque_name = map_code_to_name(bic_code, xlsx_path=xlsx_bic_path)
                    pays_iso3 = map_code_to_country(bic_code, xlsx_path=xlsx_bic_path)

                # Essayer d'autres candidats BIC si le pays n'est pas résolu
                if not pays_iso3 and HAS_BIC_UTILS:
                    for candidate in _extract_all_bic_candidates(by_order_of):
                        if candidate == bic_code:
                            continue
                        c_pays = map_code_to_country(candidate, xlsx_path=xlsx_bic_path)
                        if c_pays:
                            if not bic_code:
                                bic_code = candidate
                            if not banque_name:
                                banque_name = map_code_to_name(candidate, xlsx_path=xlsx_bic_path)
                            pays_iso3 = c_pays
                            break

                # Si pas de nom résolu, garder le texte brut
                if not banque_name:
                    banque_name = by_order_of if by_order_of else None

                # Détection fuzzy du pays CEMAC
                if not pays_iso3 and by_order_of:
                    pays_iso3 = _detect_cemac_country_fuzzy(by_order_of)

        else:
            # SORTANTS : PRIORITÉ au code de règlement dans Customer Reference
            if customer_ref and HAS_BIC_UTILS and map_reglement_code:
                # Ne pas essayer le code règlement si c'est un pattern forex
                if not _FOREX_CUST_REF_RE.match(customer_ref):
                    parts = customer_ref.split('.')
                    if parts and len(parts[0]) == 4 and parts[0].isdigit():
                        reg_code = parts[0]
                        reg_info = map_reglement_code(reg_code, xlsx_path=xlsx_bic_path)
                        if reg_info:
                            bic_code = reg_info.get("bic") or bic_code
                            banque_name = reg_info.get("name") or banque_name
                            pays_iso3 = reg_info.get("country") or pays_iso3

                    # Détecter le pays ISO2 dans la Customer Reference
                    if not pays_iso3:
                        iso2_match = _ISO2_PATTERN.search(customer_ref.upper())
                        if iso2_match:
                            pays_iso3 = CEMAC_ISO2.get(iso2_match.group(1))

            # SECONDAIRE : By Order Of pour le bénéficiaire
            if by_order_of:
                if by_order_of.startswith("1/"):
                    beneficiary = by_order_of[2:].strip()
                else:
                    # Essayer BIC matching dans By Order Of (secondaire)
                    if not bic_code:
                        bic_from_order = _extract_bic_from_text(by_order_of)
                        if bic_from_order and HAS_BIC_UTILS:
                            bic_code = bic_from_order
                            banque_name = map_code_to_name(bic_from_order, xlsx_path=xlsx_bic_path) or banque_name
                            pays_iso3 = map_code_to_country(bic_from_order, xlsx_path=xlsx_bic_path) or pays_iso3

                    if not banque_name:
                        banque_name = by_order_of

                # Détection fuzzy du pays CEMAC
                if not pays_iso3:
                    pays_iso3 = _detect_cemac_country_fuzzy(by_order_of)

        # Source tracking
        source = f"ligne N°{idx + 1} du fichier {file_name}"

        entry = {
            "correspondant": "CITIUS33XXX",
            "date_reference": date_val,
            "reference": bank_ref,
            "reference_origine": customer_ref,
            "type_MT": f"Relevé CITI USD {'entrant' if direction == 'incoming' else 'sortant'}",
            "pays_iso3": pays_iso3,
            "code_donneur_dordre": bic_code,
            "donneur_dordre": banque_name,
            "institution_name": banque_name,
            "beneficiaire": beneficiary,
            "montant": montant,
            "devise": "USD",
            "commentaires": None,
            "source_pdf": source,
            "direction": direction,
        }

        # ===== Règles d'exception (par ordre de priorité) =====

        # Règle 1 : BEACCMCX091 dans By Order Of
        is_beaccmcx091 = 'BEACCMCX091' in by_order_of.upper()

        # Règle 2 : INTEREST ou BILLING DIRECT DEBIT → exception "intérêts"
        is_interest = desc in ('INTEREST', 'BILLING DIRECT DEBIT')

        # Règle 3 : Customer Reference finissant par "BC" → exception BC
        is_bc_exception = bool(_BC_SUFFIX_RE.search(customer_ref))

        # Règle 4 : Forex pattern dans Customer Reference (entrants ET sortants)
        is_forex = bool(_FOREX_CUST_REF_RE.match(customer_ref))

        # Routage (BEACCMCX091 > intérêts > BC > forex > normal)
        if is_beaccmcx091:
            beaccmcx091_rows.append(entry)
        elif is_interest:
            entry["commentaires"] = "intérêts"
            other_exceptions_rows.append(entry)
        elif is_bc_exception:
            entry["commentaires"] = "exception BC"
            other_exceptions_rows.append(entry)
        elif is_forex:
            entry["commentaires"] = "forex"
            forex_rows.append(entry)
        else:
            rows.append(entry)

    return rows, beaccmcx091_rows, other_exceptions_rows, forex_rows


# ========================================================================================
# DISPATCHER PRINCIPAL
# ========================================================================================

CORRESPONDANT_BDF = "BDF"
CORRESPONDANT_CITI = "CITI"
CORRESPONDANT_STANDARD = "Standard"
CORRESPONDANT_CITI_USD_RELEVE = "CITI_USD_Releve"

def detect_correspondant_from_file(file_path: str) -> Optional[str]:
    """
    Détecter le correspondant à partir du nom du fichier ou de son contenu.
    """
    name = Path(file_path).name.upper()
    if 'BDF' in name or 'BANQUE DE FRANCE' in name or 'RELEVÉ BDF' in name.replace('É', 'E') or 'RELEVE BDF' in name:
        return CORRESPONDANT_BDF
    if 'CITI' in name:
        return CORRESPONDANT_CITI
    if 'SCBL' in name or 'STANDARD' in name:
        return CORRESPONDANT_STANDARD

    # Tenter de détecter depuis le contenu
    try:
        wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        # Vérifier les premières lignes
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            if row and row[0]:
                val = str(row[0]).upper()
                if "CONSULTATION" in val and "POSITION" in val:
                    wb.close()
                    return CORRESPONDANT_BDF
        # Vérifier les en-têtes CITI
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if first_row and first_row[0]:
            if 'DATE DE VALEUR' in str(first_row[0]).upper():
                wb.close()
                return CORRESPONDANT_CITI
        wb.close()
    except Exception as e:
        logger.debug("detect_correspondant_from_file: error reading %s: %s", file_path, e)

    return None


def extract_excel_files(file_paths: List[str], correspondant: str, xlsx_bic_path: Optional[str] = None) -> Tuple[List[Dict], List[Dict], List[Dict], List[Dict], List[Dict], Dict[str, set]]:
    """
    Extraire les données de plusieurs fichiers Excel du même correspondant.

    Args:
        file_paths: Liste des chemins vers les fichiers Excel
        correspondant: "BDF", "CITI", "Standard" ou "CITI_USD_Releve"
        xlsx_bic_path: Chemin vers le fichier bic_codes.xlsx

    Returns:
        (rows, beaccmcx091_rows, other_exceptions_rows, forex_rows, banque_de_france_rows, missing_codes)
    """
    all_rows = []
    all_beaccmcx091 = []
    all_other_exceptions = []
    all_forex = []
    all_bdf = []
    missing_codes = {"unmapped": set(), "empty": set()}

    for fp in file_paths:
        try:
            if correspondant == CORRESPONDANT_BDF:
                r, b, o, f = extract_bdf(fp, xlsx_bic_path=xlsx_bic_path)
                bdf = []
            elif correspondant == CORRESPONDANT_CITI:
                r, b, o, f, bdf = extract_citi(fp, xlsx_bic_path=xlsx_bic_path)
            elif correspondant == CORRESPONDANT_STANDARD:
                r, b, o, f = extract_standard(fp, xlsx_bic_path=xlsx_bic_path)
                bdf = []
            elif correspondant == CORRESPONDANT_CITI_USD_RELEVE:
                r, b, o, f = extract_citi_usd_releve(fp, xlsx_bic_path=xlsx_bic_path)
                bdf = []
            else:
                logger.warning("excel_extractor: correspondant inconnu '%s', skip %s", correspondant, fp)
                continue

            all_rows.extend(r)
            all_beaccmcx091.extend(b)
            all_other_exceptions.extend(o)
            all_forex.extend(f)
            all_bdf.extend(bdf)

            # Tracker les codes manquants
            for entry in r + b + o + f + bdf:
                code = entry.get("code_donneur_dordre")
                name = entry.get("donneur_dordre")
                if not code:
                    missing_codes["empty"].add("(vide)")
                elif name == code:
                    missing_codes["unmapped"].add(code)

        except Exception as e:
            logger.exception("excel_extractor: erreur lors de l'extraction de %s: %s", fp, e)

    return all_rows, all_beaccmcx091, all_other_exceptions, all_forex, all_bdf, missing_codes
